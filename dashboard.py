"""
运营数据分析面板（Claude 交接版）
- 保留原始中文字段命名
- BigQuery 联动
- 会员价值页补齐筛选与默认排除提示
- 实时波动页修复「时间=2026-01-31 23~24」解析
- 若会员表未补 _snapshot_month / _snapshot_date，则留存明确标记为暂未启用
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from html import escape
from typing import Dict, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from google.cloud import bigquery

PROJECT_ID = "dashboard-492601"
DATASET_ID = "ops_data"
BQ_PREFIX = f"{PROJECT_ID}.{DATASET_ID}"

# ── 版本号 ────────────────────────────────────────────────
APP_VERSION = "v2.2"          # 数据健康页 + 上传校验/内容识别/月份选 + 三份代理数据全自助(市代月度独立页+对比) + 平哥结算月报自助
APP_VERSION_DATE = "2026-06-17"

# ── 页面配置 ──────────────────────────────────────────────
st.set_page_config(page_title="运营数据分析", layout="wide")

# ── 样式 ──────────────────────────────────────────────────
st.markdown(
    """
<style>
:root {
  --bg-1: #040a16;
  --bg-2: #081226;
  --card: rgba(9, 22, 42, 0.85);
  --card-2: rgba(12, 28, 52, 0.90);
  --line: rgba(34, 211, 238, 0.18);
  --line-strong: rgba(34, 211, 238, 0.45);
  --glow: rgba(34, 211, 238, 0.30);
  --text-1: #e8f7ff;
  --text-2: #9fc1d9;
  --text-3: #6d8aa6;
  --accent: #22d3ee;
  --accent-page: #22d3ee;
  --good: #2dd4a7;
  --good-soft: rgba(45, 212, 167, 0.12);
  --good-text: #7df0c8;
  --bad: #fb7185;
  --bad-soft: rgba(251, 113, 133, 0.12);
  --bad-text: #fda4af;
  --warn: #fbbf24;
  --warn-soft: rgba(251, 191, 36, 0.12);
  --warn-text: #fde08a;
  --green: #2dd4a7;
  --red: #fb7185;
  --amber: #fbbf24;
}
html, body, [data-testid="stAppViewContainer"] {
  background:
    radial-gradient(900px 420px at 12% -8%, rgba(14, 90, 130, 0.22) 0%, rgba(14, 90, 130, 0) 60%),
    radial-gradient(1100px 500px at 88% -10%, rgba(20, 60, 130, 0.28) 0%, rgba(20, 60, 130, 0) 60%),
    linear-gradient(180deg, #030810 0%, #050d1c 100%);
}
[data-testid="stHeader"] { background: rgba(0,0,0,0); }
[data-testid="stSidebar"] { display: none; }
.block-container { padding-top: 0.9rem; padding-bottom: 2rem; max-width: 1480px; }

/* ── 导航：科技感分段控件 ── */
div[role="radiogroup"] { gap: 0.35rem 0.4rem; }
div[role="radiogroup"] label {
  background: rgba(7, 18, 36, 0.85);
  border: 1px solid var(--line);
  border-radius: 4px;
  padding: 0.34rem 0.95rem;
  transition: border-color 0.15s, background 0.15s, box-shadow 0.15s;
  clip-path: polygon(8px 0, 100% 0, 100% calc(100% - 8px), calc(100% - 8px) 100%, 0 100%, 0 8px);
}
div[role="radiogroup"] label:hover { border-color: var(--line-strong); }
div[role="radiogroup"] label:has(input:checked) {
  background: linear-gradient(135deg, rgba(34,211,238,0.22) 0%, rgba(34,211,238,0.06) 100%);
  border-color: var(--accent-page);
  box-shadow: 0 0 14px var(--glow), inset 0 0 10px rgba(34,211,238,0.08);
}
div[role="radiogroup"] label p { color: #c9e6f5; font-weight: 600; font-size: 0.88rem; }
div[role="radiogroup"] label:has(input:checked) p { color: #ffffff; text-shadow: 0 0 8px var(--glow); }

/* ── 页头：大屏式标题横幅 ── */
.hero-card {
  position: relative;
  overflow: hidden;
  background: linear-gradient(180deg, rgba(10,26,48,0.92) 0%, rgba(6,15,30,0.88) 100%);
  border: 1px solid var(--line);
  border-radius: 6px;
  padding: 0.95rem 1.3rem;
  margin-bottom: 0.9rem;
  box-shadow: inset 0 0 30px rgba(34,211,238,0.04);
}
.hero-card::before {
  content: '';
  position: absolute;
  top: 0; left: 0; right: 0;
  height: 2px;
  background: linear-gradient(90deg, rgba(0,0,0,0) 0%, var(--accent-page) 18%, var(--accent-page) 82%, rgba(0,0,0,0) 100%);
  filter: drop-shadow(0 0 6px var(--accent-page));
}
.hero-card::after {
  content: '';
  position: absolute;
  bottom: 0; left: 8%; right: 8%;
  height: 1px;
  background: linear-gradient(90deg, rgba(0,0,0,0) 0%, var(--line-strong) 50%, rgba(0,0,0,0) 100%);
}
.hero-title {
  color: var(--text-1);
  font-size: 1.22rem;
  font-weight: 800;
  letter-spacing: 0.12em;
  margin-bottom: 0.25rem;
  text-shadow: 0 0 16px var(--glow);
}
.hero-subtitle {
  color: var(--text-2);
  font-size: 0.9rem;
  line-height: 1.45;
}
.badge-row { display: flex; flex-wrap: wrap; gap: 0.5rem; margin-bottom: 0.7rem; }
.badge {
  display: inline-flex;
  align-items: center;
  gap: 0.4rem;
  background: rgba(11, 32, 58, 0.65);
  border: 1px solid var(--line);
  color: #cfeefb;
  border-radius: 3px;
  padding: 0.22rem 0.62rem;
  font-size: 0.78rem;
  font-weight: 600;
}
.badge::before { content: ''; width: 6px; height: 6px; border-radius: 50%; background: var(--accent-page); }
.badge-good { background: var(--good-soft); border-color: rgba(45,212,167,0.40); color: var(--good-text); }
.badge-good::before { background: var(--good); }
.badge-bad { background: var(--bad-soft); border-color: rgba(251,113,133,0.40); color: var(--bad-text); }
.badge-bad::before { background: var(--bad); }
.badge-warn { background: var(--warn-soft); border-color: rgba(251,191,36,0.40); color: var(--warn-text); }
.badge-warn::before { background: var(--warn); }

/* ── 顶部核心结论横幅 ── */
.kpi-banner {
  display: flex;
  flex-wrap: wrap;
  align-items: baseline;
  gap: 0.4rem 1.5rem;
  background: linear-gradient(90deg, rgba(34,211,238,0.10) 0%, rgba(6,15,30,0.85) 60%);
  border: 1px solid var(--line-strong);
  border-radius: 5px;
  padding: 0.7rem 1.2rem;
  margin-bottom: 0.9rem;
  box-shadow: 0 0 18px rgba(34,211,238,0.08), inset 0 0 24px rgba(34,211,238,0.05);
}
.kpi-banner .kb-main { font-size: 1.25rem; font-weight: 800; color: #eafcff; text-shadow: 0 0 14px var(--glow); }
.kpi-banner .kb-item { color: var(--text-2); font-size: 0.95rem; font-weight: 600; }
.kpi-banner .kb-up { color: var(--good-text); font-weight: 700; }
.kpi-banner .kb-down { color: var(--bad-text); font-weight: 700; }

/* ── 指标卡：大屏式发光数字 + 角标，tone-* 上状态色 ── */
.metric-card {
  position: relative;
  background: linear-gradient(180deg, rgba(11,28,52,0.85) 0%, rgba(6,15,30,0.92) 100%);
  border: 1px solid var(--line);
  border-radius: 5px;
  padding: 0.8rem 0.95rem 0.8rem 1.05rem;
  min-height: 118px;
  box-shadow: inset 0 0 22px rgba(34,211,238,0.04);
}
.metric-card::before {
  content: '';
  position: absolute;
  left: 0; top: 14%; bottom: 14%;
  width: 3px;
  border-radius: 0 3px 3px 0;
  background: transparent;
}
.metric-card::after {
  content: '';
  position: absolute;
  top: -1px; right: -1px;
  width: 12px; height: 12px;
  border-top: 2px solid var(--line-strong);
  border-right: 2px solid var(--line-strong);
}
.metric-card.tone-good { border-color: rgba(45,212,167,0.40); box-shadow: inset 0 0 22px rgba(45,212,167,0.06); }
.metric-card.tone-good::before { background: var(--good); box-shadow: 0 0 8px var(--good); }
.metric-card.tone-good::after { border-color: rgba(45,212,167,0.65); }
.metric-card.tone-good .metric-value { color: var(--good-text); text-shadow: 0 0 14px rgba(45,212,167,0.45); }
.metric-card.tone-bad { border-color: rgba(251,113,133,0.40); box-shadow: inset 0 0 22px rgba(251,113,133,0.06); }
.metric-card.tone-bad::before { background: var(--bad); box-shadow: 0 0 8px var(--bad); }
.metric-card.tone-bad::after { border-color: rgba(251,113,133,0.65); }
.metric-card.tone-bad .metric-value { color: var(--bad-text); text-shadow: 0 0 14px rgba(251,113,133,0.45); }
.metric-card.tone-warn { border-color: rgba(251,191,36,0.40); box-shadow: inset 0 0 22px rgba(251,191,36,0.05); }
.metric-card.tone-warn::before { background: var(--warn); box-shadow: 0 0 8px var(--warn); }
.metric-card.tone-warn::after { border-color: rgba(251,191,36,0.65); }
.metric-card.tone-warn .metric-value { color: var(--warn-text); text-shadow: 0 0 14px rgba(251,191,36,0.40); }
.metric-card.tone-accent { border-color: var(--line-strong); }
.metric-card.tone-accent::before { background: var(--accent-page); box-shadow: 0 0 8px var(--accent-page); }
.metric-label { color: #8fb6cf; font-size: 0.84rem; margin-bottom: 0.3rem; letter-spacing: 0.03em; }
.metric-value {
  color: #eafcff;
  font-size: 1.78rem;
  font-weight: 800;
  line-height: 1.12;
  font-variant-numeric: tabular-nums;
  text-shadow: 0 0 14px var(--glow);
}
.metric-delta {
  display: inline-block;
  margin-top: 0.5rem;
  padding: 0.18rem 0.55rem;
  border-radius: 999px;
  font-size: 0.8rem;
  font-weight: 700;
  color: #d8e6ff;
  background: rgba(255,255,255,0.07);
  border: 1px solid rgba(255,255,255,0.10);
}
.metric-delta.d-up { color: var(--good-text); background: var(--good-soft); border-color: rgba(45,212,167,0.35); }
.metric-delta.d-down { color: var(--bad-text); background: var(--bad-soft); border-color: rgba(251,113,133,0.35); }
.metric-help {
  color: var(--text-3);
  font-size: 0.78rem;
  line-height: 1.35;
  margin-top: 0.5rem;
  white-space: normal;
  word-break: break-word;
}

/* ── 区块标题：大屏式面板头（左标记 + 右延伸线）── */
.section-title {
  display: flex;
  align-items: center;
  gap: 0.55rem;
  color: var(--text-1);
  font-size: 1.05rem;
  font-weight: 700;
  letter-spacing: 0.05em;
  margin-top: 0.5rem;
  margin-bottom: 0.25rem;
}
.section-title::before {
  content: '';
  width: 4px;
  height: 1.0em;
  background: var(--accent-page);
  box-shadow: 0 0 8px var(--accent-page);
  flex: 0 0 auto;
}
.section-title::after {
  content: '';
  flex: 1 1 auto;
  height: 1px;
  margin-left: 0.6rem;
  background: linear-gradient(90deg, var(--line-strong) 0%, rgba(0,0,0,0) 85%);
}
.section-subtitle {
  color: var(--text-3);
  font-size: 0.84rem;
  margin-bottom: 0.55rem;
}
.info-chip {
  background: rgba(42, 55, 20, 0.65);
  border: 1px solid rgba(220, 214, 82, 0.24);
  color: #f3edbb;
  border-radius: 12px;
  padding: 0.8rem 0.95rem;
  margin: 0.45rem 0 0.9rem 0;
}
.filter-note {
  background: rgba(22, 33, 58, 0.72);
  border: 1px dashed rgba(115, 156, 255, 0.35);
  color: #d8e6ff;
  border-radius: 12px;
  padding: 0.7rem 0.85rem;
  margin: 0.35rem 0 0.8rem 0;
}
.tooltip-note {
  color: #f4e6a5;
  font-weight: 700;
}
div[data-testid="stTabs"] button {
  border-radius: 4px !important;
  border: 1px solid var(--line) !important;
  background: rgba(7, 18, 36, 0.75) !important;
  color: #c9e6f5 !important;
  padding: 0.4rem 1rem !important;
  margin-right: 0.4rem !important;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
  border-color: var(--accent-page) !important;
  box-shadow: 0 0 12px var(--glow);
}
button[kind="secondaryFormSubmit"] { border-radius: 6px !important; }
[data-testid="stDataFrame"] {
  border: 1px solid var(--line);
  border-radius: 5px;
  overflow: hidden;
}

/* ── 面板卡：st.container(border=True) 重绘成大屏 panel（角标 + 内光）── */
[data-testid="stVerticalBlockBorderWrapper"] {
  position: relative;
  background: linear-gradient(180deg, rgba(9,23,44,0.60) 0%, rgba(5,13,26,0.72) 100%);
  border: 1px solid var(--line) !important;
  border-radius: 5px !important;
  padding: 0.85rem 0.95rem 0.5rem !important;
  box-shadow: inset 0 0 28px rgba(34,211,238,0.04);
}
[data-testid="stVerticalBlockBorderWrapper"]::before {
  content: '';
  position: absolute;
  top: -1px; left: -1px;
  width: 14px; height: 14px;
  border-top: 2px solid var(--line-strong);
  border-left: 2px solid var(--line-strong);
  pointer-events: none;
}
[data-testid="stVerticalBlockBorderWrapper"]::after {
  content: '';
  position: absolute;
  bottom: -1px; right: -1px;
  width: 14px; height: 14px;
  border-bottom: 2px solid var(--line-strong);
  border-right: 2px solid var(--line-strong);
  pointer-events: none;
}
[data-testid="stVerticalBlockBorderWrapper"] [data-testid="stVerticalBlockBorderWrapper"] {
  background: rgba(6, 14, 28, 0.45);
}
[data-testid="stVerticalBlockBorderWrapper"] .metric-card {
  background: rgba(6, 14, 28, 0.55);
}

/* ── 输入控件统一暗色科技风 ── */
[data-testid="stNumberInput"] input,
[data-testid="stDateInput"] input,
[data-testid="stTextInput"] input {
  background: rgba(7, 18, 36, 0.9) !important;
  border-radius: 4px !important;
  color: var(--text-1) !important;
}
[data-baseweb="select"] > div {
  background: rgba(7, 18, 36, 0.9) !important;
  border-radius: 4px !important;
}
[data-testid="stNumberInput"] button { background: rgba(12, 28, 52, 0.85) !important; }
[data-testid="stFileUploader"] section {
  background: rgba(7, 18, 36, 0.6) !important;
  border: 1px dashed var(--line-strong) !important;
  border-radius: 5px !important;
}
[data-testid="stWidgetLabel"] p { color: var(--text-2) !important; font-size: 0.84rem !important; }
[data-testid="stCaptionContainer"] { color: var(--text-3) !important; }
hr { border-color: var(--line) !important; margin: 0.8rem 0 !important; }
</style>
""",
    unsafe_allow_html=True,
)

BLUE = "#5B8FF9"
CYAN = "#22D3EE"
GREEN = "#2DD4A7"
RED = "#FB7185"
PURPLE = "#A78BFA"
AMBER = "#FBBF24"

# ── 自定义图表主题（全站统一）：透明底融入面板卡、细网格、统一字体与悬浮样式 ──
import plotly.io as pio

_ops_template = go.layout.Template(pio.templates["plotly_dark"])
_ops_template.layout.update(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(
        family='-apple-system, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif',
        size=12,
        color="#9fc1d9",
    ),
    title=dict(font=dict(color="#e8f7ff", size=14)),
    margin=dict(l=14, r=14, t=36, b=14),
    hoverlabel=dict(
        bgcolor="#0a1d36",
        bordercolor="rgba(34,211,238,0.45)",
        font=dict(color="#e8f7ff", size=12),
    ),
    xaxis=dict(
        gridcolor="rgba(56,189,248,0.07)",
        zeroline=False,
        linecolor="rgba(56,189,248,0.22)",
        tickfont=dict(size=11),
    ),
    yaxis=dict(
        gridcolor="rgba(56,189,248,0.07)",
        zeroline=False,
        linecolor="rgba(0,0,0,0)",
        tickfont=dict(size=11),
    ),
    legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(size=11)),
    colorway=[CYAN, "#3B82F6", GREEN, PURPLE, AMBER, RED, "#38BDF8", "#34D399"],
    bargap=0.35,
)
try:
    _ops_template.layout.barcornerradius = 4  # plotly>=5.19 圆角柱状图，旧版自动跳过
except Exception:
    pass
pio.templates["ops_dark"] = _ops_template
pio.templates.default = "ops_dark"
TEMPLATE = "ops_dark"

# 各页强调色（页头色条 / 区块标记 / 导航选中态）：统一大屏冷色科技系，页间微差
PAGE_ACCENTS = {
    "经营总览": "#22D3EE",
    "红利分析": "#38BDF8",
    "红利 ROI & 代理质量": "#3B82F6",
    "代理佣金 & 退成": "#6366F1",
    "会员结构 & ARPU": "#A78BFA",
    "投注分析": "#8B5CF6",
    "客服分析": "#C084FC",
    "电访召回": "#7DD3FC",
    "实时波动 & DAU": "#2DD4A7",
    "代理团队 & 渠道": "#34D399",
    "代理 × 会员 明细": "#4ADE80",
    "游戏 & 场馆": "#10B981",
    # hero 标题别名（与导航名不同的页头标题）
    "渠道与代理": "#34D399",
    "游戏与场馆": "#10B981",
    "会员价值": "#A78BFA",
    "实时波动": "#2DD4A7",
    "代理佣金深度分析": "#6366F1",
    "投注分析（2026 年 4 月）": "#8B5CF6",
}

TEXT_COLUMNS = {
    "日期", "时间", "注册时间", "最后登录时间", "首存时间", "_imported_at",
    "代理名称", "代理类型", "代理编号", "代理账号", "会员账号", "用户名", "站点名称",
    "场馆名称", "场馆类型", "游戏类型", "游戏名称", "代理ID", "_source_file",
    "下注时间", "结算时间", "开赛时间", "min_t", "max_t",
    "venue", "venue_type", "vip_label", "member",
    "_snapshot_month", "_snapshot_date", "注册来源", "注册网址", "区号", "地区名称",
    "用户标签", "会员状态", "用户来源", "是否为代理", "VIP等级",
    "一级", "二级", "三级", "四级", "代理",
    # 红利报表
    "订单号", "上级代理", "上级代理编号", "会员等级", "红利类型", "红利标题",
    "发放方式", "领取方式", "派发平台", "是否需要流水", "申请时间", "发放时间",
    "操作人", "申请备注", "审核备注", "状态", "拦截备注", "发放钱包", "活动名称",
    # 电访召回报表（raw_winback）：月份/专员是文本，别被 normalize 转成数字 NaN
    "月份", "专员",
    # 市代月度结算（raw_agent_settlement_monthly）：「代理帐号」是异体「帐」非「账」，别被转成 NaN
    "代理帐号", "发展情况", "开户日期", "来源栏位",
    # 存取款订单（raw_finance_deposit / _withdraw）：时间/状态/渠道等都要保文本
    "订单时间", "完成时间", "支付方式", "订单状态", "取消原因", "会员等级", "接入节点",
    "存款账户ID", "银行卡信息", "风控审核完成时间", "站点审核完成时间", "订单时长",
    "风控审核人", "站点审核人", "出款人", "操作人", "支付备注", "预约奖励",
    # 代理佣金报表
    "佣金月份", "上级账号", "团队名称", "团队类型", "主线/副线", "线别",
    "是否在团队", "是否取消代理资格", "是否为主线", "成为代理时间", "加入团队时间",
    "发放人", "审核人", "审核时间", "审核状态", "调整原因", "发展人", "维护人",
    "备注", "佣金状态", "发放状态", "申请发放佣金", "发佣验证", "VIP专享",
    # 客服对话报表
    "终端", "访客ID", "对话ID", "新对话ID", "地区", "接待客服", "访客IP",
    "网站名称", "是否邀请评价", "满意度评价", "评价内容", "服务主题",
    "机器人标识", "对话内容", "_sheet", "_extracted_issue",
}

METRIC_META = {
    "公司输赢": {"source": "raw_platform_report", "fields": ["公司输赢"], "formula": "直接汇总原始字段【公司输赢】"},
    "有效投注额": {"source": "raw_platform_report", "fields": ["有效投注额"], "formula": "直接汇总原始字段【有效投注额】"},
    "存提差": {"source": "raw_platform_report / raw_finance_report", "fields": ["存提差"], "formula": "直接汇总原始字段【存提差】"},
    "实际总存款": {"source": "raw_finance_report", "fields": ["实际总存款"], "formula": "直接汇总原始字段【实际总存款】"},
    "首存转化率": {"source": "raw_platform_report", "fields": ["首存人数", "注册数"], "formula": "当前筛选范围内：首存人数 ÷ 注册数"},
    "公司收入": {"source": "raw_member_report", "fields": ["公司收入"], "formula": "当前直接汇总原始字段【公司收入】，不在前端做二次推导"},
    "TOP20投注占比": {"source": "raw_top_report", "fields": ["有效投注额", "_snapshot_month"], "formula": "当前快照月份：TOP20 有效投注额 ÷ 全部快照有效投注额"},
    "次月活跃留存率": {
        "source": "raw_member_report（需含 _snapshot_month）",
        "fields": ["会员账号", "存款额", "有效投注额", "_snapshot_month"],
        "formula": (
            "活跃定义：该月存款额 > 0 或有效投注额 > 0。"
            "计算：本月活跃会员中，次月仍为活跃的比例。"
            "公式 = COUNT(本月活跃 ∩ 次月活跃) ÷ COUNT(本月活跃)。"
            "卡片显示最近一期（如 2月→3月）的值，非平均值。"
            "注意：计算前会先套用当前筛选条件（默认：会员状态=启用、是否为代理=非代理）。"
        ),
    },
    "次月存款留存率": {
        "source": "raw_member_report（需含 _snapshot_month）",
        "fields": ["会员账号", "存款额", "_snapshot_month"],
        "formula": (
            "存款定义：该月存款额 > 0。"
            "计算：本月有存款的会员中，次月仍有存款的比例。"
            "公式 = COUNT(本月有存款 ∩ 次月有存款) ÷ COUNT(本月有存款)。"
            "卡片显示最近一期的值。"
            "注意：计算前会先套用当前筛选条件（默认：会员状态=启用、是否为代理=非代理）。"
        ),
    },
    "首存用户次月留存率": {
        "source": "raw_member_report（需含 _snapshot_month、首存时间）",
        "fields": ["会员账号", "首存时间", "存款额", "有效投注额", "_snapshot_month"],
        "formula": (
            "首存用户定义：首存时间落在该快照月份内的会员。"
            "次月活跃定义：次月存款额 > 0 或有效投注额 > 0。"
            "计算：本月首存用户中，次月仍为活跃的比例。"
            "公式 = COUNT(本月首存用户 ∩ 次月活跃) ÷ COUNT(本月首存用户)。"
            "卡片显示最近一期的值。"
            "注意：计算前会先套用当前筛选条件（默认：会员状态=启用、是否为代理=非代理）。"
        ),
    },
    "ARPU（用户均收）": {
        "source": "raw_member_report",
        "fields": ["公司收入", "会员账号"],
        "formula": (
            "ARPU（Average Revenue Per User，用户均收）= SUM(公司收入) ÷ DISTINCT(会员账号)。"
            "口径与当前筛选完全一致（含 VIP 等级、用户来源、会员状态、是否为代理等）。"
            "面板默认排除代理账号、仅看启用会员，避免被代理流水稀释。"
        ),
    },
    "日投注人次（DAU 近似）": {
        "source": "raw_realtime_bet",
        "fields": ["日期", "时段", "时段投注人数"],
        "formula": (
            "日投注人次 = 当日所有时段「时段投注人数」加总。"
            "是 DAU（Daily Active Users，日活跃用户数）的近似值——实时表是按【日期 × 时段 × 游戏类型】聚合的，"
            "同一会员若跨时段或跨游戏类型出现会被重复计入，所以严格意义不是去重后的 DAU。"
            "要拿到精准 DAU 需新增「每会员每日活跃」明细表。"
        ),
    },
    "退成（介绍人佣金分成）": {
        "source": "客服主管月报（手动汇整 / 4 月代理帐 xlsx）",
        "fields": ["业绩总额", "退成比例", "实际佣金"],
        "formula": (
            "退成（亦称「介绍人引荐分成」）：介绍人从其推荐下线代理之「业绩总额」中，"
            "按固定比例额外获取之佣金。该笔佣金由平台独立派发，"
            "与下线代理之主佣金互不影响、互不抵销。\n\n"
            "核心定义：\n"
            "• 业绩总额 = 下线代理当月所有有效业务产生之累计金额（与主佣金共用同一基础数据）\n"
            "• 主佣金 = 下线代理依其合约比例领取之常态月佣金（一般为业绩 × 55%）\n"
            "• 退成 = 平台在主佣金之外，额外支付给介绍人之分成，不从下线之主佣金中扣除\n"
            "• 叠付 = 主佣金 + 退成 皆由平台承担，平台之单笔业绩成本系两者「相加」（叠付），非「嵌套」\n\n"
            "退成比例分档（依 2026 年 4 月代理帐实际派发记录）：\n"
            "• 引荐分成（基础档）业绩 × 1% — 较远端引荐链\n"
            "• 标准退成 业绩 × 3% – 5% — 常态档位（占多数）\n"
            "• 新代理特殊安排 业绩 × 30% — 仅限特定新代理与平台之个案约定\n\n"
            "平台单笔业绩之总成本：\n"
            "• 常态：主佣金 55% + 退成 1–5% = 56% – 60%\n"
            "• 特殊：主佣金 55% + 退成 30% = 85%\n\n"
            "对照举例（2026-04 代理帐）：\n"
            "• YU → 豪Hao：业绩 246,971 → 主佣金 135,834（55%）+ 退成 12,349（5%）= 平台总支出 148,183（60%）\n"
            "• 咪娜 → 咬咬咬：业绩 285,386 → 主佣金 156,963（55%）+ 退成 8,562（3%）= 平台总支出 165,525（58%）\n"
            "• 豪Hao → 八万：业绩 374,806 → 主佣金 206,143（55%）+ 退成 3,748（1%）= 平台总支出 209,891（56%）"
        ),
    },
}


@dataclass
class FilterNotice:
    title: str
    detail: str


@st.cache_resource
def get_client():
    try:
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/bigquery"]
        )
        return bigquery.Client(project=PROJECT_ID, credentials=creds)
    except Exception:
        return bigquery.Client(project=PROJECT_ID)


@st.cache_data(ttl=300)
def query_bq(sql: str) -> pd.DataFrame:
    client = get_client()
    df = client.query(sql).to_dataframe()
    return normalize_dataframe(df)


@st.cache_data(ttl=300)
def load_table(table_name: str) -> pd.DataFrame:
    sql = f"SELECT * FROM `{BQ_PREFIX}.{table_name}`"
    return query_bq(sql)


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if col in TEXT_COLUMNS:
            continue
        if pd.api.types.is_numeric_dtype(out[col]):
            continue
        out[col] = pd.to_numeric(out[col], errors="coerce")
    return out


def clean_text_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().replace({"nan": None, "None": None, "": None})


def to_datetime_safe(series: pd.Series) -> pd.Series:
    s = series.copy()
    if s.dtype == object:
        s = s.astype(str).str.replace('="', '', regex=False).str.replace('"', '', regex=False)
    return pd.to_datetime(s, errors="coerce")


def fmt_num(v, suffix=""):
    if v is None or pd.isna(v):
        return "N/A"
    v = float(v)
    if abs(v) >= 1e8:
        return f"{v/1e8:,.2f}亿{suffix}"
    if abs(v) >= 1e4:
        return f"{v/1e4:,.2f}万{suffix}"
    if float(v).is_integer():
        return f"{int(v):,}{suffix}"
    return f"{v:,.2f}{suffix}"


def fmt_pct(v):
    if v is None or pd.isna(v):
        return "N/A"
    v = float(v)
    return f"{v*100:,.2f}%" if abs(v) <= 1 else f"{v:,.2f}%"


def safe_sum(df: pd.DataFrame, col: str) -> float:
    return float(df[col].sum()) if col in df.columns else 0.0


def safe_mean(df: pd.DataFrame, col: str) -> float:
    return float(df[col].mean()) if col in df.columns and len(df) else 0.0


def safe_nunique(df: pd.DataFrame, col: str) -> int:
    return int(df[col].nunique()) if col in df.columns else 0


def member_count(df: pd.DataFrame, account_col: str = '会员账号', agent_col: str = '代理') -> int:
    """全站统一「会员数」口径：按「会员账号 + 代理」去重计数。
    同一个账号名挂在不同代理底下 = 不同的人，只按账号会把他们并成一个、少算人头
    （跟「代理×会员」「新注册分析」一致）。没有代理列就退回按账号去重。"""
    if account_col not in df.columns:
        return 0
    if agent_col in df.columns:
        return int(df.drop_duplicates(subset=[account_col, agent_col]).shape[0])
    return int(df[account_col].nunique())



def tone_by_sign(v, invert: bool = False) -> Optional[str]:
    """按数值正负给指标卡状态色（纯展示用，不改任何计算）。正=绿、负=红；invert=True 时反过来（如成本类）。"""
    if v is None or pd.isna(v):
        return None
    v = float(v)
    if v == 0:
        return None
    positive = v > 0
    if invert:
        positive = not positive
    return 'good' if positive else 'bad'


def show_metric(card_col, label: str, value: str, delta: Optional[str] = None, help_text: Optional[str] = None,
                tone: Optional[str] = None, delta_tone: str = 'auto'):
    """指标卡。tone: good/bad/warn/accent 上状态色；delta_tone: auto(按±符号着色)/up/down/flat。"""
    tone_cls = f' tone-{tone}' if tone in ('good', 'bad', 'warn', 'accent') else ''
    delta_html = ''
    if delta:
        d = delta_tone
        if d == 'auto':
            ds = str(delta).strip()
            if ds.startswith(('+', '▲', '↑')):
                d = 'up'
            elif ds.startswith(('-', '▼', '↓')):
                d = 'down'
            else:
                d = 'flat'
        elif d in ('good',):
            d = 'up'
        elif d in ('bad',):
            d = 'down'
        elif d not in ('up', 'down'):
            d = 'flat'
        delta_html = f'<div class="metric-delta d-{d}">{escape(str(delta))}</div>'
    help_html = f'<div class="metric-help">{escape(help_text)}</div>' if help_text else ''
    html = (
        f'<div class="metric-card{tone_cls}">'
        f'<div class="metric-label">{escape(label)}</div>'
        f'<div class="metric-value">{escape(str(value))}</div>'
        f'{delta_html}'
        f'{help_html}'
        f'</div>'
    )
    with card_col:
        st.markdown(html, unsafe_allow_html=True)


def status_badge(text: str, tone: Optional[str] = None) -> str:
    """返回状态徽章 HTML（嵌进 badge-row / hero 用）。tone: good/bad/warn/None。"""
    cls = f'badge badge-{tone}' if tone in ('good', 'bad', 'warn') else 'badge'
    return f'<span class="{cls}">{escape(str(text))}</span>'


def section_header(title: str, subtitle: str = ""):
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)
    if subtitle:
        st.markdown(f'<div class="section-subtitle">{subtitle}</div>', unsafe_allow_html=True)



def hero(title: str, subtitle: str, updated_at: Optional[str] = None, extra_badges: Optional[List[str]] = None,
         basis: Optional[str] = None, detail: Optional[str] = None, source_badge: Optional[str] = None):
    """basis：标题下一行中文「数据基础」灰字（一眼看的）。
    detail：折叠「数据详情」markdown（想深入的点开，可放来源表/口径/更新方式）。
    source_badge：首个徽章文字，默认「数据库：BigQuery」；读谷歌表的页可传中文来源。"""
    accent = PAGE_ACCENTS.get(title)
    if accent:
        st.markdown(f'<style>:root {{ --accent-page: {accent}; }}</style>', unsafe_allow_html=True)
    badges = [f'<span class="badge">{escape(source_badge or "数据库：BigQuery")}</span>']
    if updated_at:
        badges.append(f'<span class="badge">最近更新：{updated_at}</span>')
    if extra_badges:
        badges.extend(extra_badges)
    basis_html = ''
    if basis:
        basis_html = (f'<div style="margin-top:.55rem;font-size:.82rem;color:#9fb3c8;'
                      f'line-height:1.55;">📊 数据基础：{escape(basis)}</div>')
    html = (
        '<div class="hero-card">'
        f'<div class="badge-row">{"".join(badges)}</div>'
        f'<div class="hero-title">{escape(title)}</div>'
        f'<div class="hero-subtitle">{escape(subtitle)}</div>'
        f'{basis_html}'
        '</div>'
    )
    st.markdown(html, unsafe_allow_html=True)
    if detail:
        with st.expander('ℹ️ 数据详情（来源 / 口径 / 更新方式）'):
            st.markdown(detail)


def latest_imported_at(*dfs: pd.DataFrame) -> str:
    vals = []
    for df in dfs:
        if '_imported_at' in df.columns:
            ts = to_datetime_safe(df['_imported_at'])
            if ts.notna().any():
                vals.append(ts.max())
    if not vals:
        return ""  # 没有 _imported_at（如早期 import 灌的红利表）→ 不显示「最近更新」徽章，比显示「未提供」干净
    return max(vals).strftime("%Y-%m-%d %H:%M")


def date_range_picker(df: pd.DataFrame, date_col: str, key_prefix: str, default_last_days: Optional[int] = None) -> Tuple[pd.DataFrame, Optional[pd.Timestamp], Optional[pd.Timestamp], Optional[str]]:
    """全站统一日期筛选：快捷预设(全部/本月/上月/近7天/近30天)或自订日期，单选不互相盖。
    跟「新注册分析」同款口径。返回 (筛后df, start_ts, end_ts, sel_month)，
    sel_month 只在选「本月/上月」这种单一自然月时给 'YYYY-MM'，其余为 None（呼叫端据此决定要不要按月切第二张表）。"""
    import datetime as _dt
    if date_col not in df.columns:
        return df, None, None, None
    out = df.copy()
    out[date_col] = to_datetime_safe(out[date_col])
    out = out[out[date_col].notna()].copy()
    if out.empty:
        return out, None, None, None
    min_d = out[date_col].min().date()
    max_d = out[date_col].max().date()
    first_this = max_d.replace(day=1)
    prev_last = first_this - _dt.timedelta(days=1)
    prev_first = prev_last.replace(day=1)
    presets = {
        '全部': (min_d, max_d, None),
        '本月': (max(min_d, first_this), max_d, first_this.strftime('%Y-%m')),
        '上月': (max(min_d, prev_first), min(max_d, prev_last), prev_first.strftime('%Y-%m')),
        '近7天': (max(min_d, max_d - _dt.timedelta(days=6)), max_d, None),
        '近30天': (max(min_d, max_d - _dt.timedelta(days=29)), max_d, None),
    }
    keys = list(presets.keys()) + ['自订日期']
    default_pick = '近30天' if default_last_days == 30 else ('近7天' if default_last_days == 7 else '全部')
    pick = st.radio('快速选择', keys, index=keys.index(default_pick),
                    horizontal=True, key=f'{key_prefix}_pick')
    if pick == '自订日期':
        c1, c2 = st.columns(2)
        with c1:
            start = st.date_input('开始日期', value=min_d, min_value=min_d, max_value=max_d, key=f'{key_prefix}_start')
        with c2:
            end = st.date_input('结束日期', value=max_d, min_value=min_d, max_value=max_d, key=f'{key_prefix}_end')
        sel_month = None
    else:
        start, end, sel_month = presets[pick]
    if start > end:
        start, end = end, start
    start_ts = pd.Timestamp(start)
    mask_end = pd.Timestamp(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    out = out[(out[date_col] >= start_ts) & (out[date_col] <= mask_end)].copy()
    st.caption(f'📅 当前显示 {start} ~ {end}，共 {len(out)} 条。改上面就立刻跟着变。')
    return out, start_ts, pd.Timestamp(end), sel_month


def apply_multiselect(df: pd.DataFrame, col: str, label: str, key: str, default_all: bool = True,
                      options_df: pd.DataFrame = None, auto_include_new: bool = True) -> pd.DataFrame:
    if col not in df.columns:
        return df
    # 选项来源用 options_df（通常是「未经其他筛选的全量」），让选项清单跨筛选/日期保持稳定，
    # 避免 Streamlit 多选框因选项集变动而保留旧选择、把新出现的值默默漏掉（会造成总数对不上）。
    src = options_df if options_df is not None else df
    if col not in src.columns:
        src = df
    options = [x for x in sorted(src[col].dropna().astype(str).unique().tolist()) if x not in ("", "nan", "None")]
    if not options:
        return df
    # auto_include_new：当数据新增了以前没有的值（如上传了带新域名的报表），
    # 自动把这些新值补进当前勾选，确保「没主动取消的东西永远默认显示」，根治 stale-default 漏算。
    if auto_include_new and default_all and key in st.session_state:
        seen_key = f'_{key}_seen'
        prev = st.session_state.get(seen_key, options)
        new_opts = [o for o in options if o not in prev]
        if new_opts:
            cur = [o for o in st.session_state.get(key, []) if o in options]
            st.session_state[key] = cur + new_opts
    st.session_state[f'_{key}_seen'] = options
    default = options if default_all else options[: min(8, len(options))]
    selected = st.multiselect(label, options, default=default, key=key)
    if selected:
        return df[df[col].astype(str).isin(selected)].copy()
    return df.iloc[0:0].copy()


def add_info_box(notices: List[FilterNotice]):
    if not notices:
        return
    lines = [f"<span class='tooltip-note'>!</span> <strong>{n.title}</strong>：{n.detail}" for n in notices]
    st.markdown(f"<div class='filter-note'>{'<br>'.join(lines)}</div>", unsafe_allow_html=True)


def tooltip_text(metric_name: str) -> str:
    meta = METRIC_META.get(metric_name)
    if not meta:
        return ""
    return f"来源：{meta['source']}"


def normalize_month_key(series: pd.Series) -> pd.Series:
    s = series.copy()
    s = s.astype(str).str.strip()
    s = s.replace({"nan": None, "None": None, "": None})
    mask = s.notna() & s.str.fullmatch(r"\d{6}")
    s.loc[mask] = s.loc[mask].str.slice(0, 4) + "-" + s.loc[mask].str.slice(4, 6)
    return s




def month_start_end(month_key: str) -> Tuple[pd.Timestamp, pd.Timestamp]:
    start = pd.Timestamp(f"{month_key}-01")
    end = (start + pd.offsets.MonthEnd(1)).normalize()
    return start, end

def get_snapshot_month(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if '_snapshot_month' in out.columns:
        out['__snapshot_month__'] = normalize_month_key(out['_snapshot_month'])
        return out
    if '_snapshot_date' in out.columns:
        dt = to_datetime_safe(out['_snapshot_date'])
        out['__snapshot_month__'] = dt.dt.strftime('%Y-%m')
        return out
    out['__snapshot_month__'] = None
    return out


def has_member_snapshot(df: pd.DataFrame) -> bool:
    if '_snapshot_month' in df.columns:
        s = normalize_month_key(df['_snapshot_month'])
        if s.notna().any():
            return True
    if '_snapshot_date' in df.columns:
        s = to_datetime_safe(df['_snapshot_date'])
        if s.notna().any():
            return True
    return False


def member_default_filters(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[FilterNotice], Dict[str, List[str]]]:
    notices: List[FilterNotice] = []
    current = {}
    out = df.copy()
    # defaults only if fields exist
    if '会员状态' in out.columns:
        opts = sorted(out['会员状态'].dropna().astype(str).unique())
        default = ['启用'] if '启用' in opts else opts
        sel = st.multiselect('会员状态', opts, default=default, key='mv_status')
        if set(sel) != set(opts):
            notices.append(FilterNotice('当前已应用会员状态筛选', f"仅保留：{' / '.join(sel)}"))
        out = out[out['会员状态'].astype(str).isin(sel)] if sel else out.iloc[0:0].copy()
        current['会员状态'] = sel
    if '是否为代理' in out.columns:
        opts = sorted(out['是否为代理'].dropna().astype(str).unique())
        default = ['非代理'] if '非代理' in opts else opts
        sel = st.multiselect('是否为代理', opts, default=default, key='mv_is_agent')
        if set(sel) != set(opts):
            notices.append(FilterNotice('当前已应用是否为代理筛选', f"仅保留：{' / '.join(sel)}"))
        out = out[out['是否为代理'].astype(str).isin(sel)] if sel else out.iloc[0:0].copy()
        current['是否为代理'] = sel
    if '用户来源' in out.columns:
        opts = sorted(out['用户来源'].dropna().astype(str).unique())
        sel = st.multiselect('用户来源', opts, default=opts, key='mv_user_source')
        out = out[out['用户来源'].astype(str).isin(sel)] if sel else out.iloc[0:0].copy()
        current['用户来源'] = sel

    with st.expander('高级筛选', expanded=False):
        cols = st.columns(3)
        with cols[0]:
            if 'VIP等级' in out.columns:
                out = apply_multiselect(out, 'VIP等级', 'VIP等级', 'mv_vip')
        with cols[1]:
            if '注册来源' in out.columns:
                out = apply_multiselect(out, '注册来源', '注册来源', 'mv_reg_source')
        with cols[2]:
            if '代理' in out.columns:
                agent_options = [x for x in sorted(out['代理'].dropna().astype(str).unique().tolist()) if x not in ('', 'nan', 'None')]
                if agent_options:
                    selected = st.multiselect('代理', agent_options, default=[], key='mv_agent')
                    if selected:
                        out = out[out['代理'].astype(str).isin(selected)].copy()
        if '用户标签' in out.columns:
            keyword = st.text_input('用户标签包含关键字', key='mv_tag_kw')
            if keyword:
                out = out[out['用户标签'].astype(str).str.contains(keyword, na=False)].copy()
    return out, notices, current


def render_metric_explainer(page_metrics: List[str]):
    with st.expander('本页指标口径说明', expanded=False):
        for name in page_metrics:
            meta = METRIC_META.get(name)
            if not meta:
                continue
            st.markdown(f"**{name}**")
            st.write(f"来源表：{meta['source']}")
            st.write(f"来源字段：{'、'.join(meta['fields'])}")
            st.write(f"口径：{meta['formula']}")
            st.markdown('---')


def compute_monthly_retention(member_df: pd.DataFrame) -> Optional[pd.DataFrame]:
    required = {'会员账号'}
    if not required.issubset(member_df.columns):
        return None
    df = member_df.copy()
    if '_snapshot_month' in df.columns:
        df['__snapshot_month__'] = normalize_month_key(df['_snapshot_month'])
    elif '_snapshot_date' in df.columns:
        df['__snapshot_month__'] = to_datetime_safe(df['_snapshot_date']).dt.strftime('%Y-%m')
    else:
        return None
    if df['__snapshot_month__'].isna().all():
        return None
    # 会员身份键含代理（同名挂不同代理=不同人），与 member_count / 矩阵口径一致
    if '代理' in df.columns:
        df['__account__'] = df['会员账号'].astype(str) + '\x01' + df['代理'].astype(str)
    else:
        df['__account__'] = df['会员账号'].astype(str)
    if '存款额' in df.columns:
        df['__has_deposit__'] = pd.to_numeric(df['存款额'], errors='coerce').fillna(0) > 0
    else:
        df['__has_deposit__'] = False
    if '有效投注额' in df.columns:
        df['__has_valid_bet__'] = pd.to_numeric(df['有效投注额'], errors='coerce').fillna(0) > 0
    else:
        df['__has_valid_bet__'] = False
    df['__is_active__'] = df['__has_deposit__'] | df['__has_valid_bet__']

    if '首存时间' in df.columns:
        first_deposit_month = to_datetime_safe(df['首存时间']).dt.strftime('%Y-%m')
        df['__is_first_deposit_month__'] = first_deposit_month == df['__snapshot_month__']
    else:
        df['__is_first_deposit_month__'] = False

    per_month = []
    months = sorted([m for m in df['__snapshot_month__'].dropna().unique().tolist()])
    for i in range(len(months) - 1):
        m = months[i]
        next_m = months[i + 1]
        cur = df[df['__snapshot_month__'] == m].copy()
        nxt = df[df['__snapshot_month__'] == next_m].copy()
        nxt_accounts = set(nxt['__account__'])
        nxt_active = set(nxt.loc[nxt['__is_active__'], '__account__'])
        nxt_deposit = set(nxt.loc[nxt['__has_deposit__'], '__account__'])

        cur_active = set(cur.loc[cur['__is_active__'], '__account__'])
        cur_deposit = set(cur.loc[cur['__has_deposit__'], '__account__'])
        cur_first = set(cur.loc[cur['__is_first_deposit_month__'], '__account__'])

        active_ret = len(cur_active & nxt_active) / len(cur_active) if cur_active else None
        deposit_ret = len(cur_deposit & nxt_deposit) / len(cur_deposit) if cur_deposit else None
        first_ret = len(cur_first & nxt_active) / len(cur_first) if cur_first else None

        per_month.append({
            '月份': m,
            '次月': next_m,
            '次月活跃留存率': active_ret,
            '次月存款留存率': deposit_ret,
            '首存用户次月留存率': first_ret,
            '本月活跃会员数': len(cur_active),
            '本月存款会员数': len(cur_deposit),
            '本月首存用户数': len(cur_first),
        })
    return pd.DataFrame(per_month)


def parse_realtime_time(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if '时间' not in out.columns:
        return out
    raw = out['时间'].astype(str).str.replace('="', '', regex=False).str.replace('"', '', regex=False).str.strip()
    # try patterns like 2026-01-31 23~24
    date_part = raw.str.extract(r'^(\d{4}-\d{2}-\d{2})')[0]
    slot_part = raw.str.extract(r'(\d{1,2}~\d{1,2})')[0]
    parsed_date = pd.to_datetime(date_part, errors='coerce')
    # fallback if already normal datetime
    fallback = pd.to_datetime(raw, errors='coerce')
    parsed_date = parsed_date.fillna(fallback.dt.normalize())
    out['日期'] = parsed_date
    out['时段'] = slot_part.fillna(fallback.dt.strftime('%H:00').where(fallback.notna()))
    return out


def _venue_category(name) -> str:
    s = str(name)
    if '真人' in s: return '真人'
    if '体育' in s or '體育' in s: return '体育'
    if '电竞' in s or '電競' in s: return '电竞'
    if '电子' in s or '電子' in s or '老虎' in s or 'PG' in s.upper() or 'PP' in s.upper(): return '电子'
    if '棋牌' in s: return '棋牌'
    if '捕鱼' in s or '捕魚' in s: return '捕鱼'
    if '哈希' in s or 'hash' in s.lower(): return '哈希'
    if '彩' in s: return '彩票'
    return '其他'


def _gsheet_client():
    import gspread
    from google.oauth2 import service_account
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly"])
    return gspread.authorize(creds)


def _recent_month_labels(n=3):
    import datetime as _dt
    today = _dt.date.today()
    y, m, labs = today.year, today.month, []
    for _ in range(n):
        labs.append(f"{y}年{m}月 运营日报")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return tuple(labs)


@st.cache_data(ttl=600)
def load_daily_ops(month_labels: tuple) -> pd.DataFrame:
    """读「运营日报」谷歌表 各月份 sheet 的「平台报表」分页，合并成每日 df。"""
    gc = _gsheet_client()
    frames = []
    for lab in month_labels:
        try:
            ws = gc.open(lab).worksheet("平台报表")
            recs = ws.get_all_records()
            if recs:
                frames.append(pd.DataFrame(recs))
        except Exception:
            continue
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    if "日期" not in df.columns:
        return pd.DataFrame()
    df = df[df["日期"].astype(str).str.match(r"\d{4}-\d{2}-\d{2}")].copy()
    for c in df.columns:
        if c == "日期":
            continue
        df[c] = pd.to_numeric(
            df[c].astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False),
            errors="coerce")
    return df.drop_duplicates(subset="日期", keep="last").sort_values("日期").reset_index(drop=True)


def render_recent_trend():
    hero("近期走势（日报）",
         "直接读每日「运营日报」谷歌表——自己拉日期看每日走势、选两段时间对比（如赛前 vs 世界杯）。"
         "数据来源：日报机器人每天自动收集；月度汇总请看「经营总览」。",
         source_badge='数据源：运营日报谷歌表（自动）',
         basis='每日「运营日报」谷歌表的「平台报表」分页（程序每日 10:00 自动写入）',
         detail=(
             '**分析范围**：每日经营走势（注册／首存／投注／公司输赢／存取款等）与两段时间对比。\n\n'
             '**数据来源**：运营日报谷歌表「平台报表」分页，自动读取近三个月。\n\n'
             '**更新方式**：程序每日 10:00 自动写入，本页约 10 分钟缓存，刷新即为最新，无需人工。\n\n'
             '**名词解释**见下方「名词说明」。完整数据来源对照见「数据说明」页。'
         ))
    import datetime as _dt
    try:
        df = load_daily_ops(_recent_month_labels(3))
    except Exception as e:
        st.error(f"读日报谷歌表失败：{str(e)[:150]}")
        return
    if df is None or df.empty:
        st.info("📭 暂时读不到日报数据（谷歌表「平台报表」分页为空，或服务账号没被授权读这份表）。")
        return
    df["_d"] = pd.to_datetime(df["日期"], errors="coerce")
    df = df[df["_d"].notna()].copy()
    min_d, max_d = df["_d"].min().date(), df["_d"].max().date()
    METRICS = [m for m in ["注册数", "首存人数", "投注人数", "有效投注额", "公司输赢",
                           "存款额", "取款额", "存提差", "公司净收入"] if m in df.columns]

    section_header("选日期范围", f"日报数据现有 {min_d} ~ {max_d}")
    c1, c2 = st.columns(2)
    with c1:
        start = st.date_input("开始日期", value=max(min_d, max_d - _dt.timedelta(days=13)),
                              min_value=min_d, max_value=max_d, key="rt_start")
    with c2:
        end = st.date_input("结束日期", value=max_d, min_value=min_d, max_value=max_d, key="rt_end")
    if start > end:
        start, end = end, start
    sel = df[(df["_d"] >= pd.Timestamp(start)) & (df["_d"] <= pd.Timestamp(end))].copy()
    st.caption(f"📅 当前显示 {start} ~ {end}，共 {len(sel)} 天。")

    def _s(c):
        return float(sel[c].sum()) if c in sel.columns and len(sel) else 0.0
    n = len(sel)
    reg, fd = _s("注册数"), _s("首存人数")
    vbet, win = _s("有效投注额"), _s("公司输赢")
    dep, wdr, net = _s("存款额"), _s("取款额"), _s("存提差")
    bonus, rebate = _s("红利"), _s("返水")
    netinc = _s("公司净收入")
    conv = (fd / reg) if reg else 0
    hold = (win / vbet) if vbet else 0

    section_header("区间总览", "合计（除标注外）")
    with st.expander("📖 名词说明（看不懂点这）"):
        st.markdown(
            "- **有效投注额**：客户实际有效的下注总额（已去掉对冲、取消、走盘那些不算的）。\n"
            "- **公司输赢**：平台对客户的输赢——客户输多少 = 公司赢多少。**正数 = 平台赢**。这是「还没扣成本」的毛赢（红利/返水/佣金/场馆费另算）。\n"
            "- **Hold%（杀率）**：公司输赢 ÷ 有效投注 = 平台从客户投注里实际赢走的比例，越高越赚。业界叫 Hold 或杀率。\n"
            "- **首存转化率**：首存人数 ÷ 新注册 = 新注册里有多少人真的存了第一笔钱。\n"
            "- **存提差（净流入）**：实际存款 − 实际取款 = 这段时间客户净流入平台多少钱。\n"
            "- **提存率**：取款 ÷ 存款，越高代表提款压力越大。\n"
            "- **公司净收入**：公司输赢扣掉红利/返水/代理佣金/场馆费等之后（口径以后台字段为准）。")
    r1 = st.columns(4)
    show_metric(r1[0], "天数", fmt_num(n))
    show_metric(r1[1], "新注册", f"{fmt_num(int(reg))}（日均 {reg/n:.0f}）" if n else "—")
    show_metric(r1[2], "首存人数", fmt_num(int(fd)))
    show_metric(r1[3], "首存转化率", fmt_pct(conv), help_text="首存人数 ÷ 新注册：新注册里多少人真的存了款", tone="accent")
    r2 = st.columns(4)
    show_metric(r2[0], "有效投注额", fmt_num(round(vbet)), help_text="客户有效下注总额（去掉对冲/取消/走盘）")
    show_metric(r2[1], "公司输赢", fmt_num(round(win)), tone=tone_by_sign(win),
                help_text="平台对客户的输赢，客户输=公司赢；正数=平台赢。未扣红利/返水/佣金/场馆费的毛赢")
    show_metric(r2[2], "Hold%（杀率）", f"{hold*100:.2f}%",
                help_text="公司输赢 ÷ 有效投注 = 平台从客户投注里实际赢走的比例，越高平台越赚（业界叫 Hold / 杀率）",
                tone="accent")
    show_metric(r2[3], "实际存款", fmt_num(round(dep)))
    r3 = st.columns(4)
    show_metric(r3[0], "存提差（净流入）", fmt_num(round(net)), tone=tone_by_sign(net),
                help_text="实际存款 − 实际取款 = 客户净流入平台的钱")
    show_metric(r3[1], "红利成本", fmt_num(round(bonus)), tone="warn")
    show_metric(r3[2], "返水成本", fmt_num(round(rebate)), tone="warn")
    show_metric(r3[3], "公司净收入", fmt_num(round(netinc)), tone=tone_by_sign(netinc),
                help_text="公司输赢扣掉红利/返水/佣金/场馆费等之后（口径以后台为准）")

    # 自动小结
    if n and "注册数" in sel.columns:
        peak_reg = sel.loc[sel["注册数"].idxmax()]
        line = (f"这 {n} 天：新注册 **{int(reg)}** 人（日均 {reg/n:.0f}）、首存 **{int(fd)}** 人（转化 **{conv*100:.1f}%**）；"
                f"有效投注 **{fmt_num(round(vbet))}**、公司输赢 **{fmt_num(round(win))}**（Hold **{hold*100:.2f}%**）；"
                f"实际存款 **{fmt_num(round(dep))}**、存提差 **{fmt_num(round(net))}**。"
                f"注册最高是 **{peak_reg['日期']}（{int(peak_reg['注册数'])} 人）**。")
        if "公司输赢" in sel.columns:
            pw, lw = sel.loc[sel["公司输赢"].idxmax()], sel.loc[sel["公司输赢"].idxmin()]
            line += f"公司输赢最高 {pw['日期']}（{fmt_num(round(pw['公司输赢']))}）、最低 {lw['日期']}（{fmt_num(round(lw['公司输赢']))}）。"
        st.info(line)

    # 关键比率每日走势
    section_header("关键比率走势", "首存转化率 / Hold% / 提存率 逐日看。")
    rate = sel.copy()
    if {"首存人数", "注册数"}.issubset(rate.columns):
        rate["首存转化率"] = (rate["首存人数"] / rate["注册数"].replace(0, pd.NA) * 100).round(1)
    if {"公司输赢", "有效投注额"}.issubset(rate.columns):
        rate["Hold%"] = (rate["公司输赢"] / rate["有效投注额"].replace(0, pd.NA) * 100).round(2)
    if {"取款额", "存款额"}.issubset(rate.columns):
        rate["提存率%"] = (rate["取款额"] / rate["存款额"].replace(0, pd.NA) * 100).round(1)
    ratecols = [c for c in ["首存转化率", "Hold%", "提存率%"] if c in rate.columns]
    if ratecols:
        rfig = go.Figure()
        for c in ratecols:
            rfig.add_trace(go.Scatter(x=rate["日期"], y=rate[c], mode="lines+markers", name=c))
        rfig.update_layout(height=300, template=TEMPLATE, margin=dict(l=10, r=10, t=10, b=10),
                           legend=dict(orientation="h", y=1.12), xaxis_title=None)
        st.plotly_chart(rfig, use_container_width=True)

    section_header("每日走势", "选要看的指标。")
    pick = st.multiselect("指标", METRICS,
                          default=[m for m in ["注册数", "投注人数", "公司输赢"] if m in METRICS],
                          key="rt_metrics")
    for m in pick:
        fig = px.bar(sel, x="日期", y=m, template=TEMPLATE, title=m, color_discrete_sequence=[BLUE])
        fig.update_layout(height=240, margin=dict(l=10, r=10, t=34, b=10), xaxis_title=None)
        st.plotly_chart(fig, use_container_width=True)

    section_header("两段时间对比", "选 A、B 两段，看日均差异（例如赛前 vs 世界杯）。")
    cc = st.columns(4)
    with cc[0]:
        a1 = st.date_input("A 开始", value=min_d, min_value=min_d, max_value=max_d, key="rt_a1")
    with cc[1]:
        a2 = st.date_input("A 结束", value=min(max_d, min_d + _dt.timedelta(days=6)),
                           min_value=min_d, max_value=max_d, key="rt_a2")
    with cc[2]:
        b1 = st.date_input("B 开始", value=max(min_d, max_d - _dt.timedelta(days=6)),
                           min_value=min_d, max_value=max_d, key="rt_b1")
    with cc[3]:
        b2 = st.date_input("B 结束", value=max_d, min_value=min_d, max_value=max_d, key="rt_b2")
    segA = df[(df["_d"] >= pd.Timestamp(a1)) & (df["_d"] <= pd.Timestamp(a2))]
    segB = df[(df["_d"] >= pd.Timestamp(b1)) & (df["_d"] <= pd.Timestamp(b2))]
    rows = []
    for m in METRICS:
        av = segA[m].mean() if len(segA) else 0
        bv = segB[m].mean() if len(segB) else 0
        chg = f"{(bv/av-1)*100:+.0f}%" if av else "—"
        rows.append({"指标": m, f"A日均({a1}~{a2})": round(av, 1),
                     f"B日均({b1}~{b2})": round(bv, 1), "变化": chg})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.caption("A、B 两段的日均对比；变化 = (B−A)/A。想比「赛前 vs 世界杯」：A 设 6/1~6/11、B 设 6/12 至今。")


def render_overview():
    platform = load_table('raw_platform_report')
    finance = load_table('raw_finance_report')
    hero('经营总览', '以平台报表与财务报表为主，查看经营结果、资金流与成本结构。', latest_imported_at(platform, finance),
         basis='经营报表＋财务报表（＋红利记录、代理结算）｜净利润＝公司输赢−红利/返水/代理佣金/集团分成',
         detail=(
             '**分析范围**：平台整体经营结果、资金流（存款／取款／净流入）、成本结构与真实净利润。\n\n'
             '**数据来源（后台导出 → 上传）**：\n'
             '- 经营报表（报表中心→经营报表）：公司输赢／注册／首存／存取款／有效投注\n'
             '- 财务报表（报表中心→财务报表）：红利／返水／代理佣金／集团分成\n'
             '- 红利记录（会员管理→VIP记录管理→红利记录）：红利构成拆分\n'
             '- 代理结算月报（客服主管提供）：实际佣金派发\n\n'
             '**计算口径**：真实平台净盈利 ＝ 公司输赢 ＋ 提前结算 ＋ 帐户调整 − 红利 − 返水 − 代理佣金 − 集团分成。\n\n'
             '**更新方式**：手动上传（每月导出后于「数据上传」页上传）。完整数据来源对照见「数据说明」页。'
         ))

    headline_slot = st.container()  # 顶部核心结论横幅：数值在页面下方算齐后回填到这里

    platform, start, end, month = date_range_picker(platform, '日期', 'ov', default_last_days=None)
    if month and '时间' in finance.columns:
        finance['时间'] = to_datetime_safe(finance['时间'])
        finance = finance[finance['时间'].dt.strftime('%Y-%m') == month].copy()
    elif '时间' in finance.columns and start is not None and end is not None:
        finance['时间'] = to_datetime_safe(finance['时间'])
        finance = finance[(finance['时间'] >= start) & (finance['时间'] < end + pd.Timedelta(days=1))].copy()

    if platform.empty:
        st.warning('当前筛选条件下无数据。')
        return

    kpi_winloss = safe_sum(platform, '公司输赢')
    kpi_diff = safe_sum(platform, '存提差')
    cols = st.columns(6)
    show_metric(cols[0], '公司输赢', fmt_num(kpi_winloss), help_text=tooltip_text('公司输赢'),
                tone=tone_by_sign(kpi_winloss))
    show_metric(cols[1], '有效投注额', fmt_num(safe_sum(platform, '有效投注额')), help_text=tooltip_text('有效投注额'))
    show_metric(cols[2], '实际总存款', fmt_num(safe_sum(finance, '实际总存款')), help_text=tooltip_text('实际总存款'))
    show_metric(cols[3], '存提差', fmt_num(kpi_diff), help_text=tooltip_text('存提差'),
                tone=tone_by_sign(kpi_diff))
    show_metric(cols[4], '注册数', fmt_num(safe_sum(platform, '注册数')))
    show_metric(cols[5], '首存转化率', fmt_pct(safe_sum(platform, '首存人数') / safe_sum(platform, '注册数') if safe_sum(platform, '注册数') else None), help_text=tooltip_text('首存转化率'),
                tone='accent')

    # ── 扣除代理红利后的调整净收入 ──
    agent_activities = st.session_state.get('agent_bonus_activities', [])
    if agent_activities and '公司净收入' in platform.columns:
        # 加载红利数据并计算代理相关红利
        bonus_for_adj = load_table('raw_bonus_report')
        if not bonus_for_adj.empty and '活动名称' in bonus_for_adj.columns and '红利金额' in bonus_for_adj.columns:
            # 对齐日期筛选
            if '申请时间' in bonus_for_adj.columns:
                bonus_for_adj['申请时间'] = to_datetime_safe(bonus_for_adj['申请时间'])
                if month:
                    bonus_for_adj = bonus_for_adj[bonus_for_adj['申请时间'].dt.strftime('%Y-%m') == month].copy()
                elif start is not None and end is not None:
                    bonus_for_adj = bonus_for_adj[
                        (bonus_for_adj['申请时间'] >= start) &
                        (bonus_for_adj['申请时间'] < end + pd.Timedelta(days=1))
                    ].copy()

            agent_mask = bonus_for_adj['活动名称'].astype(str).isin(agent_activities)
            agent_bonus_amt = float(bonus_for_adj.loc[agent_mask, '红利金额'].sum())
            total_bonus_amt = float(bonus_for_adj['红利金额'].sum())
            real_member_bonus = total_bonus_amt - agent_bonus_amt
            net_income = safe_sum(platform, '公司净收入')

            section_header('红利构成拆分', '根据红利分析页选定的代理相关活动，拆分红利支出构成。')
            adj_cols = st.columns(4)
            show_metric(adj_cols[0], '红利总支出', fmt_num(total_bonus_amt))
            show_metric(adj_cols[1], '代理相关红利', fmt_num(agent_bonus_amt),
                        help_text=f'来自 {len(agent_activities)} 个代理相关活动', tone='warn')
            show_metric(adj_cols[2], '真实会员活动红利', fmt_num(real_member_bonus),
                        help_text='红利总支出 - 代理相关红利')
            show_metric(adj_cols[3], '代理红利占比', fmt_pct(agent_bonus_amt / total_bonus_amt if total_bonus_amt else None),
                        tone='warn')

    # ── 真实平台净盈利计算 ──
    section_header('真实平台净盈利', '综合系统数据与手动输入的财务数据，还原真实盈亏。')

    # 从BQ自动读取的项目
    company_winloss = safe_sum(platform, '公司输赢')
    early_settle = safe_sum(platform, '提前结算')
    account_adjust = safe_sum(platform, '账户调整')
    rebate = safe_sum(platform, '返水')
    bonus_total = safe_sum(platform, '红利')
    tips = safe_sum(platform, '打赏收入')
    group_share = safe_sum(platform, '集团分成')
    agent_comm_sys = safe_sum(platform, '代理佣金')

    st.markdown('**系统自动读取（来自平台报表）：**')
    auto_cols = st.columns(4)
    show_metric(auto_cols[0], '公司输赢', fmt_num(company_winloss))
    show_metric(auto_cols[1], '提前结算（计营利）', fmt_num(early_settle))
    show_metric(auto_cols[2], '帐户调整（计营利）', fmt_num(account_adjust))
    show_metric(auto_cols[3], '集团分成（系统费）', fmt_num(group_share))

    auto_cols2 = st.columns(4)
    show_metric(auto_cols2[0], '红利', fmt_num(bonus_total))
    show_metric(auto_cols2[1], '返水', fmt_num(rebate))
    show_metric(auto_cols2[2], '打赏收入', fmt_num(tips))
    show_metric(auto_cols2[3], '代理佣金（系统）', fmt_num(agent_comm_sys))

    # 自动从 raw_agent_settlement_summary 拉「实际佣金派发总额」(总计发放) 当默认值
    auto_commission_default = 0.0
    settle_month_key = None
    try:
        _settle_summary = load_table('raw_agent_settlement_summary')
        if not _settle_summary.empty and '月份' in _settle_summary.columns and month:
            _sub = _settle_summary[_settle_summary['月份'].astype(str) == month]
            if not _sub.empty:
                settle_month_key = month
                # 项目 含 "总计发放" 之 row,取 abs(金额) 当默认
                _proj = _sub['项目'].astype(str).str.replace('⭐️', '', regex=False).str.strip()
                _total_row = _sub[_proj == '总计发放']
                if not _total_row.empty:
                    auto_commission_default = abs(float(_total_row['金额'].iloc[0]))
    except Exception:
        pass

    # 手动输入的项目（部分可由客服主管月报自动填）
    st.markdown('**手动输入 / 自动填入（清空则不计入）：**')
    if settle_month_key:
        st.caption(f'📌 「实际佣金派发总额」已从 raw_agent_settlement_summary（{settle_month_key} 总计发放）自动带入 {fmt_num(auto_commission_default)}。可手动覆盖。')

    manual_cols = st.columns(4)
    with manual_cols[0]:
        channel_fee = st.number_input('通道手续费（正数）', min_value=0.0, value=0.0, step=10000.0, key=f'channel_fee_{month or "all"}', help='存取款支付通道费用，填正数（目前无 BQ 来源，需手填）')
    with manual_cols[1]:
        project_adjust = st.number_input('项目调整-财务承担', value=0.0, step=10000.0, key=f'project_adjust_{month or "all"}', help='正数=平台收入，负数=平台支出（目前无 BQ 来源，需手填）')
    with manual_cols[2]:
        real_commission = st.number_input(
            '实际佣金派发总额（正数）',
            min_value=0.0, value=auto_commission_default, step=10000.0,
            key=f'real_commission_{month or "all"}',
            help='含平台直发+兑台预派的佣金总额，填正数。已从客服主管月报「总计发放」自动带入,可手动覆盖。',
        )
    with manual_cols[3]:
        agent_refund = st.number_input('代理回帐', min_value=0.0, value=0.0, step=10000.0, key=f'agent_refund_{month or "all"}', help='代理退回的金额，填正数。口径与客服主管月报存在差异（仅部分笔回款入帐），目前仍需手填')

    # 计算真实净盈利
    # 如果填了实际佣金且选了代理相关活动，红利用真实会员红利（扣掉代理红利）避免重复计算
    agent_activities = st.session_state.get('agent_bonus_activities', [])
    if real_commission > 0 and agent_activities:
        # 有选代理活动 + 有填实际佣金 → 用真实会员红利
        bonus_for_calc = load_table('raw_bonus_report')
        if not bonus_for_calc.empty and '活动名称' in bonus_for_calc.columns and '红利金额' in bonus_for_calc.columns:
            if '申请时间' in bonus_for_calc.columns:
                bonus_for_calc['申请时间'] = to_datetime_safe(bonus_for_calc['申请时间'])
                if month:
                    bonus_for_calc = bonus_for_calc[bonus_for_calc['申请时间'].dt.strftime('%Y-%m') == month].copy()
                elif start is not None and end is not None:
                    bonus_for_calc = bonus_for_calc[
                        (bonus_for_calc['申请时间'] >= start) &
                        (bonus_for_calc['申请时间'] < end + pd.Timedelta(days=1))
                    ].copy()
            agent_mask = bonus_for_calc['活动名称'].astype(str).isin(agent_activities)
            agent_bonus_in_calc = float(bonus_for_calc.loc[agent_mask, '红利金额'].sum())
            real_bonus = bonus_total - agent_bonus_in_calc
        else:
            real_bonus = bonus_total
        bonus_note = f'真实会员红利（已扣除代理红利 {fmt_num(agent_bonus_in_calc)}）'
    else:
        real_bonus = bonus_total
        bonus_note = '红利（含代理红利，建议在红利分析页选择代理活动+填入实际佣金以避免重复计算）'

    gross_income = company_winloss + early_settle + account_adjust + tips
    costs = real_bonus + rebate + group_share + channel_fee - project_adjust
    commission_total = real_commission if real_commission > 0 else agent_comm_sys
    refund = agent_refund

    real_profit = gross_income - costs - commission_total + refund

    st.markdown('---')
    result_cols = st.columns(3)
    show_metric(result_cols[0], '毛收入', fmt_num(gross_income),
                help_text='公司输赢 + 提前结算 + 帐户调整 + 打赏',
                tone=tone_by_sign(gross_income))
    show_metric(result_cols[1], '总成本', fmt_num(costs + commission_total - refund),
                help_text=bonus_note + ' + 返水 + 集团分成 + 通道手续费 - 项目调整 + 佣金 - 代理回帐',
                tone='warn')
    show_metric(result_cols[2], '真实平台净盈利', fmt_num(real_profit),
                help_text='毛收入 - 总成本',
                tone=tone_by_sign(real_profit))
    if real_commission > 0 and agent_activities:
        st.caption('✅ 已使用实际佣金 + 扣除代理红利，避免重复计算')
    elif real_commission > 0:
        st.caption('⚠️ 已使用实际佣金，但未选择代理相关活动，红利可能含代理部分导致重复扣除。请到红利分析页选择代理活动。')
    else:
        st.caption('⚠️ 佣金使用系统数据（可能偏低），建议填入实际佣金派发总额以获得准确结果')

    # ── 经营摘要：净利瀑布 + 业务线盈利贡献 + 月环比 + 摘要 ──
    valid_bet = safe_sum(platform, '有效投注额')
    fx = st.number_input('兑台汇率（人民币→台币）', min_value=0.0, value=4.35, step=0.01,
                         key=f'fx_rate_{month or "all"}', help='把人民币净利换算成台币，仅展示用')
    net_rmb = real_profit
    net_twd = net_rmb * fx
    hold = (company_winloss / valid_bet) if valid_bet else None

    section_header('经营摘要', '本期净利润、损益结构与业务线盈利贡献。')
    bcols = st.columns(4)
    show_metric(bcols[0], '平台净盈利（人民币）', fmt_num(net_rmb), help_text='即上方「真实平台净盈利」',
                tone=tone_by_sign(net_rmb))
    show_metric(bcols[1], f'平台净盈利（台币 @{fx:g}）', fmt_num(net_twd),
                tone=tone_by_sign(net_twd))
    show_metric(bcols[2], '有效投注额（流水）', fmt_num(valid_bet))
    show_metric(bcols[3], '整体盈余比例', fmt_pct(hold),
                help_text='公司输赢 ÷ 有效投注额（Hold %）',
                tone='accent')

    other_rev = early_settle + account_adjust + tips + project_adjust
    promo = real_bonus + rebate
    sys_chan = group_share + channel_fee
    agent_net = commission_total - refund

    with st.container(border=True):
        section_header('损益瀑布分析', '绿色为收入项，红色为成本项，蓝色为净利润。')
        wf = go.Figure(go.Waterfall(
            orientation='v',
            measure=['relative', 'relative', 'relative', 'relative', 'relative', 'total'],
            x=['公司输赢', '其他营收', '优惠成本<br>(红利+返水)', '平台费用<br>(系统费+通道费)', '代理佣金<br>(净额)', '净利润'],
            y=[company_winloss, other_rev, -promo, -sys_chan, -agent_net, net_rmb],
            text=[fmt_num(company_winloss), fmt_num(other_rev), fmt_num(-promo),
                  fmt_num(-sys_chan), fmt_num(-agent_net), fmt_num(net_rmb)],
            textposition='outside',
            textfont=dict(size=12),
            connector={'line': {'color': 'rgba(150,170,210,0.30)', 'width': 1}},
            increasing={'marker': {'color': GREEN}},
            decreasing={'marker': {'color': RED}},
            totals={'marker': {'color': BLUE}},
        ))
        wf.update_layout(height=430, template=TEMPLATE, showlegend=False, yaxis_title='金额')
        st.plotly_chart(wf, width='stretch')

    # 哪块业务在赚（BQ 场馆报表，仅在选定单一月份时显示）
    cat = pd.DataFrame()
    if month:
        try:
            vsql = (
                "SELECT `场馆名称` AS venue, "
                "SUM(SAFE_CAST(REPLACE(REPLACE(`公司输赢`,'=\"',''),'\"','') AS FLOAT64)) AS win "
                f"FROM `{BQ_PREFIX}.raw_game_report_venue` "
                f"WHERE `时间` LIKE '{month}%' GROUP BY venue"
            )
            vdf = get_client().query(vsql).to_dataframe()
            if not vdf.empty:
                vdf['类别'] = vdf['venue'].map(_venue_category)
                cat = vdf.groupby('类别', as_index=False)['win'].sum().sort_values('win', ascending=False)
        except Exception:
            cat = pd.DataFrame()

    if not cat.empty:
        with st.container(border=True):
            section_header('业务线盈利贡献', '各场馆类别本期公司输赢（场馆报表口径）。')
            figcat = go.Figure(go.Bar(
                x=cat['类别'], y=cat['win'],
                marker_color=[GREEN if v >= 0 else RED for v in cat['win']],
                text=[fmt_num(v) for v in cat['win']], textposition='outside',
                textfont=dict(size=12),
            ))
            figcat.update_layout(height=360, template=TEMPLATE, showlegend=False, yaxis_title='公司输赢')
            st.plotly_chart(figcat, width='stretch')

    # 月环比（BQ 平台报表）：流水 + 公司输赢
    mom_txt = ''
    bet_mom_pct = None
    win_mom_pct = None
    if month:
        try:
            yy, mm = int(month[:4]), int(month[5:7])
            py, pm = (yy, mm - 1) if mm > 1 else (yy - 1, 12)
            prev_ym = f'{py}-{pm:02d}'
            psql = (
                "SELECT SUBSTR(`日期`,1,7) AS ym, "
                "SUM(SAFE_CAST(REPLACE(REPLACE(`有效投注额`,'=\"',''),'\"','') AS FLOAT64)) AS bet, "
                "SUM(SAFE_CAST(REPLACE(REPLACE(`公司输赢`,'=\"',''),'\"','') AS FLOAT64)) AS win "
                f"FROM `{BQ_PREFIX}.raw_platform_report` "
                f"WHERE SUBSTR(`日期`,1,7) IN ('{month}','{prev_ym}') GROUP BY ym"
            )
            mdf = get_client().query(psql).to_dataframe()
            if len(mdf) == 2:
                mdf = mdf.set_index('ym')
                cb, pb = float(mdf.loc[month, 'bet']), float(mdf.loc[prev_ym, 'bet'])
                if pb:
                    bet_mom_pct = (cb - pb) / pb * 100
                    mom_txt = f' 有效投注额较上月（{prev_ym}）{"▲" if bet_mom_pct >= 0 else "▼"} {abs(bet_mom_pct):.1f}%。'
                cw, pw = float(mdf.loc[month, 'win']), float(mdf.loc[prev_ym, 'win'])
                if pw:
                    win_mom_pct = (cw - pw) / abs(pw) * 100
        except Exception:
            mom_txt = ''

    # 回填顶部核心结论横幅（口径 = 真实平台净盈利，含手动输入项）
    with headline_slot:
        def _kb_delta(p):
            cls = 'kb-up' if p >= 0 else 'kb-down'
            return f'<span class="{cls}">{"▲" if p >= 0 else "▼"} {abs(p):.1f}%</span>'
        _net_color = GREEN if net_rmb >= 0 else RED
        items = [
            f'<span class="kb-main" style="color:{_net_color};">'
            f'{escape(month) if month else "当前范围"} 平台净利润 {fmt_num(net_rmb)}</span>',
            f'<span class="kb-item">约合台币 {fmt_num(net_twd)}</span>',
        ]
        if hold is not None:
            items.append(f'<span class="kb-item">盈余率 {fmt_pct(hold)}</span>')
        if win_mom_pct is not None:
            items.append(f'<span class="kb-item">公司输赢环比 {_kb_delta(win_mom_pct)}</span>')
        if bet_mom_pct is not None:
            items.append(f'<span class="kb-item">流水环比 {_kb_delta(bet_mom_pct)}</span>')
        st.markdown(f'<div class="kpi-banner">{"".join(items)}</div>', unsafe_allow_html=True)

    def _hl(text, color=None):
        style = f'color:{color};font-weight:600;' if color else 'font-weight:600;'
        return f'<span style="{style}">{escape(str(text))}</span>'

    big3 = sorted([('代理佣金', abs(agent_net)), ('红利+返水', abs(promo)),
                   ('系统费+通道费', abs(sys_chan))], key=lambda x: -x[1])
    big3_txt = '、'.join([f'{n} {fmt_num(v)}' for n, v in big3])
    s_scale = (f'本期有效投注额 {_hl(fmt_num(valid_bet))}，公司输赢 {_hl(fmt_num(company_winloss))}'
               + (f'（盈余率 {hold*100:.2f}%）' if hold is not None else '') + '。' + mom_txt)
    s_net = f'扣除各项成本后，平台净利润 {_hl(fmt_num(net_rmb), GREEN)}（约合台币 {_hl(fmt_num(net_twd), GREEN)}）。'
    s_cost = f'主要成本项：{big3_txt}。'
    cat_txt = ''
    if not cat.empty:
        top2 = cat.head(2)
        tot_win = cat['win'].sum()
        if tot_win:
            share = top2['win'].sum() / tot_win
            cat_txt = f'盈利贡献集中于 {_hl("、".join(top2["类别"].tolist()))}，合计约占 {_hl(f"{share*100:.0f}%")}。'
    st.markdown(
        '<div class="hero-card" style="padding:1.1rem 1.4rem;line-height:2.05;">'
        f'<div>{s_scale}</div>'
        f'<div>{s_net}</div>'
        f'<div style="margin-top:0.35rem;color:#9fb0d0;">{s_cost}{(" " + cat_txt) if cat_txt else ""}</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('经营结果趋势', '公司输赢与公司净收入分开经营规模字段显示，避免量级混在一起。')
            trend_cols = [c for c in ['公司输赢', '公司净收入'] if c in platform.columns]
            if trend_cols and '日期' in platform.columns:
                daily = platform[['日期'] + trend_cols].copy()
                daily['日期'] = to_datetime_safe(daily['日期'])
                daily = daily[daily['日期'].notna()].sort_values('日期')
                daily = daily.groupby('日期', as_index=False)[trend_cols].sum()
                fig = go.Figure()
                if '公司输赢' in daily.columns:
                    fig.add_trace(go.Bar(x=daily['日期'], y=daily['公司输赢'], name='公司输赢', marker_color=BLUE, opacity=0.55))
                if '公司净收入' in daily.columns:
                    fig.add_trace(go.Scatter(x=daily['日期'], y=daily['公司净收入'], name='公司净收入', mode='lines+markers',
                                             line=dict(color=GREEN, width=2.5, shape='spline', smoothing=0.6), marker=dict(size=5)))
                fig.update_layout(
                    height=380,
                    template=TEMPLATE,
                    barmode='overlay',
                    legend=dict(orientation='h', y=-0.18),
                    hovermode='x unified',
                    xaxis_title=None,
                    yaxis_title='金额'
                )
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            section_header('资金流趋势', '字段使用财务报表原始字段。')
            if not finance.empty and '时间' in finance.columns:
                fin = finance.copy()
                fin['时间'] = to_datetime_safe(fin['时间'])
                fin = fin[fin['时间'].notna()].sort_values('时间')
                value_cols = [c for c in ['实际总存款', '实际总提款', '存提差'] if c in fin.columns]
                if value_cols:
                    melted = fin.melt(id_vars='时间', value_vars=value_cols, var_name='指标', value_name='值')
                    fig = px.bar(melted, x='时间', y='值', color='指标', barmode='group', template=TEMPLATE,
                                 color_discrete_sequence=[GREEN, RED, BLUE])
                    fig.update_layout(height=380, legend=dict(orientation='h', y=-0.18), hovermode='x unified',
                                      xaxis_title=None)
                    st.plotly_chart(fig, width='stretch')

    if '有效投注额' in platform.columns and '日期' in platform.columns:
        with st.container(border=True):
            section_header('经营规模趋势', '有效投注额单独显示，避免与经营结果共用一张图。')
            scale_df = platform[['日期', '有效投注额']].copy()
            scale_df['日期'] = to_datetime_safe(scale_df['日期'])
            scale_df = scale_df[scale_df['日期'].notna()].sort_values('日期')
            scale_df = scale_df.groupby('日期', as_index=False)['有效投注额'].sum()
            fig_scale = px.area(scale_df, x='日期', y='有效投注额', template=TEMPLATE)
            fig_scale.update_traces(line_color=CYAN, line_shape='spline', line_smoothing=0.6,
                                    fillcolor='rgba(34,211,238,0.16)')
            fig_scale.update_layout(height=300, hovermode='x unified', showlegend=False, xaxis_title=None, yaxis_title='金额')
            st.plotly_chart(fig_scale, width='stretch')

    c3, c4 = st.columns([1.1, 0.9])
    with c3:
        with st.container(border=True):
            section_header('核心日报明细')
            table_cols = [c for c in ['日期', '公司输赢', '公司净收入', '有效投注额', '注册数', '首存人数'] if c in platform.columns]
            st.dataframe(platform[table_cols].sort_values('日期', ascending=False), width='stretch', hide_index=True)
    with c4:
        with st.container(border=True):
            section_header('成本结构')
            cost_cols = [c for c in ['红利', '返水', '代理佣金'] if c in platform.columns]
            if cost_cols:
                cost_df = pd.DataFrame({'项目': cost_cols, '金额': [safe_sum(platform, c) for c in cost_cols]})
                cost_sum_total = float(cost_df['金额'].sum())
                fig = px.pie(cost_df, names='项目', values='金额', hole=0.58, template=TEMPLATE,
                             color_discrete_sequence=[BLUE, CYAN, PURPLE])
                fig.update_traces(textinfo='percent', textfont_size=12,
                                  marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(
                    height=310,
                    legend=dict(orientation='h', y=-0.12),
                    annotations=[dict(text=f'总成本<br><b>{fmt_num(cost_sum_total)}</b>',
                                      x=0.5, y=0.5, showarrow=False,
                                      font=dict(size=14, color='#f0f5ff'))],
                )
                st.plotly_chart(fig, width='stretch')
            summary = []
            if '公司输赢' in platform.columns:
                summary.append(f"• 公司输赢：{fmt_num(safe_sum(platform, '公司输赢'))}")
            if '存提差' in platform.columns:
                summary.append(f"• 存提差：{fmt_num(safe_sum(platform, '存提差'))}")
            if '首存人数' in platform.columns and '注册数' in platform.columns and safe_sum(platform, '注册数'):
                summary.append(f"• 首存转化率：{fmt_pct(safe_sum(platform, '首存人数')/safe_sum(platform, '注册数'))}")
            if cost_cols and safe_sum(platform, '公司输赢'):
                cost_total = sum(safe_sum(platform, c) for c in cost_cols)
                summary.append(f"• 成本 / 公司输赢占比：{fmt_pct(cost_total / safe_sum(platform, '公司输赢'))}")
            st.markdown('<div style="color:#aebcd9; line-height:1.9; padding:0.2rem 0.2rem 0.6rem;">' + '<br>'.join(summary) + '</div>', unsafe_allow_html=True)

    render_metric_explainer(['公司输赢', '有效投注额', '实际总存款', '存提差', '首存转化率'])


def render_channel_agent():
    agent = load_table('raw_agent_report')
    promo = load_table('raw_promotion_report')
    hero('渠道与代理', '查看代理与渠道表现；字段命名尽量保留原始字段。', latest_imported_at(agent, promo),
         basis='代理报表＋推广报表（后台导出·每月上传）',
         detail=(
             '**分析范围**：代理与渠道表现、代理分层、渠道结构。\n\n'
             '**数据来源（后台导出 → 上传）**：\n'
             '- 代理报表（报表中心→代理报表）\n'
             '- 推广报表（报表中心→推广报表）\n\n'
             '**更新方式**：手动上传（每月导出后于「数据上传」页上传）。完整对照见「数据说明」页。'
         ))

    agent, start, end, month = date_range_picker(agent, '日期', 'ag', default_last_days=None)
    if month and '日期' in promo.columns:
        promo['日期'] = to_datetime_safe(promo['日期'])
        promo = promo[promo['日期'].dt.strftime('%Y-%m') == month].copy()
    elif '日期' in promo.columns and start is not None and end is not None:
        promo['日期'] = to_datetime_safe(promo['日期'])
        promo = promo[(promo['日期'] >= start) & (promo['日期'] < end + pd.Timedelta(days=1))].copy()

    c1, c2, c3 = st.columns(3)
    with c1:
        agent = apply_multiselect(agent, '代理类型', '代理类型', 'ag_type')
    with c2:
        if '代理名称' in agent.columns:
            kw = st.text_input('搜索代理名称', key='ag_name_kw')
            if kw:
                agent = agent[agent['代理名称'].astype(str).str.contains(kw, case=False, na=False)].copy()
    with c3:
        if '一级' in promo.columns:
            lvl1 = [x for x in sorted(promo['一级'].dropna().astype(str).unique().tolist()) if x not in ('', 'nan', 'None')]
            if lvl1:
                sel = st.multiselect('一级', lvl1, default=[], key='promo_l1')
                if sel:
                    promo = promo[promo['一级'].astype(str).isin(sel)].copy()

    with st.expander('高级筛选', expanded=False):
        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            if '二级' in promo.columns:
                promo = apply_multiselect(promo, '二级', '二级', 'promo_l2', default_all=False)
        with cc2:
            if '三级' in promo.columns:
                promo = apply_multiselect(promo, '三级', '三级', 'promo_l3', 'promo_l3' if False else 'promo_l3')
        with cc3:
            if '四级' in promo.columns:
                promo = apply_multiselect(promo, '四级', '四级', 'promo_l4', default_all=False)

    add_info_box([FilterNotice('一级渠道图表口径', '当前“TOP 一级渠道”直接使用推广报表原始字段【一级】；二级 / 三级 / 四级先收在高级筛选。')])

    ca_winloss = safe_sum(agent, '公司输赢')
    cols = st.columns(5)
    show_metric(cols[0], '活跃代理数', fmt_num(safe_nunique(agent, '代理名称')))
    show_metric(cols[1], '有效投注额', fmt_num(safe_sum(agent, '有效投注额')))
    show_metric(cols[2], '公司输赢', fmt_num(ca_winloss), tone=tone_by_sign(ca_winloss))
    show_metric(cols[3], '代理佣金', fmt_num(safe_sum(agent, '代理佣金')), tone='warn')
    show_metric(cols[4], '推广收入', fmt_num(safe_sum(promo, '推广收入')))

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('TOP 代理（按有效投注额）')
            if {'代理名称', '有效投注额'}.issubset(agent.columns):
                top = agent.groupby('代理名称', as_index=False)['有效投注额'].sum().nlargest(10, '有效投注额').sort_values('有效投注额')
                fig = px.bar(top, y='代理名称', x='有效投注额', orientation='h', template=TEMPLATE,
                             color='有效投注额', color_continuous_scale='Blues')
                fig.update_layout(height=420, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            section_header('TOP 一级渠道（来源：原始字段【一级】）')
            if {'一级', '有效投注额'}.issubset(promo.columns):
                top = promo.groupby('一级', as_index=False)['有效投注额'].sum().nlargest(10, '有效投注额').sort_values('有效投注额')
                fig = px.bar(top, y='一级', x='有效投注额', orientation='h', template=TEMPLATE,
                             color='有效投注额', color_continuous_scale='teal')
                fig.update_layout(height=420, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')

    c3, c4 = st.columns(2)
    with c3:
        with st.container(border=True):
            section_header('代理分层散点')
            if {'代理名称', '有效投注额', '公司输赢'}.issubset(agent.columns):
                grp = agent.groupby('代理名称', as_index=False).agg({'有效投注额': 'sum', '公司输赢': 'sum'})
                fig = px.scatter(grp, x='有效投注额', y='公司输赢', hover_name='代理名称', template=TEMPLATE,
                                 color='公司输赢', color_continuous_scale='RdYlGn')
                fig.add_hline(y=0, line_dash='dot', line_color='rgba(150,170,210,0.5)')
                fig.update_layout(height=420, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')
    with c4:
        with st.container(border=True):
            section_header('渠道结构（Sunburst）')
            if {'一级', '二级', '三级', '四级', '有效投注额'}.issubset(promo.columns):
                tmp = promo.copy()
                tmp[['一级', '二级', '三级', '四级']] = tmp[['一级', '二级', '三级', '四级']].fillna('空值')
                fig = px.sunburst(tmp, path=['一级', '二级', '三级', '四级'], values='有效投注额', template=TEMPLATE)
                fig.update_layout(height=420)
                st.plotly_chart(fig, width='stretch')


def render_game_venue():
    venue = load_table('raw_game_report_venue')
    game = load_table('raw_game_analysis')
    hero('游戏与场馆', '查看原始字段【场馆名称】【游戏类型】【游戏名称】的规模与结果。', latest_imported_at(venue, game),
         basis='游戏报表(场馆)＋游戏分析（后台导出·每月上传）',
         detail=(
             '**分析范围**：各场馆与游戏类型的投注规模、有效投注、公司输赢。\n\n'
             '**数据来源（后台导出 → 上传）**：\n'
             '- 游戏报表(场馆)（报表中心→游戏报表(场馆)）\n'
             '- 游戏分析（报表中心→游戏分析，导出须选「日报」颗粒度）\n\n'
             '**更新方式**：手动上传。完整对照见「数据说明」页。'
         ))

    venue, start, end, month = date_range_picker(venue, '时间', 'gv', default_last_days=None)
    if month and '日期' in game.columns:
        game['日期'] = to_datetime_safe(game['日期'])
        game = game[game['日期'].dt.strftime('%Y-%m') == month].copy()
    elif '日期' in game.columns and start is not None and end is not None:
        game['日期'] = to_datetime_safe(game['日期'])
        game = game[(game['日期'] >= start) & (game['日期'] < end + pd.Timedelta(days=1))].copy()

    c1, c2 = st.columns(2)
    with c1:
        venue = apply_multiselect(venue, '场馆名称', '场馆名称', 'gv_venue')
    with c2:
        game = apply_multiselect(game, '游戏类型', '游戏类型', 'gv_type')

    notices = []
    if '站点名称' in venue.columns and safe_nunique(venue, '站点名称') <= 1:
        notices.append(FilterNotice('站点名称未设为主筛选', '当前样本中站点名称仅有单一值，不作为主筛选条件。'))
    add_info_box(notices)

    gv_winloss = safe_sum(venue, '公司输赢')
    cols = st.columns(4)
    show_metric(cols[0], '投注人数', fmt_num(safe_sum(venue, '投注人数')))
    show_metric(cols[1], '有效投注额', fmt_num(safe_sum(venue, '有效投注额')))
    show_metric(cols[2], '公司输赢', fmt_num(gv_winloss), tone=tone_by_sign(gv_winloss))
    show_metric(cols[3], '场馆数', fmt_num(safe_nunique(venue, '场馆名称')))

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('场馆有效投注额')
            if {'场馆名称', '有效投注额'}.issubset(venue.columns):
                grp = venue.groupby('场馆名称', as_index=False)['有效投注额'].sum().nlargest(15, '有效投注额').sort_values('有效投注额')
                fig = px.bar(grp, y='场馆名称', x='有效投注额', orientation='h', template=TEMPLATE,
                             color='有效投注额', color_continuous_scale='blues')
                fig.update_layout(height=430, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            section_header('游戏类型结构（Treemap）')
            if {'游戏类型', '场馆名称', '有效投注额'}.issubset(game.columns):
                tmp = game.groupby(['游戏类型', '场馆名称'], as_index=False)['有效投注额'].sum()
                fig = px.treemap(tmp, path=['游戏类型', '场馆名称'], values='有效投注额', template=TEMPLATE)
                fig.update_layout(height=430)
                st.plotly_chart(fig, width='stretch')

    with st.container(border=True):
        section_header('游戏明细（原始字段）')
        table_cols = [c for c in ['日期', '游戏类型', '场馆名称', '游戏名称', '投注人数', '有效投注额', '公司输赢'] if c in game.columns]
        if table_cols:
            st.dataframe(game[table_cols].sort_values(table_cols[0], ascending=False), width='stretch', hide_index=True)


def render_member_value():
    member = load_table('raw_member_report')
    top = load_table('raw_top_report')
    hero('会员价值', '查看会员结构、高价值用户集中度、VIP 分布与月快照状态。', latest_imported_at(member, top),
         basis='会员报表＋TOP报表（后台导出·每月上传）',
         detail=(
             '**分析范围**：会员结构、VIP 分布、ARPU、高价值用户集中度与月度留存。\n\n'
             '**数据来源（后台导出 → 上传）**：\n'
             '- 会员报表（报表中心→会员报表，须按注册时间、完整日期、全部页数导出）\n'
             '- TOP报表（报表中心→TOP报表）\n\n'
             '**计算口径**：会员身份＝会员账号＋代理（同名挂不同代理为不同人）；首存金额为一次性属性，统计前需去重。\n\n'
             '**更新方式**：手动上传。完整对照见「数据说明」页。'
         ))

    # 日期筛选统一用全站同款预设(全部/本月/上月/近7天/近30天/自订)，按「会员注册时间」筛
    if '注册时间' in member.columns:
        st.markdown('**按会员注册时间筛选**')
        member, _mv_s, _mv_e, _mv_m = date_range_picker(member, '注册时间', 'mv')

    top = get_snapshot_month(top)
    snap_options = [x for x in sorted(top['__snapshot_month__'].dropna().unique().tolist())]
    selected_snapshot = st.selectbox('TOP 快照月份', snap_options if snap_options else ['未提供快照字段'], key='mv_top_snapshot')
    if snap_options:
        top = top[top['__snapshot_month__'] == selected_snapshot].copy()

    member, notices, current = member_default_filters(member)
    # 统一会员去重键：账号+代理（同名挂不同代理=不同人），跨快照月也只算一次
    if '会员账号' in member.columns:
        if '代理' in member.columns:
            member['__member_key__'] = member['会员账号'].astype(str) + ' @ ' + member['代理'].astype(str)
        else:
            member['__member_key__'] = member['会员账号'].astype(str)
    add_info_box(notices)
    current_line = []
    for k, vals in current.items():
        if vals:
            current_line.append(f"{k}={' / '.join(vals)}")
    if current_line:
        st.caption('当前口径：' + '｜'.join(current_line))

    cols = st.columns(6)
    total_members = member_count(member)
    total_income = safe_sum(member, '公司收入')
    arpu_value = (total_income / total_members) if total_members else None
    # 首存金额是「一次性属性」(每月快照重复同一个值)，按行数会把同一人算 N 次→必须先按人去重再数
    member_unique = member.drop_duplicates('__member_key__') if '__member_key__' in member.columns else member
    first_dep_n = int((pd.to_numeric(member_unique['首存金额'], errors='coerce').fillna(0) > 0).sum()) if '首存金额' in member_unique.columns else 0
    show_metric(cols[0], '会员总数', fmt_num(total_members))
    show_metric(cols[1], '首存人数', fmt_num(first_dep_n))
    show_metric(cols[2], '总有效投注额', fmt_num(safe_sum(member, '有效投注额')))
    show_metric(cols[3], '公司收入', fmt_num(total_income), help_text=tooltip_text('公司收入'),
                tone=tone_by_sign(total_income))
    show_metric(cols[4], 'ARPU（用户均收）', fmt_num(arpu_value), help_text=tooltip_text('ARPU（用户均收）'),
                tone='accent')
    top_total = safe_sum(top, '有效投注额')
    top20_total = float(top.nlargest(20, '有效投注额')['有效投注额'].sum()) if {'有效投注额'}.issubset(top.columns) and not top.empty else 0
    top20_share = (top20_total / top_total) if top_total else None
    show_metric(cols[5], 'TOP20投注占比', fmt_pct(top20_share), help_text=tooltip_text('TOP20投注占比'))
    st.caption('口径：会员总数 / 首存人数 按「会员账号+代理」去重（每人算一次）；总有效投注额 / 公司收入 / ARPU '
               '为所选会员在数据库各月快照的**累计**值（投注额、公司收入是逐月数字，跨月相加=区间累计）。')

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('用户来源分布（原始字段）')
            if '用户来源' in member.columns and not member.empty:
                _md = member.drop_duplicates('__member_key__') if '__member_key__' in member.columns else member
                source_df = _md['用户来源'].fillna('空值').value_counts().reset_index()
                source_df.columns = ['用户来源', '会员数']
                fig = px.pie(source_df, names='用户来源', values='会员数', hole=0.58, template=TEMPLATE,
                             color_discrete_sequence=[PURPLE, BLUE, CYAN, GREEN, AMBER, RED])
                fig.update_traces(textinfo='percent', textfont_size=12,
                                  marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(
                    height=360,
                    legend=dict(orientation='h', y=-0.12),
                    annotations=[dict(text=f'会员<br><b>{fmt_num(int(source_df["会员数"].sum()))}</b>',
                                      x=0.5, y=0.5, showarrow=False,
                                      font=dict(size=14, color='#f0f5ff'))],
                )
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            section_header('VIP等级分布（原始字段）')
            if 'VIP等级' in member.columns and not member.empty:
                _md = member.drop_duplicates('__member_key__') if '__member_key__' in member.columns else member
                vip_df = _md['VIP等级'].fillna('空值').astype(str).value_counts().reset_index()
                vip_df.columns = ['VIP等级', '会员数']
                fig = px.bar(vip_df, x='VIP等级', y='会员数', template=TEMPLATE, color='会员数', color_continuous_scale='Purples')
                fig.update_layout(height=360, coloraxis_showscale=False, xaxis_title=None)
                st.plotly_chart(fig, width='stretch')

    section_header('ARPU（用户均收）按 VIP 等级 / 用户来源切片',
                   'ARPU = SUM(公司收入) ÷ DISTINCT(会员账号+代理)。分析各层级会员与来源渠道的人均价值。')
    c_arpu1, c_arpu2 = st.columns(2)
    with c_arpu1, st.container(border=True):
        if {'VIP等级', '公司收入', '会员账号'}.issubset(member.columns) and not member.empty:
            tmp_vip = member[['VIP等级', '公司收入', '__member_key__']].copy()
            tmp_vip['VIP等级'] = tmp_vip['VIP等级'].fillna('空值').astype(str)
            arpu_vip = tmp_vip.groupby('VIP等级', as_index=False).agg(
                公司收入=('公司收入', 'sum'),
                会员数=('__member_key__', 'nunique'),
            )
            arpu_vip['ARPU'] = arpu_vip.apply(
                lambda r: r['公司收入'] / r['会员数'] if r['会员数'] else 0, axis=1
            )
            arpu_vip = arpu_vip.sort_values('VIP等级')
            fig = px.bar(
                arpu_vip, x='VIP等级', y='ARPU', template=TEMPLATE,
                color='ARPU', color_continuous_scale='Tealgrn',
                hover_data={'公司收入': ':,.0f', '会员数': ':,.0f', 'ARPU': ':,.0f'},
                title='按 VIP 等级',
            )
            fig.update_layout(height=320, coloraxis_showscale=False)
            st.plotly_chart(fig, width='stretch')
        else:
            st.caption('当前数据缺少 VIP等级 或 公司收入 字段，跳过此切片。')
    with c_arpu2, st.container(border=True):
        if {'用户来源', '公司收入', '会员账号'}.issubset(member.columns) and not member.empty:
            tmp_src = member[['用户来源', '公司收入', '__member_key__']].copy()
            tmp_src['用户来源'] = tmp_src['用户来源'].fillna('空值').astype(str)
            arpu_src = tmp_src.groupby('用户来源', as_index=False).agg(
                公司收入=('公司收入', 'sum'),
                会员数=('__member_key__', 'nunique'),
            )
            arpu_src['ARPU'] = arpu_src.apply(
                lambda r: r['公司收入'] / r['会员数'] if r['会员数'] else 0, axis=1
            )
            arpu_src = arpu_src.sort_values('ARPU', ascending=False)
            fig = px.bar(
                arpu_src, x='用户来源', y='ARPU', template=TEMPLATE,
                color='ARPU', color_continuous_scale='Blues',
                hover_data={'公司收入': ':,.0f', '会员数': ':,.0f', 'ARPU': ':,.0f'},
                title='按用户来源',
            )
            fig.update_layout(height=320, coloraxis_showscale=False)
            st.plotly_chart(fig, width='stretch')
        else:
            st.caption('当前数据缺少 用户来源 或 公司收入 字段，跳过此切片。')

    section_header('月快照状态')
    member_snapshot_ready = has_member_snapshot(member)
    status_cols = st.columns(1)
    show_metric(status_cols[0], 'TOP快照月份', selected_snapshot if snap_options else '未提供')

    retention = compute_monthly_retention(member)
    if member_snapshot_ready and retention is not None and not retention.empty:
        cols = st.columns(3)
        last_row = retention.iloc[-1]
        show_metric(cols[0], '次月活跃留存率', fmt_pct(last_row['次月活跃留存率']), tone='accent')
        show_metric(cols[1], '次月存款留存率', fmt_pct(last_row['次月存款留存率']), tone='accent')
        show_metric(cols[2], '首存用户次月留存率', fmt_pct(last_row['首存用户次月留存率']), tone='accent')
        with st.container(border=True):
            fig = px.line(
                retention.melt(
                    id_vars=['月份', '次月'],
                    value_vars=['次月活跃留存率', '次月存款留存率', '首存用户次月留存率'],
                    var_name='指标',
                    value_name='值'
                ),
                x='月份', y='值', color='指标', template=TEMPLATE
            )
            fig.update_traces(line=dict(width=2.5, shape='spline', smoothing=0.6))
            fig.update_layout(height=360, hovermode='x unified', xaxis_title=None,
                              legend=dict(orientation='h', y=-0.18))
            st.plotly_chart(fig, width='stretch')
        st.caption(
            f'留存数据来源：raw_member_report（会员报表 × 多月快照）。'
            f'显示期间：{last_row["月份"]}→{last_row["次月"]}。'
            f'筛选条件已套用（见上方"当前口径"）。详细计算方式请展开下方「本页指标口径说明」。'
            f'⚠️ 若最新月份（{last_row["次月"]}）的快照是月中导入、还不是整月数据，'
            f'最后一个数据点的留存率会偏低，仅供参考，等整月导齐再看准。'
        )

    with st.expander('公司收入口径与月度校验', expanded=False):
        st.write('当前“公司收入”直接汇总会员报表原始字段【公司收入】，前端不做二次推导。')
        if '注册时间' in member.columns and '公司收入' in member.columns and not member.empty:
            tmp = member.copy()
            tmp['注册月份'] = tmp['注册时间'].dt.strftime('%Y-%m')
            check = tmp.groupby('注册月份', as_index=False).agg(
                会员数=('__member_key__', 'nunique'),
                有效投注额=('有效投注额', 'sum'),
                公司收入=('公司收入', 'sum'),
            )
            st.dataframe(check, width='stretch', hide_index=True)

    render_metric_explainer(['公司收入', 'ARPU（用户均收）', 'TOP20投注占比', '次月活跃留存率', '次月存款留存率', '首存用户次月留存率'])


def render_realtime():
    rt = load_table('raw_realtime_bet')
    hero('实时波动', '查看实时投注分时波动、热度分布与异常变化。', latest_imported_at(rt),
         basis='即时注单（后台导出·上传）',
         detail=(
             '**分析范围**：实时投注分时波动、DAU 近似、时段热度与异常监测。\n\n'
             '**数据来源**：即时注单（报表中心→即时注单）。\n\n'
             '**更新方式**：手动上传（为时点快照，按需更新）。完整对照见「数据说明」页。'
         ))
    rt = parse_realtime_time(rt)
    if '日期' not in rt.columns or rt['日期'].isna().all():
        st.warning('实时投注表当前无法从字段【时间】拆出可用日期。常见原始格式应类似 `2026-01-31 23~24`；若 BigQuery 中该字段已被改写，请先确认原始值。')
        return

    rt, start, end, month = date_range_picker(rt, '日期', 'rt', default_last_days=None)
    if '游戏类型' in rt.columns:
        rt = apply_multiselect(rt, '游戏类型', '游戏类型', 'rt_type', default_all=True)
    if rt.empty:
        st.warning('当前筛选条件下无数据。')
        return

    daily_bettors_series = (
        rt.groupby('日期')['时段投注人数'].sum()
        if '时段投注人数' in rt.columns else pd.Series(dtype=float)
    )
    avg_daily_bettors = float(daily_bettors_series.mean()) if not daily_bettors_series.empty else None

    cols = st.columns(5)
    show_metric(
        cols[0], '日投注人次（DAU 近似）', fmt_num(avg_daily_bettors),
        help_text=tooltip_text('日投注人次（DAU 近似）'),
    )
    show_metric(cols[1], '时段投注人数', fmt_num(safe_sum(rt, '时段投注人数')))
    show_metric(cols[2], '投注金额', fmt_num(safe_sum(rt, '投注金额')))
    show_metric(cols[3], '有效投注额', fmt_num(safe_sum(rt, '有效投注额')))
    rt_winloss = safe_sum(rt, '公司输赢')
    show_metric(cols[4], '公司输赢', fmt_num(rt_winloss), tone=tone_by_sign(rt_winloss))

    # ── 时段异常监测：最新一天各时段 vs 前 7 日同时段均值，偏离 ±20% 自动标示 ──
    if {'日期', '时段', '有效投注额'}.issubset(rt.columns):
        with st.container(border=True):
            section_header('时段异常监测', '最新一天各时段有效投注额，对比前 7 日同时段均值；下跌 ≥20% 标红、上涨 ≥20% 标绿。')
            slot_daily = rt.groupby(['日期', '时段'], as_index=False)['有效投注额'].sum()
            latest_d = slot_daily['日期'].max()
            base = slot_daily[(slot_daily['日期'] < latest_d) &
                              (slot_daily['日期'] >= latest_d - pd.Timedelta(days=7))]
            cur = slot_daily[slot_daily['日期'] == latest_d]
            if base.empty or cur.empty:
                st.caption('当前筛选范围内数据不足（最新一天之前需要至少一天数据才能对比）。')
            else:
                base_avg = base.groupby('时段', as_index=False)['有效投注额'].mean().rename(columns={'有效投注额': '前7日均值'})
                cmp_df = pd.merge(cur[['时段', '有效投注额']], base_avg, on='时段', how='inner').rename(columns={'有效投注额': '最新一天'})
                cmp_df['偏离%'] = (cmp_df['最新一天'] - cmp_df['前7日均值']) / cmp_df['前7日均值'].replace(0, pd.NA) * 100
                cmp_df = cmp_df.dropna(subset=['偏离%']).copy()
                cmp_df['_h'] = pd.to_numeric(cmp_df['时段'].astype(str).str.extract(r'(\d{1,2})')[0], errors='coerce')
                cmp_df = cmp_df.sort_values('_h')
                drops = cmp_df[cmp_df['偏离%'] <= -20]
                spikes = cmp_df[cmp_df['偏离%'] >= 20]
                if drops.empty and spikes.empty:
                    st.markdown(
                        status_badge(f'{latest_d:%m-%d} 各时段流水均在前 7 日均值 ±20% 内，未见异常', 'good'),
                        unsafe_allow_html=True,
                    )
                else:
                    badges = []
                    for _, r in drops.iterrows():
                        badges.append(status_badge(
                            f'{r["时段"]} 时段流水 ▼{abs(r["偏离%"]):.0f}%（{fmt_num(r["最新一天"])} / 均值 {fmt_num(r["前7日均值"])}）', 'bad'))
                    for _, r in spikes.iterrows():
                        badges.append(status_badge(f'{r["时段"]} 时段流水 ▲{r["偏离%"]:.0f}%', 'good'))
                    st.markdown(f'<div class="badge-row">{"".join(badges)}</div>', unsafe_allow_html=True)
                figd = go.Figure()
                figd.add_trace(go.Bar(
                    x=cmp_df['时段'], y=cmp_df['最新一天'], name=f'{latest_d:%m-%d}',
                    marker_color=[RED if v <= -20 else (GREEN if v >= 20 else BLUE) for v in cmp_df['偏离%']],
                ))
                figd.add_trace(go.Scatter(
                    x=cmp_df['时段'], y=cmp_df['前7日均值'], name='前 7 日均值', mode='lines',
                    line=dict(color=AMBER, width=2, dash='dot'),
                ))
                figd.update_layout(height=300, template=TEMPLATE, hovermode='x unified',
                                   legend=dict(orientation='h', y=-0.25), xaxis_title=None, yaxis_title='有效投注额')
                st.plotly_chart(figd, width='stretch')

    if not daily_bettors_series.empty:
        with st.container(border=True):
            section_header('日投注人次趋势（DAU 近似）',
                           '按日加总各时段「时段投注人数」，叠加 7 日移动平均看趋势。注意：会员可能跨时段重复计入，是 DAU 的近似。')
            daily_b = daily_bettors_series.reset_index()
            daily_b.columns = ['日期', '日投注人次']
            daily_b = daily_b.sort_values('日期')
            daily_b['7日均值'] = daily_b['日投注人次'].rolling(window=7, min_periods=1).mean()
            fig_dau = go.Figure()
            fig_dau.add_trace(go.Bar(
                x=daily_b['日期'], y=daily_b['日投注人次'],
                name='日投注人次', marker_color=CYAN, opacity=0.55,
            ))
            fig_dau.add_trace(go.Scatter(
                x=daily_b['日期'], y=daily_b['7日均值'],
                name='7 日均值', mode='lines',
                line=dict(color=AMBER, width=2.5, shape='spline', smoothing=0.6),
            ))
            fig_dau.update_layout(
                height=340, template=TEMPLATE, hovermode='x unified',
                legend=dict(orientation='h', y=-0.2),
                xaxis_title=None, yaxis_title='人次',
            )
            st.plotly_chart(fig_dau, width='stretch')

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('按日期趋势')
            daily = rt.groupby('日期', as_index=False).agg({'有效投注额': 'sum', '公司输赢': 'sum'})
            fig = px.line(daily.melt(id_vars='日期', var_name='指标', value_name='值'), x='日期', y='值', color='指标', template=TEMPLATE,
                          color_discrete_sequence=[CYAN, RED])
            fig.update_traces(line=dict(width=2.5, shape='spline', smoothing=0.6))
            fig.update_layout(height=380, hovermode='x unified', xaxis_title=None,
                              legend=dict(orientation='h', y=-0.18))
            st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            section_header('时段热度图')
            if '时段' in rt.columns and '游戏类型' in rt.columns:
                heat = rt.groupby(['游戏类型', '时段'], as_index=False)['有效投注额'].sum()
                if not heat.empty:
                    pivot = heat.pivot(index='游戏类型', columns='时段', values='有效投注额').fillna(0)

                    def _hour_key(val):
                        s = str(val)
                        m = re.match(r'(\d{1,2})', s)
                        return int(m.group(1)) if m else 99

                    ordered_cols = sorted(list(pivot.columns), key=_hour_key)
                    pivot = pivot.reindex(columns=ordered_cols)

                    fig = go.Figure(
                        data=go.Heatmap(
                            z=pivot.values,
                            x=list(pivot.columns),
                            y=list(pivot.index),
                            colorscale='YlOrRd',
                            hovertemplate='时段:%{x}<br>游戏类型:%{y}<br>有效投注额:%{z:,.0f}<extra></extra>',
                            colorbar=dict(title='金额', thickness=14),
                            xgap=1,
                            ygap=1,
                        )
                    )
                    fig.update_layout(
                        template=TEMPLATE,
                        height=450,
                        margin=dict(l=20, r=20, t=10, b=20),
                        xaxis=dict(title='时段', tickangle=-45, automargin=True),
                        yaxis=dict(title='游戏类型', automargin=True),
                    )
                    st.plotly_chart(fig, width='stretch')


def render_bonus_analysis():
    bonus = load_table('raw_bonus_report')
    hero('红利分析', '按活动名称、红利类型分析红利发放情况。', latest_imported_at(bonus),
         basis='红利记录（会员管理→VIP记录管理→红利记录·上传）',
         detail=(
             '**分析范围**：红利按活动名称、类型的发放分布、每日趋势与 TOP 领取会员。\n\n'
             '**数据来源**：红利记录（会员管理→VIP记录管理→红利记录，状态＝成功）。\n\n'
             '**计算口径**：区分「代理相关红利」与「真实会员活动红利」。\n\n'
             '**更新方式**：手动上传（按订单号去重，只补新订单）。完整对照见「数据说明」页。'
         ))

    if bonus.empty:
        st.warning('暂无红利数据')
        return

    # Date filter
    if '申请时间' in bonus.columns:
        bonus['申请时间'] = to_datetime_safe(bonus['申请时间'])
        bonus['日期'] = bonus['申请时间'].dt.date

    bonus, start, end, month = date_range_picker(bonus, '申请时间', 'bn', default_last_days=None)

    # 口径与「红利ROI / 代理质量」页一致：只算发放成功的红利（失败/驳回不是成本）
    if '状态' in bonus.columns:
        _before_n = len(bonus)
        bonus = bonus[bonus['状态'].astype(str).str.strip() == '成功'].copy()
        _excluded = _before_n - len(bonus)
        if _excluded:
            st.caption(f'已排除 {_excluded} 笔非「成功」状态的红利（与红利ROI页口径一致，只统计成功发放）。')

    # ── 代理相关红利分类 ──
    all_activity_names = []
    if '活动名称' in bonus.columns:
        all_activity_names = sorted(bonus['活动名称'].dropna().astype(str).unique().tolist())
        all_activity_names = [n for n in all_activity_names if n not in ('', 'nan', 'None')]

    if 'agent_bonus_activities' not in st.session_state:
        st.session_state['agent_bonus_activities'] = []

    # 安全过滤：确保 default 中的值都在当前 options 内
    safe_default = [n for n in st.session_state['agent_bonus_activities'] if n in all_activity_names]

    section_header('代理相关红利分类', '选择属于代理相关的活动名称，用于拆分红利归属并调整经营总览净收入。')
    selected_agent_activities = st.multiselect(
        '选择代理相关红利活动',
        options=all_activity_names,
        default=safe_default,
        key='agent_bonus_activities_select',
        help='选中的活动将被归类为"代理相关红利"，其余为"真实会员活动红利"',
    )
    st.session_state['agent_bonus_activities'] = selected_agent_activities

    # 计算拆分金额
    if selected_agent_activities and '活动名称' in bonus.columns:
        agent_mask = bonus['活动名称'].astype(str).isin(selected_agent_activities)
        agent_bonus_total = float(bonus.loc[agent_mask, '红利金额'].sum()) if '红利金额' in bonus.columns else 0.0
    else:
        agent_bonus_total = 0.0

    # KPIs
    cols = st.columns(4)
    total_amount = safe_sum(bonus, '红利金额')
    total_count = len(bonus)
    unique_members = member_count(bonus)
    avg_per_member = total_amount / unique_members if unique_members else 0
    show_metric(cols[0], '红利总金额', fmt_num(total_amount), tone='warn',
                help_text='成本项，重点监控')
    show_metric(cols[1], '红利笔数', fmt_num(total_count))
    show_metric(cols[2], '领取会员数', fmt_num(unique_members))
    show_metric(cols[3], '人均红利', fmt_num(avg_per_member), tone='accent')

    # 代理 vs 真实会员红利拆分
    if selected_agent_activities:
        real_member_bonus_total = total_amount - agent_bonus_total
        split_cols = st.columns(3)
        show_metric(split_cols[0], '代理相关红利', fmt_num(agent_bonus_total),
                    help_text=f'来自 {len(selected_agent_activities)} 个代理相关活动', tone='warn')
        show_metric(split_cols[1], '真实会员活动红利', fmt_num(real_member_bonus_total),
                    help_text='红利总金额 - 代理相关红利', tone='good')
        if total_amount:
            show_metric(split_cols[2], '代理红利占比', fmt_pct(agent_bonus_total / total_amount),
                        help_text='代理相关红利 / 红利总金额', tone='warn')

    # By activity name
    with st.container(border=True):
        section_header('按活动名称统计')
        if '活动名称' in bonus.columns and '红利金额' in bonus.columns:
            by_activity = bonus.groupby('活动名称', as_index=False).agg(
                红利金额=('红利金额', 'sum'),
                笔数=('红利金额', 'count')
            ).sort_values('红利金额', ascending=False)

            fig = px.bar(by_activity.head(20), x='活动名称', y='红利金额', text='笔数',
                         template=TEMPLATE, color='红利金额', color_continuous_scale='Oranges')
            fig.update_layout(height=420, xaxis_tickangle=-45, coloraxis_showscale=False, xaxis_title=None)
            fig.update_traces(textposition='outside', textfont_size=10)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(by_activity, use_container_width=True, hide_index=True)

    c1, c2 = st.columns(2)
    with c1:
        # By type
        with st.container(border=True):
            section_header('按红利类型分布')
            if '红利类型' in bonus.columns:
                by_type = bonus.groupby('红利类型', as_index=False).agg(
                    红利金额=('红利金额', 'sum'),
                    笔数=('红利金额', 'count')
                ).sort_values('红利金额', ascending=False)
                fig = px.pie(by_type, names='红利类型', values='红利金额', hole=0.58,
                             template=TEMPLATE, color_discrete_sequence=[AMBER, BLUE, PURPLE, CYAN, GREEN, RED])
                fig.update_traces(textinfo='percent', textfont_size=12,
                                  marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(
                    height=360,
                    legend=dict(orientation='h', y=-0.12),
                    annotations=[dict(text=f'红利<br><b>{fmt_num(float(by_type["红利金额"].sum()))}</b>',
                                      x=0.5, y=0.5, showarrow=False,
                                      font=dict(size=14, color='#f0f5ff'))],
                )
                st.plotly_chart(fig, use_container_width=True)

    with c2:
        # By date trend
        with st.container(border=True):
            section_header('每日红利趋势')
            if '日期' in bonus.columns:
                by_date = bonus.groupby('日期', as_index=False).agg(红利金额=('红利金额', 'sum'))
                by_date = by_date.sort_values('日期')
                fig = px.area(by_date, x='日期', y='红利金额', template=TEMPLATE)
                fig.update_traces(line_color=AMBER, line_shape='spline', line_smoothing=0.6,
                                  fillcolor='rgba(251,191,36,0.14)')
                fig.update_layout(height=360, hovermode='x unified', xaxis_title=None)
                st.plotly_chart(fig, use_container_width=True)

    # TOP members
    with st.container(border=True):
        section_header('TOP 领取会员')
        if '会员账号' in bonus.columns and '红利金额' in bonus.columns:
            top_members = bonus.groupby('会员账号', as_index=False).agg(
                红利金额=('红利金额', 'sum'),
                笔数=('红利金额', 'count'),
                会员等级=('会员等级', 'first')
            ).sort_values('红利金额', ascending=False).head(20)
            st.dataframe(top_members, use_container_width=True, hide_index=True)

    with st.expander('本页指标口径说明', expanded=False):
        st.write('数据来源：raw_bonus_report（红利记录导出）')
        st.write('活动名称：红利标题为空时，自动取用申请备注')
        st.write('仅统计状态=成功的记录')


def source_note(origin_html: str):
    """页面顶部「数据来源」说明 — 给接手的人看：后台从哪导、怎么导、多久一次。
    （延续 Miru『我走了也要有人能运作』原则：每页都该标清楚来源，不靠记忆/不靠问人。）"""
    st.markdown(
        '<div style="background:rgba(56,189,248,0.08);border-left:3px solid #38bdf8;'
        'border-radius:6px;padding:0.55rem 0.85rem;margin:0.1rem 0 0.9rem;'
        'font-size:0.86rem;line-height:1.65;color:#b6c5e1;">'
        '📂 <b style="color:#e2e8f0;">数据来源（怎么更新）</b>：' + origin_html +
        '</div>', unsafe_allow_html=True)


def render_agent_commission():
    """代理佣金深度分析 — 基于代理佣金单线版 + 团队版两张 BigQuery 表"""
    hero(
        '代理佣金深度分析',
        '基于后台的代理佣金报表（单线版 + 团队版），分析代理产值、赤字分布、手续费净成本与代理层级结构。',
        '',
        basis='代理佣金单线＋团队（代理管理→佣金管理→发放佣金）＋代理结算月报（客服主管提供）',
        detail=(
            '**分析范围**：代理产值与赤字、手续费净成本、代理层级结构，及客服主管月报的结算明细与退成。\n\n'
            '**数据来源**：\n'
            '- 代理佣金（代理管理→佣金管理→发放佣金，切单线／团队，设「佣金月份」，导出 Csv）\n'
            '- 代理结算月报（客服主管提供「X月代理帐.xlsx」，现可拖数据上传页自助入库）\n\n'
            '**更新方式**：佣金按佣金月份手动上传；结算月报拖上传页选月份入库。市代月度见独立「市代月度结算」页。完整对照见「数据说明」页。'
        )
    )
    source_note(
        '后台 <b>代理管理 → 佣金管理 → 发放佣金</b>，切换「<b>单线佣金 / 团队佣金</b>」，'
        '把「佣金月份」选到目标月份 → 按「<b>导出 Csv</b>」下载（单线、团队各导一份）。每月导一次。'
        '<br>导出后**直接拖进面板顶部「🗂 数据上传」页**即可入库（自动识别单线/团队、按月刷新当月、保留其他月，密码 zip 也行）。'
        '<br>对应 BigQuery 表：<code>raw_agent_commission_single</code>（单线）/ '
        '<code>raw_agent_commission_team</code>（团队）。'
    )

    try:
        single = load_table('raw_agent_commission_single')
        team = load_table('raw_agent_commission_team')
    except Exception as e:
        st.error(f'代理佣金数据尚未导入 BigQuery。错误：{e}')
        st.info('运行 `import_agent_commission.py` 导入最新月份数据后再查看此页面。')
        return

    if single.empty:
        st.warning('尚无代理佣金单线版数据')
        return

    months = []
    if '佣金月份' in single.columns:
        months = sorted(single['佣金月份'].dropna().astype(str).unique().tolist())
    month_label = months[-1] if months else '未知'

    # 月份选择器（始终显示，即使只有一个月）
    month_options = months if months else ['未知']
    col_m1, col_m2 = st.columns([1, 4])
    with col_m1:
        sel_month = st.selectbox(
            '📅 佣金月份',
            month_options,
            index=len(month_options) - 1,
            key='ac_month',
            disabled=(len(month_options) <= 1 and month_options[0] == '未知'),
        )
    with col_m2:
        st.markdown(
            f'<div style="padding-top:1.8rem;color:#b6c5e1;">'
            f'当前查看：<b>{sel_month}</b> | 本页共有数据月份：<b>{", ".join(months) if months else "尚未导入"}</b>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.caption('📊 本页上半部（手续费 / 产值赤字 / 层级 / 代理明细）全部用「**代理佣金 单线·团队**」数据，'
               '跟着上面的「佣金月份」走。下半部「客服主管月报」是平哥手动结算（另一份、自己的月份）。'
               '「市代月度」已独立成 🅲 代理/渠道 →「市代月度结算」页。')

    # 月份过滤
    if months and sel_month in months:
        single = single[single['佣金月份'].astype(str) == sel_month].copy()
        team = team[team['佣金月份'].astype(str) == sel_month].copy()
    month_label = sel_month

    st.info(
        f'📌 **重要提示**：此报表的「佣金」字段仅记录系统派发部分；红利派发、兑台等其他派发方式不在内。'
        f'所以"佣金"列常为 0，但"冲正后净输赢"、"上月结余（赤字）"、"手续费基数"等字段是真实数据。'
    )

    # ── 顶部 KPI ─────────────────────────────
    total_agents = len(single)
    active_agents = int((single['冲正后净输赢'].fillna(0) != 0).sum()) if '冲正后净输赢' in single.columns else 0
    total_net = safe_sum(single, '冲正后净输赢')
    deficit_mask = (single['上月结余'].fillna(0) < 0) if '上月结余' in single.columns else None
    deficit_agents = int(deficit_mask.sum()) if deficit_mask is not None else 0
    total_deficit = float(single.loc[deficit_mask, '上月结余'].sum()) if deficit_mask is not None else 0.0
    dep_base = safe_sum(single, '存款手续费基数')
    wd_base = safe_sum(single, '提款手续费基数')
    total_base = dep_base + wd_base
    charged_to_agent = total_base * 0.015
    paid_to_yabo = total_base * 0.016
    platform_extra = total_base * 0.001

    cols = st.columns(4)
    show_metric(cols[0], '代理总数', fmt_num(total_agents), f'活跃 {active_agents}')
    show_metric(cols[1], '冲正后净输赢（总）', fmt_num(total_net), help_text='正值=平台盈利，负值=平台亏损',
                tone=tone_by_sign(total_net))
    show_metric(cols[2], '赤字代理数', fmt_num(deficit_agents), f'合计欠 {fmt_num(total_deficit)}',
                help_text='上月结余 < 0 的代理数量',
                tone='bad' if deficit_agents else None, delta_tone='down' if deficit_agents else 'flat')
    show_metric(cols[3], '手续费净差额（0.1%）', fmt_num(platform_extra),
                f'代理端 {fmt_num(charged_to_agent)} / 系统方 {fmt_num(paid_to_yabo)}',
                help_text='代理手续费 1.5%，系统方手续费 1.6%，差额 0.1%', tone='warn')

    # ── 手续费结构图 ─────────────────────────────
    section_header('手续费结构分析')
    c1, c2 = st.columns([1, 1])
    with c1:
        with st.container(border=True):
            fee_df = pd.DataFrame({
                '类别': ['存款手续费基数', '提款手续费基数'],
                '金额(万)': [dep_base / 10000, wd_base / 10000],
            })
            fig = px.bar(fee_df, x='类别', y='金额(万)', template=TEMPLATE,
                         color='类别', color_discrete_sequence=[BLUE, CYAN])
            fig.update_traces(hovertemplate='%{x}<br>%{y:,.2f} 万<extra></extra>')
            fig.update_yaxes(title_text='金额（万元）')
            fig.update_layout(height=320, showlegend=False, title='存提款手续费基数', xaxis_title=None)
            st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            compare_df = pd.DataFrame({
                '方向': ['代理端 1.5%', '系统方 1.6%', '差额 0.1%'],
                '金额(万)': [charged_to_agent / 10000, paid_to_yabo / 10000, platform_extra / 10000],
                'color': ['代理端', '系统方', '差额'],
            })
            fig = px.bar(compare_df, x='方向', y='金额(万)', template=TEMPLATE,
                         color='color', color_discrete_map={'代理端': GREEN, '系统方': BLUE, '差额': AMBER})
            fig.update_traces(hovertemplate='%{x}<br>%{y:,.2f} 万<extra></extra>')
            fig.update_yaxes(title_text='金额（万元）')
            fig.update_layout(height=320, showlegend=False, title='手续费结构对比', xaxis_title=None)
            st.plotly_chart(fig, width='stretch')

    st.markdown(
        f'<div class="hero-card" style="margin-top:0.5rem;">'
        f'<div class="hero-title">{month_label} 手续费结构摘要</div>'
        f'<div class="hero-subtitle">'
        f'存提款总基数 <b>{fmt_num(total_base)}</b>；代理端手续费 <b>{fmt_num(charged_to_agent)}</b>，'
        f'系统方手续费 <b>{fmt_num(paid_to_yabo)}</b>，<b style="color:{AMBER}">差额 {fmt_num(platform_extra)}</b>。'
        f'按此速率年化约 <b>{fmt_num(platform_extra * 12)}</b>。'
        f'</div></div>',
        unsafe_allow_html=True
    )

    # ── 代理产值 & 赤字排行 ─────────────────────────────
    section_header('代理产值 vs 赤字排行')
    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            st.markdown('**Top 20 产值代理（按冲正后净输赢，正值=平台盈利）**')
            if {'代理账号', '冲正后净输赢'}.issubset(single.columns):
                pos = single[single['冲正后净输赢'].fillna(0) > 0].copy()
                top = pos.nlargest(20, '冲正后净输赢')[['代理账号', '冲正后净输赢', '下级人数']].sort_values('冲正后净输赢')
                top['净输赢(万)'] = top['冲正后净输赢'] / 10000
                fig = px.bar(top, y='代理账号', x='净输赢(万)', orientation='h', template=TEMPLATE,
                             color='净输赢(万)', color_continuous_scale='Greens',
                             hover_data=['下级人数'])
                fig.update_traces(hovertemplate='代理：%{y}<br>净输赢：%{x:,.2f} 万<extra></extra>')
                fig.update_xaxes(title_text='净输赢（万元）')
                fig.update_layout(height=520, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            st.markdown('**Top 20 赤字代理（按上月结余欠款）**')
            if {'代理账号', '上月结余'}.issubset(single.columns):
                neg = single[single['上月结余'].fillna(0) < 0].copy()
                worst = neg.nsmallest(20, '上月结余')[['代理账号', '上月结余', '冲正后净输赢']].sort_values('上月结余', ascending=False)
                worst['结余(万)'] = worst['上月结余'] / 10000
                fig = px.bar(worst, y='代理账号', x='结余(万)', orientation='h', template=TEMPLATE,
                             color='结余(万)', color_continuous_scale='Reds_r',
                             hover_data=['冲正后净输赢'])
                fig.update_traces(hovertemplate='代理：%{y}<br>上月结余：%{x:,.2f} 万<extra></extra>')
                fig.update_xaxes(title_text='上月结余（万元）')
                fig.update_layout(height=520, coloraxis_showscale=False)
                st.plotly_chart(fig, width='stretch')

    # ── 赤字分布直方图 ─────────────────────────────
    section_header('赤字代理分布（大客户专账方案的关键数据）')
    c1, c2 = st.columns([2, 1])
    with c1:
        with st.container(border=True):
            if '上月结余' in single.columns:
                neg = single[single['上月结余'].fillna(0) < 0].copy()
                neg['赤字金额(万)'] = -neg['上月结余'] / 10000
                fig = px.histogram(neg, x='赤字金额(万)', nbins=30, template=TEMPLATE,
                                   color_discrete_sequence=[RED])
                fig.update_traces(hovertemplate='赤字区间：%{x} 万<br>代理数：%{y}<extra></extra>')
                fig.update_xaxes(title_text='赤字金额（万元）')
                fig.update_yaxes(title_text='代理数')
                fig.update_layout(height=380, title='赤字金额分布直方图')
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            if '上月结余' in single.columns:
                neg = single[single['上月结余'].fillna(0) < 0].copy()
                bins = [
                    ('< 1w', neg[neg['上月结余'] > -10000]),
                    ('1w-5w', neg[(neg['上月结余'] <= -10000) & (neg['上月结余'] > -50000)]),
                    ('5w-10w', neg[(neg['上月结余'] <= -50000) & (neg['上月结余'] > -100000)]),
                    ('10w+', neg[neg['上月结余'] <= -100000]),
                ]
                bin_df = pd.DataFrame([
                    {'区间': name, '代理数': len(d), '赤字合计': float(d['上月结余'].sum())}
                    for name, d in bins
                ])
                st.markdown('**按赤字规模分区**')
                for _, row in bin_df.iterrows():
                    st.markdown(
                        f'<div style="padding:0.4rem 0.6rem;background:var(--bad-soft);'
                        f'border-left:3px solid var(--bad);margin-bottom:0.4rem;border-radius:4px;">'
                        f'<b>{row["区间"]}</b>：{int(row["代理数"])} 个代理，合计 {fmt_num(row["赤字合计"])}'
                        f'</div>',
                        unsafe_allow_html=True
                    )

    # ── 赤字代理完整名单（可排序 / 搜索 / 下载）─────────────────────────────
    section_header('赤字代理完整名单',
                   '上月结余 < 0 的全部代理。点列头可排序，悬停右上角可搜索 / 下载 CSV。')
    if {'代理账号', '上月结余'}.issubset(single.columns):
        neg = single[single['上月结余'].fillna(0) < 0].copy()
        show_cols = [c for c in ['代理账号', '上月结余', '冲正后净输赢', '下级人数', '活跃人数']
                     if c in neg.columns]
        detail = neg[show_cols].sort_values('上月结余').reset_index(drop=True)
        st.dataframe(detail, width='stretch', hide_index=True)
        st.caption(
            f'共 {len(detail)} 个赤字代理（{month_label}）。默认按欠款最多排在最前；'
            f'点任意列头可改排序，悬停表格右上角的图标可搜索或下载。')
    else:
        st.info('此月份数据缺「代理账号」或「上月结余」字段，无法列出名单。')

    # ── 主副线结构（团队版数据） ─────────────────────────────
    if not team.empty and '线别' in team.columns:
        section_header('代理层级结构（来自团队版数据）')
        c1, c2 = st.columns(2)
        with c1:
            with st.container(border=True):
                line_stats = team.groupby('线别').agg(
                    代理数=('代理账号', 'count'),
                    总净输赢=('冲正后净输赢', lambda x: float(x.sum())),
                    总赤字=('上月结余', lambda x: float(x[x < 0].sum())),
                ).reset_index()
                fig = px.bar(line_stats, x='线别', y='代理数', template=TEMPLATE,
                             color='线别', color_discrete_map={'主线': BLUE, '副线': CYAN},
                             title='主线 vs 副线 代理数量')
                fig.update_layout(height=320, xaxis_title=None)
                st.plotly_chart(fig, width='stretch')
        with c2:
            with st.container(border=True):
                line_stats['总净输赢(万)'] = line_stats['总净输赢'] / 10000
                fig = px.bar(line_stats, x='线别', y='总净输赢(万)', template=TEMPLATE,
                             color='线别', color_discrete_map={'主线': BLUE, '副线': CYAN},
                             title='主线 vs 副线 净输赢对比')
                fig.update_traces(hovertemplate='%{x}<br>净输赢：%{y:,.2f} 万<extra></extra>')
                fig.update_yaxes(title_text='净输赢（万元）')
                fig.update_layout(height=320, xaxis_title=None)
                st.plotly_chart(fig, width='stretch')

        if '团队名称' in team.columns:
            with st.container(border=True):
                section_header('Top 10 团队（按净输赢合计）')
                team_agg = team.groupby('团队名称', as_index=False).agg(
                    代理数=('代理账号', 'count'),
                    净输赢=('冲正后净输赢', 'sum'),
                    赤字=('上月结余', lambda x: float(x[x < 0].sum())),
                ).sort_values('净输赢', ascending=False).head(10)
                st.dataframe(team_agg, width='stretch', hide_index=True)

    # ── 代理活跃度分析 ─────────────────────────────
    section_header('代理活跃度分布')
    c1, c2, c3 = st.columns(3)
    with c1:
        with st.container(border=True):
            if '活跃人数' in single.columns:
                active_dist = single['活跃人数'].fillna(0).astype(int)
                bins = [
                    ('0 人（无活跃）', int((active_dist == 0).sum())),
                    ('1-5 人', int(((active_dist >= 1) & (active_dist <= 5)).sum())),
                    ('6-20 人', int(((active_dist >= 6) & (active_dist <= 20)).sum())),
                    ('21-100 人', int(((active_dist >= 21) & (active_dist <= 100)).sum())),
                    ('100+ 人', int((active_dist > 100).sum())),
                ]
                bin_df = pd.DataFrame(bins, columns=['活跃会员区间', '代理数'])
                fig = px.bar(bin_df, x='活跃会员区间', y='代理数', template=TEMPLATE,
                             color='活跃会员区间', color_discrete_sequence=px.colors.sequential.Tealgrn)
                fig.update_layout(height=320, showlegend=False, title='按名下活跃会员数分档', xaxis_title=None)
                st.plotly_chart(fig, width='stretch')
    with c2:
        with st.container(border=True):
            if '代理类型' in single.columns:
                type_dist = single['代理类型'].value_counts().reset_index()
                type_dist.columns = ['代理类型', '代理数']
                fig = px.pie(type_dist, names='代理类型', values='代理数', template=TEMPLATE,
                             hole=0.55, title='代理类型构成')
                fig.update_traces(textinfo='percent', marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(height=320)
                st.plotly_chart(fig, width='stretch')
    with c3:
        with st.container(border=True):
            if '是否在团队' in single.columns:
                in_team = single['是否在团队'].astype(str).value_counts().reset_index()
                in_team.columns = ['是否在团队', '代理数']
                fig = px.pie(in_team, names='是否在团队', values='代理数', template=TEMPLATE,
                             hole=0.55, title='是否在团队（散户识别）')
                fig.update_traces(textinfo='percent', marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(height=320)
                st.plotly_chart(fig, width='stretch')

    # ── 明细表 ─────────────────────────────
    section_header('代理明细（可筛选、可下载）')
    with st.expander('筛选器', expanded=False):
        f1, f2, f3 = st.columns(3)
        df_filter = single.copy()
        with f1:
            if '代理类型' in df_filter.columns:
                df_filter = apply_multiselect(df_filter, '代理类型', '代理类型', 'ac_type')
        with f2:
            if '是否在团队' in df_filter.columns:
                df_filter = apply_multiselect(df_filter, '是否在团队', '是否在团队', 'ac_in_team')
        with f3:
            kw = st.text_input('搜索代理账号或备注', key='ac_kw')
            if kw:
                mask = pd.Series(False, index=df_filter.index)
                for col in ['代理账号', '备注']:
                    if col in df_filter.columns:
                        mask = mask | df_filter[col].astype(str).str.contains(kw, case=False, na=False)
                df_filter = df_filter[mask]

    show_cols = [c for c in [
        '代理账号', '代理类型', '上级账号', '下级人数', '活跃人数', '存款金额', '提款金额',
        '冲正后净输赢', '上月结余', '存款手续费基数', '提款手续费基数', '佣金比例',
        '是否在团队', '是否为主线', '发展人', '备注',
    ] if c in df_filter.columns]

    display_df = df_filter[show_cols].copy()
    if '冲正后净输赢' in display_df.columns:
        display_df = display_df.sort_values('冲正后净输赢', ascending=False)
    st.dataframe(display_df, width='stretch', hide_index=True, height=400)

    csv = display_df.to_csv(index=False).encode('utf-8-sig')
    st.download_button('📥 下载当前筛选结果（CSV）', csv,
                       file_name=f'agent_commission_{month_label}.csv',
                       mime='text/csv')

    # ── 客服主管月报（代理结算手动汇整）─────────────────
    st.markdown('---')
    section_header('客服主管月报 · 代理结算明细',
                   '由客服主管平哥手动汇整，按 8 种适用待遇分类。含累计挂帐、本月发放、退成明细等系统报表外的细节。')

    try:
        settle_summary = load_table('raw_agent_settlement_summary')
        settle_detail = load_table('raw_agent_settlement_detail')
    except Exception as e:
        st.info(f'尚无代理结算手动月报数据。请运行 `import_agent_settlement.py <代理帐.xlsx> <YYYY-MM>` 导入。错误：{e}')
    else:
        if settle_summary.empty and settle_detail.empty:
            st.info('代理结算手动月报：暂无数据。')
        else:
            settle_months = sorted(set(
                settle_summary['月份'].dropna().astype(str).unique().tolist() +
                settle_detail['月份'].dropna().astype(str).unique().tolist()
            ))
            st.caption(f'📌 数据来源：客服主管月报（手动汇整 xlsx 导入），目前涵盖月份：{", ".join(settle_months) if settle_months else "无"}')
            sel_set_month = st.selectbox(
                '📅 代理结算月份',
                settle_months if settle_months else ['无'],
                index=(len(settle_months) - 1) if settle_months else 0,
                key='settle_month',
            )
            sub_sum = settle_summary[settle_summary['月份'].astype(str) == sel_set_month].copy() if not settle_summary.empty else pd.DataFrame()
            sub_det = settle_detail[settle_detail['月份'].astype(str) == sel_set_month].copy() if not settle_detail.empty else pd.DataFrame()

            # 摘要 KPI
            if not sub_sum.empty:
                summary_map = dict(zip(
                    sub_sum['项目'].astype(str).str.replace('⭐️', '', regex=False).str.strip(),
                    sub_sum['金额']
                ))
                ks = st.columns(4)
                show_metric(ks[0], '累计挂帐金额', fmt_num(summary_map.get('累计挂帐金额')),
                            help_text='截至本月底之累计代理挂帐余额（含历史滚结）', tone='warn')
                show_metric(ks[1], '本月新增挂帐', fmt_num(summary_map.get('本月新增挂帐')), tone='warn')
                show_metric(ks[2], '红利佣金派发', fmt_num(summary_map.get('红利佣金派发')))
                show_metric(ks[3], '总计发放', fmt_num(summary_map.get('总计发放')),
                            help_text='本月平台向代理实际派发之佣金 + 红利合计（红利佣金派发 + 佣金派发）',
                            tone='accent')

            # 按 8 种待遇 拆分
            if not sub_det.empty and '适用待遇' in sub_det.columns:
                section_header('按适用待遇分类（笔数 + 实际佣金）')
                by_treatment = sub_det.groupby('适用待遇', as_index=False).agg(
                    笔数=('实际佣金', 'count'),
                    实际佣金=('实际佣金', lambda x: float(pd.to_numeric(x, errors='coerce').fillna(0).sum())),
                ).sort_values('实际佣金', ascending=False)
                c_t1, c_t2 = st.columns([1.2, 1])
                with c_t1:
                    with st.container(border=True):
                        fig = px.bar(by_treatment, x='适用待遇', y='实际佣金',
                                     text='实际佣金', template=TEMPLATE,
                                     color='实际佣金', color_continuous_scale='Tealgrn',
                                     hover_data={'笔数': True, '实际佣金': ':,.0f'})
                        fig.update_layout(height=360, coloraxis_showscale=False, xaxis_tickangle=-15, xaxis_title=None)
                        fig.update_traces(texttemplate='%{text:,.0f}', textposition='outside')
                        st.plotly_chart(fig, width='stretch')
                with c_t2:
                    with st.container(border=True):
                        st.dataframe(by_treatment, width='stretch', hide_index=True)

                # 退成 明细
                rebate = sub_det[sub_det['适用待遇'].astype(str).str.contains('退成', na=False)].copy()
                if not rebate.empty:
                    section_header(f'退成（介绍人佣金分成）明细 — {len(rebate)} 笔',
                                   '详细的引荐关系 / 比例 / 实际派发金额。退成口径见下方「本页指标口径说明」。')
                    rebate_disp = rebate[['名称', '总代账号', '业绩总计', '比例', '实际佣金']].copy()
                    rebate_disp = rebate_disp.sort_values('实际佣金', ascending=False)
                    st.dataframe(
                        rebate_disp, width='stretch', hide_index=True,
                        column_config={
                            '业绩总计': st.column_config.NumberColumn(format='%.2f'),
                            '比例': st.column_config.NumberColumn(format='%.2%'),
                            '实际佣金': st.column_config.NumberColumn(format='%.0f'),
                        },
                    )

            render_metric_explainer(['退成（介绍人佣金分成）'])

    # 市代月度结算已独立成「市代月度结算」页（🅲 代理 / 渠道），本页不再重复。

    with st.expander('本页指标口径说明', expanded=False):
        st.markdown('''
- **数据来源**：`raw_agent_commission_single`（单线版）+ `raw_agent_commission_team`（团队版）+ `raw_agent_settlement_summary` / `raw_agent_settlement_detail`（客服主管手动月报）
- **冲正后净输赢**：代理名下所有会员的净盈亏，**正值=平台盈利，负值=平台亏损**
- **上月结余**：从上月滚下来的赤字/盈余，**负值=代理端尚有赤字待抵扣**
- **手续费 0.1% 净支出**：存款/提款手续费基数 × (1.6% 付系统方 − 1.5% 收代理) 的结构性差额
- **佣金字段限制**：此报表的「佣金」仅含系统派发部分，**不含红利派发/兑台等其他渠道**
- **大客户专账方案**：赤字分布数据直接用于校准 v2 方案的门槛金额
- **客服主管月报**：客服主管平哥手动汇整之 xlsx，含累计挂帐、退成、新代理特殊安排等系统外的细节；月度更新，需手动 import
''')



# ── 投注分析(月度注单明细,每月一张 raw_bet_detail_YYYY_MM 表) ──────────────────────────

def render_bet_analysis():
    hero(
        '投注分析',
        '基于当月全部投注注单，按场馆、游戏类型、VIP 等级三大维度分析。所有汇总通过 BigQuery 聚合查询，不下载明细。',
        '',
        basis='注单明细（按月表）；来源待确认，目前库内仅含 4 月',
        detail=(
            '**分析范围**：按场馆、游戏类型、VIP 等级三维度的注单聚合分析。\n\n'
            '**数据来源**：注单明细按月表（raw_bet_detail_YYYY_MM），经 BigQuery 聚合，不下载明细。\n\n'
            '**注意**：此报表来源尚未确认、目前库内仅含 2026 年 4 月。需补全请先确认后台导出位置。完整对照见「数据说明」页。'
        )
    )

    # 动态发现已导入的月份表(raw_bet_detail_YYYY_MM);新月份注单进 BigQuery 后自动出现在选单
    BET_TABLE = 'raw_bet_detail_2026_04'  # 兜底默认
    bet_month_label = '2026-04'
    try:
        # 直接走 client：query_bq 的 normalize 会把 table_name 字串列强转数字成 NaN
        _tbls = get_client().query(
            f"SELECT table_name FROM `{BQ_PREFIX}.INFORMATION_SCHEMA.TABLES` "
            "WHERE table_name LIKE 'raw_bet_detail_%' ORDER BY table_name"
        ).to_dataframe()
        _month_map = {}
        for _t in _tbls['table_name'].astype(str).tolist():
            _m = re.match(r'raw_bet_detail_(\d{4})_(\d{2})$', _t)
            if _m:
                _month_map[f'{_m.group(1)}-{_m.group(2)}'] = _t
        if _month_map:
            _opts = sorted(_month_map.keys())
            mc1, mc2 = st.columns([1, 4])
            with mc1:
                bet_month_label = st.selectbox('📅 注单月份', _opts, index=len(_opts) - 1, key='bet_month')
            with mc2:
                st.markdown(
                    f'<div style="padding-top:1.8rem;color:#9fc1d9;">已导入月份：<b>{", ".join(_opts)}</b>'
                    '（新月份注单导入 BigQuery 后自动出现在选单）</div>',
                    unsafe_allow_html=True,
                )
            BET_TABLE = _month_map[bet_month_label]
            # 切换月份时清掉旧筛选状态,避免选项与新表不一致报错
            if st.session_state.get('_bet_month_prev') != bet_month_label:
                for _k in ('bet_type', 'bet_vip', 'bet_venue', 'bet_date'):
                    st.session_state.pop(_k, None)
                st.session_state['_bet_month_prev'] = bet_month_label
    except Exception:
        pass

    # 表是否存在 — 时间戳用 SQL 端 FORMAT 成字串,避免 normalize_dataframe 把 timestamp 转成数字再 lossy 还原
    try:
        meta = query_bq(
            "SELECT COUNT(*) AS n, "
            "FORMAT_TIMESTAMP('%Y-%m-%d', MIN(`下注时间`)) AS min_t, "
            "FORMAT_TIMESTAMP('%Y-%m-%d', MAX(`下注时间`)) AS max_t "
            f"FROM `{BQ_PREFIX}.{BET_TABLE}`"
        )
    except Exception as e:
        st.error(f'投注详情数据尚未导入 BigQuery（{BET_TABLE}）。错误：{e}')
        return

    if meta.empty or int(meta.iloc[0]['n']) == 0:
        st.warning('暂无投注数据')
        return

    total_bets = int(meta.iloc[0]['n'])
    min_t = meta.iloc[0]['min_t']
    max_t = meta.iloc[0]['max_t']

    range_str = ''
    if min_t and max_t and str(min_t) != 'nan' and str(max_t) != 'nan':
        range_str = f"{min_t} ~ {max_t}"
    st.info(
        f'📌 **数据范围：{range_str}** | 📌 **注单数：{total_bets:,}** | '
        f'📌 **盈亏口径**：负数 = 平台赢；正数 = 平台输（玩家赢）'
    )

    # ── 筛选条件 ──────────────────────────────────────────
    section_header('筛选条件', '下方所有图表 / 表格 都依据此筛选条件计算（直接走 BigQuery WHERE，速度不变）')

    # 取候选清单
    opt_q = (
        f"SELECT `场馆类型` AS venue_type, `VIP等级` AS vip_label, `场馆名称` AS venue "
        f"FROM `{BQ_PREFIX}.{BET_TABLE}` GROUP BY `场馆类型`, `VIP等级`, `场馆名称`"
    )
    opt_df = query_bq(opt_q)

    type_opts = sorted([t for t in opt_df['venue_type'].dropna().unique() if t])
    vip_opts = sorted(
        [v for v in opt_df['vip_label'].dropna().unique() if v],
        key=lambda x: int(x) if str(x).isdigit() else 99
    )
    venue_opts = sorted([v for v in opt_df['venue'].dropna().unique() if v])

    # session_state 默认值初始化(只在第一次)
    if 'bet_type' not in st.session_state:
        st.session_state['bet_type'] = list(type_opts)
    if 'bet_vip' not in st.session_state:
        st.session_state['bet_vip'] = list(vip_opts)
    if 'bet_venue' not in st.session_state:
        st.session_state['bet_venue'] = []

    # 全选 / 全清 快捷按钮(需先于 multiselect 渲染,以便修改 session_state)
    bc1, bc2, bc3, bc4, bc5, bc6 = st.columns([1, 1, 1, 1, 1, 1])
    with bc1:
        if st.button('类型 全选', use_container_width=True, key='bet_type_all_btn'):
            st.session_state['bet_type'] = list(type_opts)
            st.rerun()
    with bc2:
        if st.button('类型 全清', use_container_width=True, key='bet_type_clear_btn'):
            st.session_state['bet_type'] = []
            st.rerun()
    with bc3:
        if st.button('VIP 全选', use_container_width=True, key='bet_vip_all_btn'):
            st.session_state['bet_vip'] = list(vip_opts)
            st.rerun()
    with bc4:
        if st.button('VIP 全清', use_container_width=True, key='bet_vip_clear_btn'):
            st.session_state['bet_vip'] = []
            st.rerun()
    with bc5:
        if st.button('场馆 全选', use_container_width=True, key='bet_venue_all_btn'):
            st.session_state['bet_venue'] = list(venue_opts)
            st.rerun()
    with bc6:
        if st.button('场馆 全清', use_container_width=True, key='bet_venue_clear_btn'):
            st.session_state['bet_venue'] = []
            st.rerun()

    fc1, fc2, fc3, fc4 = st.columns([1.2, 1.2, 1.4, 1.4])
    with fc1:
        # 日期 range 取自 BQ 的 min/max
        try:
            mn = pd.to_datetime(min_t).date() if min_t else None
            mx = pd.to_datetime(max_t).date() if max_t else None
        except Exception:
            mn, mx = None, None
        date_range = st.date_input(
            '下注日期范围', value=(mn, mx) if mn and mx else None,
            min_value=mn, max_value=mx, key='bet_date'
        )
    with fc2:
        sel_types = st.multiselect('场馆类型', type_opts, key='bet_type')
    with fc3:
        sel_vips = st.multiselect('VIP 等级', vip_opts, key='bet_vip')
    with fc4:
        sel_venues = st.multiselect('场馆名称（不选 = 全部）', venue_opts, key='bet_venue')

    # 组 WHERE
    where_parts = []
    if isinstance(date_range, tuple) and len(date_range) == 2 and all(date_range):
        where_parts.append(
            f"DATE(`下注时间`) BETWEEN DATE('{date_range[0]}') AND DATE('{date_range[1]}')"
        )
    def _sql_in(vals):  # 转义单引号，避免值含 ' 时断句/注入
        return ', '.join("'" + str(v).replace("'", "''") + "'" for v in vals)
    if sel_types and len(sel_types) < len(type_opts):
        where_parts.append(f"`场馆类型` IN ({_sql_in(sel_types)})")
    if sel_vips and len(sel_vips) < len(vip_opts):
        where_parts.append(f"`VIP等级` IN ({_sql_in(sel_vips)})")
    if sel_venues:
        where_parts.append(f"`场馆名称` IN ({_sql_in(sel_venues)})")

    where_clause = (' WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    if where_parts:
        # 重新算 KPI(命中筛选范围) 并提示
        sub_q = f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}"
        sub_n = int(query_bq(sub_q).iloc[0]['n'])
        st.success(f'当前筛选命中 **{sub_n:,}** 笔注单（占全月 {sub_n/total_bets*100:.1f}%）')

    # ── 总览 KPI ───────────────────────────────────────────
    kpi_q = f"""
        SELECT
          COUNT(*) AS bets,
          SUM(`有效投注`) AS valid_bet,
          SUM(`下注金额`) AS bet_amount,
          SUM(`盈亏`) AS pnl,
          COUNT(DISTINCT `会员账号`) AS players
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
    """
    kpi = query_bq(kpi_q)
    bets_n = int(kpi.iloc[0]['bets'] or 0)  # 随筛选变动（与下面 4 张卡同口径）
    valid_bet = float(kpi.iloc[0]['valid_bet'] or 0)
    pnl = float(kpi.iloc[0]['pnl'] or 0)
    players = int(kpi.iloc[0]['players'] or 0)
    hold_pct = (-pnl / valid_bet * 100) if valid_bet else 0.0

    cols = st.columns(5)
    show_metric(cols[0], '注单总数', fmt_num(bets_n),
                help_text=None if where_parts else '全月（未加筛选）')
    show_metric(cols[1], '有效投注总额', fmt_num(valid_bet))
    show_metric(cols[2], '平台净盈亏',
                fmt_num(-pnl), delta='平台赢' if pnl < 0 else '平台输',
                tone=tone_by_sign(-pnl), delta_tone='up' if pnl < 0 else 'down')
    show_metric(cols[3], 'Hold % (净盈/有效投注)', f'{hold_pct:.2f}%', tone='accent')
    show_metric(cols[4], '参与玩家数', fmt_num(players))

    # ── 模块 1 — 场馆分析 ───────────────────────────────────
    section_header('模块 1 ─ 场馆分析', '按 场馆名称 切片，看流水分布、Hold%、玩家数')

    venue_q = f"""
        SELECT
          `场馆名称` AS venue,
          `场馆类型` AS venue_type,
          COUNT(*) AS bets,
          SUM(`有效投注`) AS valid_bet,
          SUM(`盈亏`) AS pnl,
          COUNT(DISTINCT `会员账号`) AS players,
          SAFE_DIVIDE(SUM(`有效投注`), COUNT(*)) AS avg_bet
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
        GROUP BY `场馆名称`, `场馆类型`
        ORDER BY valid_bet DESC
    """
    venue_df = query_bq(venue_q)
    venue_df = venue_df.rename(columns={
        'venue': '场馆', 'venue_type': '类型', 'bets': '注单数',
        'valid_bet': '有效投注', 'pnl': '盈亏', 'players': '玩家数', 'avg_bet': '单注均值'
    })
    venue_df['平台净盈亏'] = -venue_df['盈亏']
    venue_df['Hold%'] = (venue_df['平台净盈亏'] / venue_df['有效投注'] * 100).round(3)
    venue_df['流水占比%'] = (venue_df['有效投注'] / venue_df['有效投注'].sum() * 100).round(2)

    c1, c2 = st.columns([1.4, 1])
    with c1:
        with st.container(border=True):
            section_header('Top 15 场馆 — 有效投注 与 Hold%')
            top15 = venue_df.head(15).copy()
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=top15['场馆'],
                y=top15['有效投注'],
                name='有效投注',
                marker_color=BLUE,
                yaxis='y',
                hovertemplate='%{x}<br>有效投注 %{y:,.0f}<extra></extra>'
            ))
            fig.add_trace(go.Scatter(
                x=top15['场馆'],
                y=top15['Hold%'],
                name='Hold%',
                mode='lines+markers',
                marker_color=AMBER,
                yaxis='y2',
                hovertemplate='%{x}<br>Hold %{y:.2f}%<extra></extra>'
            ))
            fig.update_layout(
                template=TEMPLATE,
                height=440,
                xaxis_tickangle=-30,
                yaxis=dict(title='有效投注'),
                yaxis2=dict(title='Hold %', overlaying='y', side='right'),
                legend=dict(orientation='h', y=1.05, x=0.5, xanchor='center'),
                margin=dict(t=40, b=10),
            )
            st.plotly_chart(fig, use_container_width=True)

    with c2:
        with st.container(border=True):
            section_header('类型流水占比')
            type_pie = venue_df.groupby('类型', as_index=False)['有效投注'].sum().sort_values('有效投注', ascending=False)
            fig = px.pie(
                type_pie, names='类型', values='有效投注', hole=0.58,
                template=TEMPLATE,
                color_discrete_sequence=[BLUE, GREEN, PURPLE, CYAN, AMBER, RED, '#7AB7FF'],
            )
            fig.update_traces(textinfo='label+percent', textfont_size=12,
                              marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
            fig.update_layout(height=440, legend=dict(orientation='h', y=-0.05))
            st.plotly_chart(fig, use_container_width=True)

    with st.container(border=True):
        section_header('全部场馆明细')
        show_df = venue_df[['场馆', '类型', '注单数', '有效投注', '平台净盈亏', 'Hold%', '玩家数', '单注均值', '流水占比%']].copy()
        st.dataframe(show_df, use_container_width=True, hide_index=True,
                     column_config={
                         '注单数': st.column_config.NumberColumn(format='%d'),
                         '有效投注': st.column_config.NumberColumn(format='%.0f'),
                         '平台净盈亏': st.column_config.NumberColumn(format='%.0f'),
                         '玩家数': st.column_config.NumberColumn(format='%d'),
                         '单注均值': st.column_config.NumberColumn(format='%.2f'),
                     })

    # ── 模块 2 — 游戏类型分析 ───────────────────────────────
    section_header('模块 2 ─ 游戏类型分析', '体育 / 真人 / 电子 / 棋牌 / 彩票 / 电竞 / 捕鱼 大类汇总')

    type_q = f"""
        SELECT
          `场馆类型` AS venue_type,
          COUNT(*) AS bets,
          SUM(`有效投注`) AS valid_bet,
          SUM(`盈亏`) AS pnl,
          COUNT(DISTINCT `会员账号`) AS players
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
        GROUP BY `场馆类型`
        ORDER BY valid_bet DESC
    """
    type_df = query_bq(type_q)
    type_df = type_df.rename(columns={
        'venue_type': '类型', 'bets': '注单数', 'valid_bet': '有效投注',
        'pnl': '盈亏', 'players': '玩家数'
    })
    type_df['平台净盈亏'] = -type_df['盈亏']
    type_df['Hold%'] = (type_df['平台净盈亏'] / type_df['有效投注'] * 100).round(3)
    type_df['人均流水'] = (type_df['有效投注'] / type_df['玩家数']).round(0)
    type_df['流水占比%'] = (type_df['有效投注'] / type_df['有效投注'].sum() * 100).round(2)

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('类型 × Hold% 对比')
            fig = px.bar(
                type_df, x='类型', y='Hold%', text='Hold%',
                color='Hold%', color_continuous_scale='RdYlGn_r',
                template=TEMPLATE,
            )
            fig.update_traces(texttemplate='%{text:.2f}%', textposition='outside')
            fig.update_layout(height=380, coloraxis_showscale=False, xaxis_title=None)
            st.plotly_chart(fig, use_container_width=True)

    with c2:
        with st.container(border=True):
            section_header('类型 × 人均流水(玩家深度)')
            fig = px.bar(
                type_df, x='类型', y='人均流水', text='玩家数',
                color='人均流水', color_continuous_scale='Blues',
                template=TEMPLATE,
            )
            fig.update_traces(texttemplate='%{text:,d} 人', textposition='outside')
            fig.update_layout(height=380, coloraxis_showscale=False, xaxis_title=None)
            st.plotly_chart(fig, use_container_width=True)

    st.dataframe(
        type_df[['类型', '注单数', '有效投注', '平台净盈亏', 'Hold%', '玩家数', '人均流水', '流水占比%']],
        use_container_width=True, hide_index=True,
        column_config={
            '注单数': st.column_config.NumberColumn(format='%d'),
            '有效投注': st.column_config.NumberColumn(format='%.0f'),
            '平台净盈亏': st.column_config.NumberColumn(format='%.0f'),
            '玩家数': st.column_config.NumberColumn(format='%d'),
            '人均流水': st.column_config.NumberColumn(format='%.0f'),
        }
    )

    # ── 模块 3 — VIP 等级分析 ───────────────────────────────
    section_header('模块 3 ─ VIP 等级分析', 'VIP 0~10 流水分布、活动让利 cap 校准、Top 大户')

    vip_q = f"""
        SELECT
          SAFE_CAST(`VIP等级` AS INT64) AS vip_n,
          `VIP等级` AS vip_label,
          COUNT(*) AS bets,
          SUM(`有效投注`) AS valid_bet,
          SUM(`盈亏`) AS pnl,
          COUNT(DISTINCT `会员账号`) AS players
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
        GROUP BY `VIP等级`
        ORDER BY vip_n
    """
    vip_df = query_bq(vip_q)
    vip_df = vip_df.rename(columns={
        'vip_label': 'VIP', 'bets': '注单数', 'valid_bet': '有效投注',
        'pnl': '盈亏', 'players': '玩家数'
    })
    vip_df['平台净盈亏'] = -vip_df['盈亏']
    vip_df['Hold%'] = (vip_df['平台净盈亏'] / vip_df['有效投注'] * 100).round(3)
    vip_df['人均流水'] = (vip_df['有效投注'] / vip_df['玩家数']).round(0)
    vip_df['流水占比%'] = (vip_df['有效投注'] / vip_df['有效投注'].sum() * 100).round(2)

    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            section_header('VIP × 流水占比 + 玩家数')
            fig = make_subplots(specs=[[{'secondary_y': True}]])
            fig.add_trace(
                go.Bar(x=vip_df['VIP'].astype(str), y=vip_df['有效投注'], name='有效投注', marker_color=BLUE),
                secondary_y=False
            )
            fig.add_trace(
                go.Scatter(x=vip_df['VIP'].astype(str), y=vip_df['玩家数'], name='玩家数',
                           mode='lines+markers', marker_color=AMBER),
                secondary_y=True
            )
            fig.update_xaxes(title='VIP 等级')
            fig.update_yaxes(title='有效投注', secondary_y=False)
            fig.update_yaxes(title='玩家数', secondary_y=True)
            fig.update_layout(template=TEMPLATE, height=400,
                              legend=dict(orientation='h', y=1.05, x=0.5, xanchor='center'))
            st.plotly_chart(fig, use_container_width=True)

    with c2:
        with st.container(border=True):
            section_header('VIP × Hold% 走势')
            fig = px.line(
                vip_df, x='VIP', y='Hold%', markers=True, text='Hold%',
                template=TEMPLATE,
            )
            fig.update_traces(line_color=GREEN, marker_size=10, texttemplate='%{text:.2f}%', textposition='top center')
            fig.add_hline(y=0, line_color='rgba(150,170,210,0.5)', line_dash='dash',
                          annotation_text='平台不输不赢 (Hold=0)', annotation_position='right')
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)

    st.dataframe(
        vip_df[['VIP', '注单数', '有效投注', '平台净盈亏', 'Hold%', '玩家数', '人均流水', '流水占比%']],
        use_container_width=True, hide_index=True,
        column_config={
            '注单数': st.column_config.NumberColumn(format='%d'),
            '有效投注': st.column_config.NumberColumn(format='%.0f'),
            '平台净盈亏': st.column_config.NumberColumn(format='%.0f'),
            '玩家数': st.column_config.NumberColumn(format='%d'),
            '人均流水': st.column_config.NumberColumn(format='%.0f'),
        }
    )

    # ── VIP × 类型 交叉(老板让利 cap 校准用) ─────────────
    section_header('VIP × 场馆类型 流水交叉表',
                   '让利 cap 校准:体育 0.5% / 真人 0.3% / 老虎机 0.4% — 这张表 看 高 VIP 主要玩什么场馆')

    cross_q = f"""
        SELECT
          SAFE_CAST(`VIP等级` AS INT64) AS vip_n,
          `VIP等级` AS vip_label,
          `场馆类型` AS venue_type,
          SUM(`有效投注`) AS valid_bet
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
        GROUP BY `VIP等级`, `场馆类型`
        ORDER BY vip_n, venue_type
    """
    cross_df = query_bq(cross_q)
    cross_df = cross_df.rename(columns={'vip_label': 'VIP', 'venue_type': '类型', 'valid_bet': '有效投注'})
    cross_pivot = cross_df.pivot_table(
        index='VIP', columns='类型', values='有效投注', aggfunc='sum', fill_value=0
    )
    # row total 排序 by VIP 数字
    cross_pivot.index = pd.Categorical(
        cross_pivot.index,
        categories=sorted(cross_pivot.index, key=lambda x: int(x) if str(x).isdigit() else 99)
    )
    cross_pivot = cross_pivot.sort_index()

    with st.container(border=True):
        fig = px.imshow(
            cross_pivot.values,
            x=cross_pivot.columns.tolist(),
            y=[f'VIP{v}' for v in cross_pivot.index.tolist()],
            color_continuous_scale='Blues',
            aspect='auto',
            text_auto='.2s',
        )
        fig.update_layout(
            template=TEMPLATE, height=460,
            title='VIP × 场馆类型 — 有效投注热力图',
        )
        st.plotly_chart(fig, use_container_width=True)
    # pandas 2.1+ DataFrame.applymap deprecated → 改 .map(DataFrame.map 是 2.1+ 新增)
    try:
        cross_display = cross_pivot.map(lambda v: f'{v:,.0f}')
    except (AttributeError, TypeError):
        cross_display = cross_pivot.applymap(lambda v: f'{v:,.0f}')
    st.dataframe(cross_display, use_container_width=True)

    # ── Top 100 大户 ─────────────────────────────────────
    section_header(f'Top 100 大户(按 {bet_month_label} 有效投注)')
    top_q = f"""
        SELECT
          `会员账号` AS member,
          `VIP等级` AS vip_label,
          COUNT(*) AS bets,
          SUM(`有效投注`) AS valid_bet,
          SUM(`盈亏`) AS pnl,
          COUNT(DISTINCT `场馆名称`) AS venues
        FROM `{BQ_PREFIX}.{BET_TABLE}`{where_clause}
        GROUP BY `会员账号`, `VIP等级`
        ORDER BY valid_bet DESC
        LIMIT 100
    """
    top_df = query_bq(top_q)
    top_df = top_df.rename(columns={
        'member': '会员', 'vip_label': 'VIP', 'bets': '注单数',
        'valid_bet': '有效投注', 'pnl': '盈亏', 'venues': '场馆数'
    })
    top_df['平台净盈亏'] = -top_df['盈亏']
    top_df['Hold%'] = (top_df['平台净盈亏'] / top_df['有效投注'] * 100).round(2)

    # Top 大户 流水集中度
    top_total = top_df['有效投注'].sum()
    pct_concentration = top_total / valid_bet * 100 if valid_bet else 0
    st.markdown(
        f'**Top 100 大户 合计 {fmt_num(top_total)} 流水 占全平台 {pct_concentration:.1f}%** '
        f'(全平台 {fmt_num(valid_bet)})'
    )

    st.dataframe(
        top_df[['会员', 'VIP', '注单数', '有效投注', '平台净盈亏', 'Hold%', '场馆数']],
        use_container_width=True, hide_index=True,
        column_config={
            '注单数': st.column_config.NumberColumn(format='%d'),
            '有效投注': st.column_config.NumberColumn(format='%.0f'),
            '平台净盈亏': st.column_config.NumberColumn(format='%.0f'),
        }
    )

    # ── 口径说明 ──────────────────────────────────────────
    with st.expander('本页指标口径说明', expanded=False):
        st.markdown(f'''
- **数据来源**：`{BQ_PREFIX}.{BET_TABLE}`({bet_month_label} 全月投注注单, {total_bets:,} 笔)
- **月份选单**：自动扫描 BigQuery 中的 `raw_bet_detail_YYYY_MM` 表,新月份注单导入后即出现,无需改代码
- **盈亏口径**:原始 `盈亏` 列 是从 玩家视角 看(玩家赢=正,玩家输=负)。 面板「平台净盈亏」 = − `盈亏` (平台赢=正)
- **Hold %** = 平台净盈亏 / 有效投注 × 100
- **有效投注** ≠ 下注金额(扣 走盘 / 和局退本金 / 不计入活动 玩法)
- **VIP** = 该笔注单产生时玩家的 VIP 等级(月内升降级会按当时记录)
- **VIP × 类型 交叉表** 直接对照 老板 5/5 让利 cap:体育 0.5% / 真人 0.3% / 老虎机 0.4% — 看 高 VIP 主玩什么 → 让利成本最大重心在哪
- **Top 100 大户**:全平台当月总流水的集中度,集中度高表示「少数大户撑起大多数流水」
''')


def render_bonus_roi_agent_quality():
    """红利 ROI & 代理质量 — 6/4 新增 (Miru 决策面板:砍 / 升活动 + 代理筛选)"""
    bonus = load_table('raw_bonus_report')
    agent = load_table('raw_agent_report')

    hero('红利 ROI & 代理质量', '红利成本效率与代理质量分析，支持活动调整与代理管理决策。',
         latest_imported_at(bonus, agent),
         basis='红利记录＋代理报表（后台导出·每月上传）',
         detail=(
             '**分析范围**：红利成本效率（ROI）与代理质量（拉新、首存转化、红利依赖、亏损识别）。\n\n'
             '**数据来源（后台导出 → 上传）**：\n'
             '- 红利记录（会员管理→VIP记录管理→红利记录）\n'
             '- 代理报表（报表中心→代理报表）\n\n'
             '**更新方式**：手动上传。完整对照见「数据说明」页。'
         ))

    if bonus.empty or agent.empty:
        st.warning('暂无 红利 或 代理 数据')
        return

    # 准备数据
    bonus['申请时间'] = to_datetime_safe(bonus['申请时间'])
    bonus['日期'] = bonus['申请时间'].dt.date
    if '状态' in bonus.columns:
        bonus_succ = bonus[bonus['状态'].astype(str) == '成功'].copy()
    else:
        bonus_succ = bonus.copy()
    agent['日期'] = to_datetime_safe(agent['日期']).dt.date

    tabs = st.tabs(['🎯 红利 ROI', '👥 代理质量'])

    # ━━━━━━━━ Tab A: 红利 ROI ━━━━━━━━
    with tabs[0]:
        b_filt, b_start, b_end, _ = date_range_picker(bonus_succ, '申请时间', 'bn_roi', default_last_days=30)
        if b_filt.empty:
            st.info('该范围无红利数据')
        else:
            # KPI
            total_amt = safe_sum(b_filt, '红利金额')
            total_cnt = len(b_filt)
            unique_mem = member_count(b_filt)
            avg_per_mem = total_amt / unique_mem if unique_mem else 0

            # 同日期范围 agent 公司输赢
            if b_start and b_end:
                a_range = agent[(agent['日期'] >= b_start.date()) & (agent['日期'] <= b_end.date())]
            else:
                a_range = agent
            co_winloss = safe_sum(a_range, '公司输赢')
            bonus_share = (total_amt / abs(co_winloss) * 100) if co_winloss else 0
            co_income = safe_sum(a_range, '公司收入')

            c1, c2, c3, c4 = st.columns(4)
            show_metric(c1, '红利总成本', fmt_num(total_amt), tone='warn')
            show_metric(c2, '红利笔数', fmt_num(total_cnt))
            show_metric(c3, '涉及会员数', fmt_num(unique_mem))
            show_metric(c4, '人均红利', fmt_num(avg_per_mem), tone='accent')

            c5, c6 = st.columns(2)
            show_metric(c5, '同期公司输赢', fmt_num(co_winloss),
                        help_text='代理报表口径，正值=平台盈利，负值=平台亏损',
                        tone=tone_by_sign(co_winloss))
            show_metric(c6, '同期公司收入(净)', fmt_num(co_income),
                        help_text='扣除红利/返水/佣金后的净收入',
                        tone=tone_by_sign(co_income))

            # 时间趋势
            with st.container(border=True):
                section_header('每日红利成本 vs 公司收入', '红利成本与净收入的相对趋势。')
                daily_bn = b_filt.groupby('日期', as_index=False)['红利金额'].sum().rename(columns={'红利金额': '红利成本'})
                daily_co = a_range.groupby('日期', as_index=False)['公司收入'].sum().rename(columns={'公司收入': '公司收入(净)'})
                merged = pd.merge(daily_bn, daily_co, on='日期', how='outer').fillna(0).sort_values('日期')
                if not merged.empty:
                    fig = make_subplots(specs=[[{'secondary_y': True}]])
                    fig.add_trace(go.Bar(x=merged['日期'], y=merged['红利成本'], name='红利成本',
                                  marker_color=RED, opacity=0.65), secondary_y=False)
                    fig.add_trace(go.Scatter(x=merged['日期'], y=merged['公司收入(净)'], name='公司收入(净)',
                                  line=dict(color=CYAN, width=2.5, shape='spline', smoothing=0.6),
                                  mode='lines+markers'), secondary_y=True)
                    fig.update_layout(height=400, hovermode='x unified', xaxis_title=None, template=TEMPLATE,
                                      margin=dict(l=40, r=40, t=30, b=40))
                    fig.update_yaxes(title_text='红利成本 (元)', secondary_y=False)
                    fig.update_yaxes(title_text='公司收入 (元)', secondary_y=True)
                    st.plotly_chart(fig, use_container_width=True)

            # 红利类型分布
            if '红利类型' in b_filt.columns:
                section_header('红利类型分布', '按类型分组,看各类成本占比')
                type_grp = b_filt.groupby('红利类型').agg(
                    笔数=('红利金额', 'count'),
                    总金额=('红利金额', 'sum'),
                    会员数=('会员账号', 'nunique')
                ).sort_values('总金额', ascending=False).reset_index()
                type_grp['总金额占比'] = (type_grp['总金额'] / total_amt * 100).round(2).astype(str) + '%'
                type_grp['总金额'] = type_grp['总金额'].round(0)
                st.dataframe(type_grp, use_container_width=True, hide_index=True)

            # 红利标题 Top 30
            if '红利标题' in b_filt.columns:
                section_header('红利标题 Top 30 排行', '按总成本排序，用于识别高成本、低回报的活动。')
                title_grp = b_filt.groupby('红利标题').agg(
                    笔数=('红利金额', 'count'),
                    总金额=('红利金额', 'sum'),
                    平均金额=('红利金额', 'mean'),
                    涉及会员=('会员账号', 'nunique'),
                ).sort_values('总金额', ascending=False).head(30).reset_index()
                title_grp['总金额'] = title_grp['总金额'].round(0)
                title_grp['平均金额'] = title_grp['平均金额'].round(2)
                title_grp['人均红利'] = (title_grp['总金额'] / title_grp['涉及会员'].replace(0, 1)).round(0)
                st.dataframe(title_grp, use_container_width=True, hide_index=True,
                             column_config={
                                 '总金额': st.column_config.NumberColumn(format='%d'),
                                 '平均金额': st.column_config.NumberColumn(format='%.2f'),
                             })

            # 高频领取监测
            section_header('高频领取监测', '单日领取 ≥3 笔的会员（按累计红利降序），可作为流水门槛与风控名单的参考。')
            if '会员账号' in b_filt.columns:
                mem_daily = b_filt.groupby(['会员账号', '日期']).agg(
                    单日笔数=('红利金额', 'count'),
                    单日总额=('红利金额', 'sum')
                ).reset_index()
                multi = mem_daily[mem_daily['单日笔数'] >= 3]
                if multi.empty:
                    st.info('该范围内无单日 >=3 笔的会员')
                else:
                    multi_mem = multi.groupby('会员账号').agg(
                        高频天数=('日期', 'nunique'),
                        累积笔数=('单日笔数', 'sum'),
                        累积红利=('单日总额', 'sum'),
                        最高单日笔数=('单日笔数', 'max'),
                        最高单日金额=('单日总额', 'max')
                    ).sort_values('累积红利', ascending=False).head(30).reset_index()
                    multi_mem['累积红利'] = multi_mem['累积红利'].round(0)
                    multi_mem['最高单日金额'] = multi_mem['最高单日金额'].round(0)
                    st.dataframe(multi_mem, use_container_width=True, hide_index=True)

            # 流水门槛分析
            if '是否需要流水' in b_filt.columns and '流水倍数' in b_filt.columns:
                section_header('流水门槛分析', '识别无流水要求或低流水倍数的红利，评估套利风险敞口。')
                wr_grp = b_filt.groupby(['是否需要流水', '流水倍数']).agg(
                    笔数=('红利金额', 'count'),
                    总金额=('红利金额', 'sum'),
                    平均金额=('红利金额', 'mean')
                ).reset_index().sort_values('总金额', ascending=False).head(25)
                wr_grp['总金额'] = wr_grp['总金额'].round(0)
                wr_grp['平均金额'] = wr_grp['平均金额'].round(2)
                st.dataframe(wr_grp, use_container_width=True, hide_index=True)

            # VIP 等级 × 红利成本
            if '会员等级' in b_filt.columns:
                section_header('VIP 等级 × 红利成本', '各 VIP 等级的红利成本分布。')
                vip_grp = b_filt.groupby('会员等级').agg(
                    笔数=('红利金额', 'count'),
                    总金额=('红利金额', 'sum'),
                    会员数=('会员账号', 'nunique')
                ).reset_index().sort_values('总金额', ascending=False)
                vip_grp['人均红利'] = (vip_grp['总金额'] / vip_grp['会员数'].replace(0, 1)).round(0)
                vip_grp['总金额'] = vip_grp['总金额'].round(0)
                with st.container(border=True):
                    fig = px.bar(vip_grp, x='会员等级', y='总金额',
                                 text='总金额', hover_data=['笔数', '会员数', '人均红利'],
                                 color='总金额', color_continuous_scale='Reds')
                    fig.update_layout(height=380, showlegend=False, margin=dict(l=40, r=20, t=20, b=40), xaxis_title=None)
                    fig.update_traces(texttemplate='%{text:,.0f}', textposition='outside')
                    st.plotly_chart(fig, use_container_width=True)
                st.dataframe(vip_grp, use_container_width=True, hide_index=True)

    # ━━━━━━━━ Tab B: 代理质量 ━━━━━━━━
    with tabs[1]:
        a_filt, a_start, a_end, _ = date_range_picker(agent, '日期', 'ag_q', default_last_days=30)
        if a_filt.empty:
            st.info('该范围无代理数据')
        else:
            # KPI
            total_agents = safe_nunique(a_filt, '代理编号')
            agents_with_signup = a_filt[a_filt['注册人数'] > 0]['代理编号'].nunique() if '注册人数' in a_filt.columns else 0
            total_bonus = safe_sum(a_filt, '红利')
            total_co_income = safe_sum(a_filt, '公司收入')
            total_signup = safe_sum(a_filt, '注册人数')
            total_ftd = safe_sum(a_filt, '首存人数')

            c1, c2, c3, c4 = st.columns(4)
            show_metric(c1, '活跃代理数', fmt_num(total_agents))
            show_metric(c2, '有拉新代理', fmt_num(agents_with_signup))
            show_metric(c3, '累积拉新', fmt_num(total_signup))
            show_metric(c4, '累积首存', fmt_num(total_ftd))

            c5, c6, c7 = st.columns(3)
            show_metric(c5, '代理口径红利成本', fmt_num(total_bonus), tone='warn')
            show_metric(c6, '公司净收入', fmt_num(total_co_income), tone=tone_by_sign(total_co_income))
            cv = total_ftd / total_signup * 100 if total_signup else 0
            show_metric(c7, '注册→首存转化', f'{cv:.1f}%', tone='accent')

            # 代理类型分布
            if '代理类型' in a_filt.columns:
                section_header('代理类型分布', '看各类代理对收入的贡献占比')
                type_grp = a_filt.groupby('代理类型').agg(
                    代理数=('代理编号', 'nunique'),
                    累积注册=('注册人数', 'sum'),
                    累积红利=('红利', 'sum'),
                    累积公司收入=('公司收入', 'sum'),
                ).reset_index().sort_values('累积公司收入', ascending=False)
                st.dataframe(type_grp, use_container_width=True, hide_index=True)

            # 代理聚合 (per 代理)
            if '代理编号' in a_filt.columns:
                group_cols = ['代理编号']
                if '代理名称' in a_filt.columns: group_cols.append('代理名称')
                if '代理类型' in a_filt.columns: group_cols.append('代理类型')
                agg_per_agent = a_filt.groupby(group_cols).agg(
                    累积注册=('注册人数', 'sum'),
                    累积首存=('首存人数', 'sum'),
                    累积投注=('有效投注额', 'sum'),
                    累积红利=('红利', 'sum'),
                    累积返水=('返水', 'sum'),
                    累积公司收入=('公司收入', 'sum'),
                ).reset_index()

                # 红利依赖度 = 红利 / 有效投注
                agg_per_agent['红利依赖度'] = (
                    agg_per_agent['累积红利'] / agg_per_agent['累积投注'].replace(0, 1)
                ).clip(upper=1).round(4)
                active = agg_per_agent[(agg_per_agent['累积注册'] > 0) | (agg_per_agent['累积投注'] > 0)].copy()

                # 散点图: 拉新 vs 公司收入 × 红利依赖度
                if not active.empty:
                    with st.container(border=True):
                        section_header('代理质量散点图', 'X=累积拉新 / Y=累积公司收入 / 颜色=红利依赖度（越红依赖度越高）')
                        scatter_df = active[active['累积注册'] > 0].copy()  # 只画有拉新的
                        if not scatter_df.empty:
                            hover = ['累积首存', '累积投注', '累积红利', '红利依赖度']
                            if '代理名称' in scatter_df.columns: hover.insert(0, '代理名称')
                            fig = px.scatter(
                                scatter_df, x='累积注册', y='累积公司收入',
                                color='红利依赖度',
                                color_continuous_scale='RdYlGn_r',
                                hover_data=hover,
                                height=500
                            )
                            fig.add_hline(y=0, line_dash='dot', line_color='rgba(150,170,210,0.5)')
                            fig.update_layout(margin=dict(l=40, r=40, t=20, b=40))
                            st.plotly_chart(fig, use_container_width=True)

                # 优质代理 Top 20
                section_header('优质代理 Top 20', '按公司净收入降序')
                top_q = active.sort_values('累积公司收入', ascending=False).head(20).copy()
                for c in ['累积投注', '累积红利', '累积返水', '累积公司收入']:
                    if c in top_q.columns:
                        top_q[c] = top_q[c].round(0)
                st.dataframe(top_q, use_container_width=True, hide_index=True)

                # 高红利依赖代理识别
                section_header('高红利依赖代理识别',
                              '红利依赖度 > 0.5（红利金额/投注金额）且 累积公司收入 ≤ 0，建议重点审视。')
                bonus_eaters = active[
                    (active['红利依赖度'] > 0.5) & (active['累积公司收入'] <= 0)
                ].sort_values('累积红利', ascending=False).head(30).copy()
                if bonus_eaters.empty:
                    st.info('未发现高红利依赖代理（标准：红利/投注 > 50% 且 公司收入 ≤ 0）')
                else:
                    for c in ['累积投注', '累积红利', '累积返水', '累积公司收入']:
                        if c in bonus_eaters.columns:
                            bonus_eaters[c] = bonus_eaters[c].round(0)
                    st.dataframe(bonus_eaters, use_container_width=True, hide_index=True)

                # 倒数 20: 公司收入最差代理
                section_header('亏损代理 Bottom 20', '按公司收入升序，亏损额最大的代理。')
                bottom = active.sort_values('累积公司收入').head(20).copy()
                for c in ['累积投注', '累积红利', '累积返水', '累积公司收入']:
                    if c in bottom.columns:
                        bottom[c] = bottom[c].round(0)
                st.dataframe(bottom, use_container_width=True, hide_index=True)

    st.markdown('---')
    with st.expander('ℹ️ 字段说明 / 计算口径'):
        st.markdown('''
- **红利总成本**: 仅含「状态=成功」的红利,失败 / 拒绝不算
- **同期公司收入(净)**: 来自 代理报表 (`raw_agent_report`),已扣红利/返水/佣金后的净收入
- **红利依赖度** = 累积红利 / 累积有效投注额 (越高代表代理拉来的会员靠红利,不靠真投注)
- **高红利依赖代理**: 红利依赖度 > 50% **AND** 公司收入 ≤ 0 (双条件,避免误判高产代理)
- **高频领取会员**: 单日领 >= 3 笔的会员 (按累积红利降序)
- **代理报表** = 每个代理每日 KPIs,这里按选定时间累加
- **数据范围**: 默认显示近 30 天,可在筛选器调整 / 切月
''')


def render_agent_member_matrix():
    """代理 × 会员 明细 — 6/4 新增 (Miru 风控决策面板)"""
    member = load_table('raw_member_report')

    hero('代理 × 会员 明细', '各代理名下会员 KPI 与套利特征识别，支持风控决策。',
         latest_imported_at(member),
         basis='会员报表＋代理报表（后台导出·每月上传）',
         detail=(
             '**分析范围**：各代理名下会员 KPI、累积投注与套利特征识别。\n\n'
             '**数据来源（后台导出 → 上传）**：会员报表（报表中心→会员报表）＋代理报表（报表中心→代理报表）。\n\n'
             '**计算口径**：会员身份＝会员账号＋代理；累积投注、公司净为逐月数值，跨快照相加为区间累计。\n\n'
             '**更新方式**：手动上传。完整对照见「数据说明」页。'
         ))

    if member.empty:
        st.warning('暂无会员数据')
        return

    # 准备数据 (合最新 snapshot 累积口径)
    df = member.copy()
    for col in ['存款额', '取款额', '有效投注额', '公司输赢', '红利', '返水', '公司收入', '首存金额']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # 代理为空标 (直客)
    df['代理'] = df['代理'].fillna('(直客)').replace('', '(直客)').astype(str)

    # ━━ 月份多选筛选 ━━
    if '_snapshot_month' in df.columns:
        all_months = sorted(df['_snapshot_month'].dropna().astype(str).unique().tolist())
        c1, c2 = st.columns([3, 1])
        with c1:
            sel_months = st.multiselect(
                '📅 月份筛选 (默认全部累积; 选某几月则只计该范围)',
                options=all_months, default=all_months, key='agent_matrix_months',
                help='会员报表是月度 snapshot, 所以筛选粒度是月, 不是日'
            )
        with c2:
            st.markdown(f'**已选**: {len(sel_months)} / {len(all_months)} 月')
        if sel_months:
            df = df[df['_snapshot_month'].astype(str).isin(sel_months)]
        else:
            st.warning('请至少选一个月份')
            return
        if df.empty:
            st.info('所选月份无数据')
            return

    # 累积到 (会员账号 × 代理) 层 — 跨月加总
    grp_member = df.groupby(['会员账号', '代理'], as_index=False).agg(
        VIP等级=('VIP等级', 'max'),
        累积存款=('存款额', 'sum'),
        累积取款=('取款额', 'sum'),
        累积有效投注=('有效投注额', 'sum'),
        累积公司输赢=('公司输赢', 'sum'),
        累积红利=('红利', 'sum'),
        累积返水=('返水', 'sum'),
        累积公司净收入=('公司收入', 'sum'),
        会员状态=('会员状态', 'last') if '会员状态' in df.columns else ('会员账号', 'count'),
        用户标签=('用户标签', 'last') if '用户标签' in df.columns else ('会员账号', 'count'),
        跨月数=('_snapshot_month', 'nunique') if '_snapshot_month' in df.columns else ('会员账号', 'count'),
    )

    # 红利依赖度 (红利+返水 / 投注)
    grp_member['红+返/投注%'] = (
        (grp_member['累积红利'] + grp_member['累积返水']) /
        grp_member['累积有效投注'].replace(0, 1) * 100
    ).round(2)
    grp_member['holding%'] = (
        grp_member['累积公司输赢'] / grp_member['累积有效投注'].replace(0, 1) * 100
    ).round(3)

    # 套利 flag
    grp_member['套利特征'] = (
        (grp_member['累积有效投注'] > 1_000_000) &
        (grp_member['累积公司输赢'].abs() / grp_member['累积有效投注'].replace(0, 1) < 0.005) &
        ((grp_member['累积红利'] + grp_member['累积返水']) > 10_000)
    )
    if '用户标签' in grp_member.columns:
        TAG_KEYWORDS = ['套利', '高风险', '多平台', '跨平台对打', '骗分', '专业玩家', '软件投注']
        # 避开 pandas 3.x + pyarrow 的 apply / map 不兼容 — 用 list comprehension
        grp_member['tag_命中'] = [
            ','.join([k for k in TAG_KEYWORDS if k in (str(x) if pd.notna(x) else '')]) or ''
            for x in grp_member['用户标签']
        ]
    else:
        grp_member['tag_命中'] = ''

    # ━━━━━━━━ 1. 代理总览 ━━━━━━━━
    section_header('1. 代理总览', '每个代理底下: 会员数 / 累积投注 / 累积净 / 红利占比 / 套利户数')
    grp_agent = grp_member.groupby('代理', as_index=False).agg(
        会员数=('会员账号', 'nunique'),
        累积投注=('累积有效投注', 'sum'),
        累积输赢=('累积公司输赢', 'sum'),
        累积红利=('累积红利', 'sum'),
        累积返水=('累积返水', 'sum'),
        累积公司净=('累积公司净收入', 'sum'),
        套利户数=('套利特征', 'sum'),
    )
    grp_agent['红+返/投注%'] = (
        (grp_agent['累积红利'] + grp_agent['累积返水']) /
        grp_agent['累积投注'].replace(0, 1) * 100
    ).round(2)
    grp_agent['holding%'] = (
        grp_agent['累积输赢'] / grp_agent['累积投注'].replace(0, 1) * 100
    ).round(3)
    # 排序: 累积投注 desc
    grp_agent = grp_agent.sort_values('累积投注', ascending=False)

    # KPI 摘要
    c1, c2, c3, c4 = st.columns(4)
    _arb_cnt = int(grp_agent['套利户数'].sum())
    show_metric(c1, '代理总数', fmt_num(grp_agent['代理'].nunique()))
    show_metric(c2, '会员总数', fmt_num(grp_agent['会员数'].sum()))
    show_metric(c3, '累积投注合计', fmt_num(grp_agent['累积投注'].sum()))
    show_metric(c4, '套利特征户数', fmt_num(_arb_cnt), tone='bad' if _arb_cnt else 'good',
                help_text='需要关注的风险会员数' if _arb_cnt else '未发现套利特征会员')

    # ── 套利特征会员明细（谁、归谁、为什么）──
    if _arb_cnt > 0:
        section_header(f'⚠️ 套利特征会员明细（{_arb_cnt} 个，就是上面那个数字的实际名单）',
                       '判定条件：累积有效投注 > 100 万　且　|holding| < 0.5%（庄家在这人身上几乎不赢不输）　且　红利+返水 > 1 万。'
                       '下面是被标记的会员、归哪个代理、关键数据——直接拿去查 / 处理。')
        arb = grp_member[grp_member['套利特征']].copy()
        arb['红利+返水'] = (arb['累积红利'].fillna(0) + arb['累积返水'].fillna(0))
        arb_cols = ['会员账号', '代理', 'VIP等级', '累积有效投注', 'holding%', '红利+返水',
                    '累积公司输赢', '累积公司净收入', 'tag_命中', '会员状态']
        arb_cols = [c for c in arb_cols if c in arb.columns]
        arb_show = arb[arb_cols].sort_values('累积有效投注', ascending=False)
        for c in ['累积有效投注', '红利+返水', '累积公司输赢', '累积公司净收入']:
            if c in arb_show.columns:
                arb_show[c] = arb_show[c].fillna(0).round(0).astype('int64')
        st.dataframe(arb_show, use_container_width=True, hide_index=True)
        st.caption('holding% 越接近 0 = 庄家在这人身上几乎没赢（套利典型）；红利+返水 = 这人累计拿走的红利+返水；'
                   'tag_命中 = 用户标签里命中的风险词（套利/高风险/多平台等）；空白代表只靠数据特征命中、标签没标。')

    # 搜索 / 排序
    search_q = st.text_input('🔍 搜代理名', '', key='agent_matrix_search', placeholder='例: newbee888')
    display_agent = grp_agent.copy()
    if search_q.strip():
        display_agent = display_agent[display_agent['代理'].str.contains(search_q.strip(), case=False, na=False)]

    # 排序
    sort_by = st.selectbox('排序', ['累积投注', '累积公司净', '会员数', '红+返/投注%', '套利户数'],
                          index=0, key='agent_matrix_sort')
    asc = st.checkbox('升序', value=False, key='agent_matrix_asc')
    display_agent = display_agent.sort_values(sort_by, ascending=asc)

    # 格式化数字
    show_df = display_agent.copy()
    for col in ['累积投注', '累积输赢', '累积红利', '累积返水', '累积公司净']:
        show_df[col] = show_df[col].round(0)
    st.markdown(f'**显示 {len(show_df)} / {len(grp_agent)} 个代理**')
    st.dataframe(show_df, use_container_width=True, hide_index=True, height=420,
                 column_config={
                     '累积投注': st.column_config.NumberColumn(format='%d'),
                     '累积输赢': st.column_config.NumberColumn(format='%d'),
                     '累积红利': st.column_config.NumberColumn(format='%d'),
                     '累积返水': st.column_config.NumberColumn(format='%d'),
                     '累积公司净': st.column_config.NumberColumn(format='%d'),
                 })

    # ━━━━━━━━ 2. 查询单一代理或会员 ━━━━━━━━
    section_header('2. 单查 代理 / 会员', '贴代理名 → 看该代理底下全部会员;贴会员账号 → 直接定位单一会员')

    c1, c2 = st.columns(2)
    with c1:
        sel_agent_input = st.text_input('🔍 输代理名 (例: newbee888)', '',
                                         key='agent_lookup', placeholder='留空则用下方下拉框选')
    with c2:
        sel_member_input = st.text_input('🔍 输会员账号 (例: yjno888)', '',
                                          key='member_lookup', placeholder='直接查会员明细')

    # 备用下拉框 (默认 100 大代理)
    agent_options = ['(用上方搜索框 / 选这里)'] + display_agent['代理'].head(200).tolist()
    sel_agent_dropdown = st.selectbox('或下拉选代理 (top 200 by 投注量)', agent_options, index=0,
                                       key='agent_matrix_select')

    # 决定要查的代理 (优先输入框 > 下拉)
    sel_agent = sel_agent_input.strip() if sel_agent_input.strip() else (
        sel_agent_dropdown if sel_agent_dropdown != '(用上方搜索框 / 选这里)' else None
    )

    # 单查会员模式 (优先级最高)
    if sel_member_input.strip():
        section_header(f'🎯 单一会员: {sel_member_input.strip()}', '')
        mem_q = sel_member_input.strip()
        mem_match = grp_member[grp_member['会员账号'].str.contains(mem_q, case=False, na=False)].copy()
        if mem_match.empty:
            st.warning(f'没找到会员账号 含 "{mem_q}" 的记录')
        else:
            st.markdown(f'找到 **{len(mem_match)}** 笔匹配:')
            cols = ['会员账号', '代理', 'VIP等级', '累积存款', '累积取款', '累积有效投注',
                    '累积公司输赢', '累积红利', '累积返水', '累积公司净收入',
                    'holding%', '红+返/投注%', '套利特征', '会员状态', '用户标签']
            cols = [c for c in cols if c in mem_match.columns]
            for c in ['累积存款', '累积取款', '累积有效投注', '累积公司输赢', '累积红利', '累积返水', '累积公司净收入']:
                if c in mem_match.columns:
                    mem_match[c] = mem_match[c].round(0)
            st.dataframe(mem_match[cols], use_container_width=True, hide_index=True)

            # 显示该会员 月度趋势 (跨月 snapshot)
            mem_acct_list = mem_match['会员账号'].unique().tolist()
            monthly = df[df['会员账号'].isin(mem_acct_list)].copy()
            if not monthly.empty and '_snapshot_month' in monthly.columns:
                section_header('该会员月度趋势', '')
                m_cols = ['_snapshot_month', '会员账号', 'VIP等级', '存款额', '取款额',
                          '有效投注额', '公司输赢', '红利', '返水', '公司收入', '会员状态']
                m_cols = [c for c in m_cols if c in monthly.columns]
                monthly_show = monthly[m_cols].sort_values(['会员账号', '_snapshot_month'])
                for c in ['存款额', '取款额', '有效投注额', '公司输赢', '红利', '返水', '公司收入']:
                    if c in monthly_show.columns:
                        monthly_show[c] = monthly_show[c].round(0)
                st.dataframe(monthly_show, use_container_width=True, hide_index=True)

    # 代理底下会员明细模式
    sub = None
    if sel_agent and sel_agent != '(用上方搜索框 / 选这里)':
        section_header(f'代理: {sel_agent} 名下全部会员', '')
        sub = grp_member[grp_member['代理'].str.contains(sel_agent, case=False, na=False)].copy()
        if sub.empty:
            st.warning(f'没找到代理 含 "{sel_agent}" 的会员')
            sub = None

    if sub is not None and not sub.empty:
        # 该代理 KPI
        c1, c2, c3, c4, c5 = st.columns(5)
        _sub_net = float(sub['累积公司净收入'].sum())
        _sub_arb = int(sub['套利特征'].sum())
        show_metric(c1, '会员数', fmt_num(sub['会员账号'].nunique()))
        show_metric(c2, '累积投注', fmt_num(sub['累积有效投注'].sum()))
        show_metric(c3, '累积公司净', fmt_num(_sub_net), tone=tone_by_sign(_sub_net))
        show_metric(c4, '套利户数', fmt_num(_sub_arb), tone='bad' if _sub_arb else None)
        tagged = (sub['tag_命中'].astype(str).str.len() > 0).sum() if 'tag_命中' in sub.columns else 0
        show_metric(c5, 'tag 命中数', fmt_num(tagged), tone='warn' if tagged else None)

        # 会员明细表
        cols_to_show = ['会员账号', 'VIP等级', '累积存款', '累积取款', '累积有效投注',
                        '累积公司输赢', '累积红利', '累积返水', '累积公司净收入',
                        'holding%', '红+返/投注%', '套利特征', '会员状态', '用户标签']
        cols_to_show = [c for c in cols_to_show if c in sub.columns]
        display_sub = sub[cols_to_show].copy()
        for c in ['累积存款', '累积取款', '累积有效投注', '累积公司输赢', '累积红利', '累积返水', '累积公司净收入']:
            if c in display_sub.columns:
                display_sub[c] = display_sub[c].round(0)
        display_sub = display_sub.sort_values('累积有效投注', ascending=False)
        st.dataframe(display_sub, use_container_width=True, hide_index=True)

        # ─── 该代理的「官方代理报表 月度 KPI」(直接拉 raw_agent_report) ───
        # 这才是跟包网商 customer service 提供的格式一致的口径
        st.markdown('---')
        st.markdown(f'**📊 该代理「官方代理报表」月度 KPI** (raw_agent_report 直接读取,跟包网商提供的格式一致)')
        try:
            agent_rep = load_table('raw_agent_report')
            if not agent_rep.empty:
                mask = (agent_rep['代理名称'].astype(str).str.lower() == sel_agent.lower()) | \
                       (agent_rep['代理编号'].astype(str).str.lower() == sel_agent.lower())
                ar = agent_rep[mask].copy()
                if not ar.empty:
                    ar['月'] = ar['日期'].astype(str).str[:7]
                    for col in ['注册人数','首存人数','投注人数','存款额','取款额',
                                '有效投注额','公司输赢','红利','返水','代理佣金',
                                '公司收入','提前结算','场馆费']:
                        if col in ar.columns:
                            ar[col] = pd.to_numeric(ar[col], errors='coerce').fillna(0)
                    grpcols = ['注册人数','首存人数','投注人数','存款额','取款额',
                               '有效投注额','公司输赢','红利','返水','代理佣金',
                               '公司收入','提前结算']
                    grpcols = [c for c in grpcols if c in ar.columns]
                    monthly = ar.groupby('月', as_index=False)[grpcols].sum()
                    for c in ['存款额','取款额','有效投注额','公司输赢','红利','返水',
                              '代理佣金','公司收入','提前结算']:
                        if c in monthly.columns:
                            monthly[c] = monthly[c].round(2)
                    st.dataframe(monthly, use_container_width=True, hide_index=True)
                    st.caption('注: 此处「公司收入」= 公司输赢 - 红利 - 返水 - 佣金 + 系统调整 + 提前结算 (后台口径). '
                               '包网商若再扣 场馆费 + 手续费 = 平台净盈利 (会小于公司收入).')
                else:
                    st.info(f'代理报表无 {sel_agent} 数据')
        except Exception as e:
            st.warning(f'代理报表查询失败: {e}')

    # ━━━━━━━━ 3. 全部会员表 ━━━━━━━━
    section_header('3. 全部会员表',
                  f'{len(grp_member)} 个会员 (跨所选月份),可搜寻 / 排序 / 筛 VIP / 筛代理')

    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        mem_search = st.text_input('🔍 搜会员账号 / 代理名 (会员账号或代理名任一含关键词都显示)',
                                    '', key='all_mem_search',
                                    placeholder='例: yjno888 或 newbee888')
    with c2:
        vip_filter = st.selectbox('VIP 等级筛选', ['(全部)'] + sorted(grp_member['VIP等级'].dropna().astype(str).unique().tolist()),
                                   key='all_mem_vip')
    with c3:
        status_filter = st.selectbox('会员状态', ['(全部)', '启用', '禁用'],
                                      key='all_mem_status') if '会员状态' in grp_member.columns else '(全部)'

    all_mem = grp_member.copy()
    if mem_search.strip():
        q = mem_search.strip()
        all_mem = all_mem[
            all_mem['会员账号'].astype(str).str.contains(q, case=False, na=False) |
            all_mem['代理'].astype(str).str.contains(q, case=False, na=False)
        ]
    if vip_filter != '(全部)':
        all_mem = all_mem[all_mem['VIP等级'].astype(str) == vip_filter]
    if status_filter != '(全部)' and '会员状态' in all_mem.columns:
        all_mem = all_mem[all_mem['会员状态'].astype(str) == status_filter]

    mem_sort_by = st.selectbox(
        '排序', ['累积有效投注', '累积公司净收入', '累积红利', '累积存款', '累积取款', 'holding%'],
        index=0, key='all_mem_sort'
    )
    mem_asc = st.checkbox('升序', value=False, key='all_mem_asc')
    all_mem = all_mem.sort_values(mem_sort_by, ascending=mem_asc)

    st.markdown(f'**显示 {len(all_mem)} / {len(grp_member)} 个会员**')

    mem_cols = ['会员账号', '代理', 'VIP等级', '累积存款', '累积取款', '累积有效投注',
                '累积公司输赢', '累积红利', '累积返水', '累积公司净收入',
                'holding%', '红+返/投注%', '套利特征', '会员状态', '用户标签']
    mem_cols = [c for c in mem_cols if c in all_mem.columns]
    all_mem_show = all_mem[mem_cols].copy()
    for c in ['累积存款', '累积取款', '累积有效投注', '累积公司输赢',
              '累积红利', '累积返水', '累积公司净收入']:
        if c in all_mem_show.columns:
            all_mem_show[c] = all_mem_show[c].round(0)
    st.dataframe(all_mem_show, use_container_width=True, hide_index=True, height=500)

    # ━━━━━━━━ 4. 系统已 tag 但状态启用 ━━━━━━━━
    section_header('4. 系统已标记但状态启用的会员',
                  '用户标签命中 套利 / 高风险 / 多平台 / 跨平台对打 / 骗分 / 专业玩家 / 软件投注 的会员')

    if 'tag_命中' in grp_member.columns:
        tagged_df = grp_member[grp_member['tag_命中'].astype(str).str.len() > 0].copy()
        # 会员状态 = 启用
        if '会员状态' in tagged_df.columns:
            tagged_df = tagged_df[tagged_df['会员状态'].astype(str) == '启用']

        # 按累积投注降序
        tagged_df = tagged_df.sort_values('累积有效投注', ascending=False).head(100)

        st.markdown(f'共 **{len(tagged_df)}** 笔会员命中 tag 且状态启用 (top 100):')

        cols = ['会员账号', '代理', 'VIP等级', '累积有效投注', '累积公司输赢',
                '累积红利', '累积返水', '累积公司净收入', 'holding%',
                '红+返/投注%', '套利特征', 'tag_命中', '用户标签']
        cols = [c for c in cols if c in tagged_df.columns]
        for c in ['累积有效投注', '累积公司输赢', '累积红利', '累积返水', '累积公司净收入']:
            if c in tagged_df.columns:
                tagged_df[c] = tagged_df[c].round(0)
        st.dataframe(tagged_df[cols], use_container_width=True, hide_index=True)

        # 套利特征 + 启用 但 公司净亏 的清单
        st.markdown('---')
        st.markdown('**套利特征 + 状态启用 + 公司净亏损 的会员（建议优先处理）：**')
        action_df = tagged_df[
            (tagged_df['套利特征']) &
            (tagged_df['累积公司净收入'] < 0)
        ].sort_values('累积公司净收入').head(50)
        if action_df.empty:
            st.info('无符合条件 (套利特征 + 状态启用 + 公司净亏损)')
        else:
            st.dataframe(action_df[cols], use_container_width=True, hide_index=True)

    st.markdown('---')
    with st.expander('ℹ️ 字段说明 / 计算口径'):
        st.markdown('''
- **代理**: 取自 `raw_member_report.代理` 字段, 跨月相同 = 同一会员
- **累积**: 跨所有月份 snapshot 加总 (会员可能多月出现)
- **holding%** = 累积公司输赢 / 累积有效投注 × 100, **平台水位 1-1.5% 是正常**
- **红+返/投注%** = (累积红利 + 累积返水) / 累积有效投注 × 100, **正常会员 < 1.5%**
- **套利特征**: 累积投注 > 100 万 **且** |holding| < 0.5% **且** 红+返 > 1 万
- **tag 命中关键词**: 套利 / 高风险 / 多平台 / 跨平台对打 / 骗分 / 专业玩家 / 软件投注
- **状态启用**: 即「会员状态 = 启用」, 表示账户没被禁用
- **直客**: `代理` 字段为空 / null, 归入「(直客)」
''')


# 对话内容按「发言人>时间戳 :」切分成轮次
_CS_TURN_RE = re.compile(r'\n\s*([^\n>]{1,40})>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\s*:')


def cs_member_text(content, agent) -> str:
    """只抽取「会员(访客)」说的话，剔除 系统 自动消息 + 接待客服 的话术/模板开场。
    每段对话都以一长串系统提醒 + 客服欢迎模板开头(含 提款/返水/VIP/体育/下载/注册 等词),
    若直接扫全文,这些词会在每笔对话里命中,把热点图顶到天花板。去掉非会员发言后才是真实问题。"""
    s = str(content)
    parts = _CS_TURN_RE.split(s)
    # parts[0] = 开场白("对话开始>>..."), 之后是 [发言人, 内容, 发言人, 内容, ...]
    if len(parts) < 3:
        return ''  # 解析不出轮次,宁可空也不要把模板算进去
    agent = str(agent).strip()
    keep = []
    for i in range(1, len(parts) - 1, 2):
        speaker = parts[i].strip()
        if speaker == '系统' or (agent and speaker == agent):
            continue
        keep.append(parts[i + 1])
    return ' '.join(keep)


def render_cs_analysis():
    """客服对话分析 — 6/5 新增 (Miru 客服质量决策面板)"""
    cs = load_table('raw_cs_conversations')
    hero('客服分析', '5 月包网商前线客服对话数据 — 客服效率 / 满意度 / 时段 / 会员侧。',
         latest_imported_at(cs),
         basis='客服对话（客服系统导出）＋会员报表（VIP 关联）',
         detail=(
             '**分析范围**：客服效率、满意度、时段／星期分布、服务主题与会员侧 VIP 关联。\n\n'
             '**数据来源**：\n'
             '- 客服对话（非后台报表，由客服系统导出 xlsx）\n'
             '- 会员报表（用于会员侧 VIP 关联）\n\n'
             '**更新方式**：人工提供后于「数据上传」页上传。完整对照见「数据说明」页。'
         ))

    if cs.empty:
        st.warning('暂无客服对话数据')
        return

    # 准备时间
    if '开始时间' in cs.columns:
        cs['开始时间'] = to_datetime_safe(cs['开始时间'])
        cs['日期'] = cs['开始时间'].dt.date
        cs['小时'] = cs['开始时间'].dt.hour
        cs['星期'] = cs['开始时间'].dt.dayofweek
        cs['星期'] = cs['星期'].map({0: '周一', 1: '周二', 2: '周三', 3: '周四', 4: '周五', 5: '周六', 6: '周日'})

    for c in ['首次响应', '平均响应', '总时长', '访客消息数', '客服消息数', '对话回合数']:
        if c in cs.columns:
            cs[c] = pd.to_numeric(cs[c], errors='coerce')

    # 日期筛选
    cs_filt, _, _, _ = date_range_picker(cs, '开始时间', 'cs', default_last_days=None)
    if cs_filt.empty:
        st.info('该范围无对话数据')
        return

    # KPI
    total_conv = len(cs_filt)
    unique_agents = cs_filt['接待客服'].nunique() if '接待客服' in cs_filt.columns else 0
    unique_members = member_count(cs_filt)
    avg_first_resp = safe_mean(cs_filt, '首次响应')
    avg_total_time = safe_mean(cs_filt, '总时长')
    rated_mask = cs_filt['满意度评价'].astype(str) != '未评价' if '满意度评价' in cs_filt.columns else None
    rated_count = rated_mask.sum() if rated_mask is not None else 0
    unhappy_count = (cs_filt['满意度评价'].astype(str) == '非常不满意').sum() if '满意度评价' in cs_filt.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    show_metric(c1, '对话总数', fmt_num(total_conv))
    show_metric(c2, '接待客服数', fmt_num(unique_agents))
    show_metric(c3, '涉及会员数', fmt_num(unique_members))
    show_metric(c4, '平均首次响应 (秒)', f'{avg_first_resp:.1f}')

    c5, c6, c7, c8 = st.columns(4)
    show_metric(c5, '平均总时长 (秒)', f'{avg_total_time:.0f}')
    show_metric(c6, '已评价对话', fmt_num(rated_count))
    rate_pct = (rated_count / total_conv * 100) if total_conv else 0
    show_metric(c7, '评价率', f'{rate_pct:.1f}%')
    show_metric(c8, '非常不满意', fmt_num(unhappy_count), tone='bad' if unhappy_count else None)

    # ━━━ 满意度分布 ━━━
    if '满意度评价' in cs_filt.columns:
        with st.container(border=True):
            section_header('满意度分布', '')
            sat_grp = cs_filt.groupby('满意度评价').size().reset_index(name='笔数').sort_values('笔数', ascending=False)
            sat_grp['占比%'] = (sat_grp['笔数'] / sat_grp['笔数'].sum() * 100).round(2)
            c1, c2 = st.columns([1, 2])
            with c1:
                st.dataframe(sat_grp, use_container_width=True, hide_index=True)
            with c2:
                fig = px.pie(sat_grp, names='满意度评价', values='笔数', height=300, hole=0.5,
                             color='满意度评价',
                             color_discrete_map={'非常满意': GREEN, '满意': CYAN, '一般': BLUE,
                                                 '不满意': AMBER, '非常不满意': RED, '未评价': '#64748b'})
                fig.update_traces(textinfo='percent', marker=dict(line=dict(color='rgba(7,15,30,0.9)', width=2)))
                fig.update_layout(margin=dict(l=20, r=20, t=20, b=20))
                st.plotly_chart(fig, use_container_width=True)

    # ━━━ 客服效率排行 ━━━
    if '接待客服' in cs_filt.columns:
        section_header('客服效率排行', '按接待量降序，对比各客服的响应速度、对话时长与满意度。')
        agg = cs_filt.groupby('接待客服').agg(
            接待量=('对话ID', 'count'),
            平均首次响应秒=('首次响应', 'mean'),
            平均响应秒=('平均响应', 'mean'),
            平均总时长秒=('总时长', 'mean'),
            平均对话回合=('对话回合数', 'mean'),
        ).reset_index()
        # 满意度
        sat = cs_filt[cs_filt['满意度评价'].astype(str) != '未评价'].groupby('接待客服').agg(
            已评价数=('对话ID', 'count'),
            非常满意=('满意度评价', lambda x: (x == '非常满意').sum()),
            非常不满意=('满意度评价', lambda x: (x == '非常不满意').sum()),
        ).reset_index()
        agg = pd.merge(agg, sat, on='接待客服', how='left').fillna(0)
        agg['好评率%'] = (agg['非常满意'] / agg['已评价数'].replace(0, 1) * 100).round(1)
        for c in ['平均首次响应秒', '平均响应秒', '平均总时长秒', '平均对话回合']:
            agg[c] = agg[c].round(1)
        for c in ['已评价数', '非常满意', '非常不满意']:
            agg[c] = agg[c].astype(int)
        agg = agg.sort_values('接待量', ascending=False)
        st.dataframe(agg, use_container_width=True, hide_index=True)

    # ━━━ 时段分布 ━━━
    hr_c1, hr_c2 = st.columns(2)
    with hr_c1:
        if '小时' in cs_filt.columns:
            with st.container(border=True):
                section_header('小时分布', '接客高峰时段，对照客服排班是否合理。')
                hr_grp = cs_filt.groupby('小时').size().reset_index(name='对话数')
                fig = px.bar(hr_grp, x='小时', y='对话数', color='对话数', color_continuous_scale='Blues', height=300)
                fig.update_layout(margin=dict(l=40, r=40, t=20, b=40), coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)
    with hr_c2:
        if '星期' in cs_filt.columns:
            with st.container(border=True):
                section_header('星期分布', '')
                wk_grp = cs_filt.groupby('星期').size().reset_index(name='对话数')
                order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
                wk_grp['order'] = wk_grp['星期'].map({d: i for i, d in enumerate(order)})
                wk_grp = wk_grp.sort_values('order')
                fig = px.bar(wk_grp, x='星期', y='对话数', color='对话数', color_continuous_scale='Greens', height=300)
                fig.update_layout(margin=dict(l=40, r=40, t=20, b=40), coloraxis_showscale=False, xaxis_title=None)
                st.plotly_chart(fig, use_container_width=True)

    # ━━━ 服务主题 ━━━
    if '服务主题' in cs_filt.columns:
        section_header('服务主题分布 (Top 20)', '客服主动 / 系统打的主题标签')
        theme_grp = cs_filt[cs_filt['服务主题'].notna()].groupby('服务主题').size().reset_index(name='笔数').sort_values('笔数', ascending=False).head(20)
        if not theme_grp.empty:
            st.dataframe(theme_grp, use_container_width=True, hide_index=True)
        else:
            st.info('5 月对话基本没标服务主题 — 这本身是 QC 问题点 (无法按主题分流追踪)')

    # ━━━ 不满意案件 deep dive ━━━
    if '满意度评价' in cs_filt.columns:
        section_header('非常不满意 / 不满意 案件',
                      '主因 = 服务主题(优先) > 评价内容关键词 > 对话内容(去模板) ;点开看完整对话')
        bad = cs_filt[cs_filt['满意度评价'].astype(str).isin(['非常不满意', '不满意'])].copy()
        if bad.empty:
            st.info('该范围无不满意案件')
        else:
            # 主因分布
            if '_extracted_issue' in bad.columns:
                from collections import Counter
                issue_cnt = Counter()
                for s in bad['_extracted_issue'].dropna().astype(str):
                    if s and s != '(未匹配)':
                        for cat in s.split(','):
                            issue_cnt[cat.strip()] += 1
                if issue_cnt:
                    st.markdown('**主因分布:**')
                    iss_df = pd.DataFrame(issue_cnt.most_common(), columns=['主因', '案件数'])
                    c1, c2 = st.columns([1, 1])
                    with c1:
                        st.dataframe(iss_df, use_container_width=True, hide_index=True, height=240)
                    with c2:
                        fig = px.bar(iss_df.head(10), x='案件数', y='主因', orientation='h',
                                     color='案件数', color_continuous_scale='Reds', height=240)
                        fig.update_layout(margin=dict(l=20, r=20, t=10, b=10),
                                          yaxis={'categoryorder': 'total ascending'})
                        st.plotly_chart(fig, use_container_width=True)

            # 表格
            cols = ['开始时间', '会员账号', '接待客服', '满意度评价', '_extracted_issue',
                    '评价内容', '服务主题', '首次响应', '平均响应', '总时长', '对话回合数', '对话ID']
            cols = [c for c in cols if c in bad.columns]
            bad_show = bad[cols].copy()
            if '_extracted_issue' in bad_show.columns:
                bad_show = bad_show.rename(columns={'_extracted_issue': '主因'})
            if '开始时间' in bad_show.columns:
                bad_show = bad_show.sort_values('开始时间', ascending=False)
            st.dataframe(bad_show, use_container_width=True, hide_index=True)

            # 完整对话 — 支持: 1) 贴对话ID(查全部对话) 2) 下拉选(只看不满意)
            if '对话ID' in bad.columns and '对话内容' in bad.columns:
                st.markdown('**👇 查看完整对话内容** (贴对话 ID 或下拉选)')

                bad_sorted = bad.sort_values('开始时间', ascending=False).reset_index(drop=True)
                option_labels = []
                for i in range(len(bad_sorted)):
                    r = bad_sorted.iloc[i]
                    t_str = r['开始时间'].strftime('%m-%d %H:%M') if pd.notna(r['开始时间']) else '?'
                    label = f"{i+1}. {t_str} | {r['会员账号'] or '匿名'} | {r['接待客服']} | {r.get('_extracted_issue') or r.get('服务主题') or '(无主因)'}"
                    option_labels.append(label)

                c_id, c_dd = st.columns([2, 3])
                with c_id:
                    paste_id = st.text_input('贴对话 ID (可查全部 2535 笔,不限不满意)', '',
                                              key='cs_paste_id',
                                              placeholder='例: fabf3ffbb53b47de8b56d076181cebe2')
                with c_dd:
                    sel = st.selectbox('或从「不满意」下拉选', ['(请选择)'] + option_labels, key='cs_bad_select')

                # 找 row
                selected_row = None
                convo_key_id = 'none'
                if paste_id.strip():
                    q = paste_id.strip()
                    match = cs_filt[cs_filt['对话ID'].astype(str).str.contains(q, case=False, na=False)]
                    if match.empty:
                        st.warning(f'没找到对话 ID 含 "{q}" 的对话')
                    else:
                        if len(match) > 1:
                            st.info(f'找到 {len(match)} 笔匹配,显示第 1 笔. ID 完整些可以唯一定位')
                        selected_row = match.iloc[0]
                        convo_key_id = f'paste_{q[:20]}'
                elif sel != '(请选择)':
                    idx = option_labels.index(sel)
                    selected_row = bad_sorted.iloc[idx]
                    convo_key_id = f'dropdown_{idx}'

                if selected_row is not None:
                    c1, c2, c3 = st.columns(3)
                    with c1: st.metric('会员账号', selected_row['会员账号'] or '匿名')
                    with c2: st.metric('接待客服', selected_row['接待客服'])
                    with c3: st.metric('满意度', selected_row['满意度评价'])
                    c4, c5, c6 = st.columns(3)
                    with c4: st.metric('首次响应 (秒)', f"{selected_row['首次响应']:.0f}" if pd.notna(selected_row['首次响应']) else '-')
                    with c5: st.metric('总时长 (秒)', f"{selected_row['总时长']:.0f}" if pd.notna(selected_row['总时长']) else '-')
                    with c6: st.metric('对话回合', f"{selected_row['对话回合数']:.0f}" if pd.notna(selected_row['对话回合数']) else '-')
                    if selected_row.get('评价内容'):
                        st.markdown(f"**客户评价内容:** {selected_row['评价内容']}")
                    if selected_row.get('服务主题'):
                        st.markdown(f"**服务主题:** {selected_row['服务主题']}")
                    issue = selected_row.get('_extracted_issue')
                    if issue and str(issue).strip() and str(issue) != '(未匹配)':
                        st.markdown(f"**抽取主因:** {issue}")
                    st.markdown('---')
                    st.markdown('**对话全文：**')
                    convo = str(selected_row.get('对话内容') or '').replace('&nbsp;', ' ')
                    st.text_area('对话全文', value=convo, height=400, label_visibility='collapsed',
                                 key=f'cs_convo_{convo_key_id}')

    # ━━━ 会员侧 join (会员账号 → BQ raw_member_report) ━━━
    if '会员账号' in cs_filt.columns:
        section_header('会员侧分析 — VIP 等级 × 问题量', '对话会员账号关联会员表，分析各 VIP 等级的客诉分布。')
        try:
            member = load_table('raw_member_report')
            if not member.empty and 'VIP等级' in member.columns:
                # 取最新月份的 VIP
                latest_vip = member.sort_values('_snapshot_month').groupby('会员账号').tail(1)[['会员账号', 'VIP等级', '代理', '用户标签']]
                merged_cs = pd.merge(cs_filt, latest_vip, on='会员账号', how='left')
                vip_grp = merged_cs.groupby('VIP等级', dropna=False).size().reset_index(name='对话数').sort_values('对话数', ascending=False)
                vip_grp['VIP等级'] = vip_grp['VIP等级'].fillna('(未匹配)').astype(str)
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.dataframe(vip_grp, use_container_width=True, hide_index=True)
                with c2:
                    fig = px.bar(vip_grp, x='VIP等级', y='对话数', color='对话数',
                                 color_continuous_scale='Oranges', height=320)
                    fig.update_layout(margin=dict(l=40, r=20, t=20, b=40))
                    st.plotly_chart(fig, use_container_width=True)
                match_rate = (vip_grp[vip_grp['VIP等级'] != '(未匹配)']['对话数'].sum() / vip_grp['对话数'].sum() * 100) if vip_grp['对话数'].sum() else 0
                st.caption(f'会员账号匹配率: {match_rate:.1f}% (未匹配多半是访客未登入)')
            else:
                st.info('会员表无 VIP等级 字段,无法 join')
        except Exception as e:
            st.warning(f'会员侧 join 失败: {e}')

    # ━━━ 关键词热点 ━━━
    if '对话内容' in cs_filt.columns:
        section_header('对话关键词热点 (Top 25)', '只扫会员发言(已剔除系统提醒+客服模板开场),看真实问题集中在哪')
        KEYWORDS = ['充值', '提款', '红利', '彩金', '活动', '登录', '登入', '密码', '注册',
                    '验证', '体育', '真人', '电子', '老虎机', '棋牌', '彩票',
                    '风控', '冻结', '禁用', '解封', '套利', '客诉', '投诉',
                    '代理', '佣金', 'VIP', '返水', 'USDT', '虚拟币',
                    'APP', '下载', '链接']
        member_text = cs_filt.apply(
            lambda r: cs_member_text(r['对话内容'], r['接待客服'] if '接待客服' in cs_filt.columns else ''),
            axis=1)
        kw_counts = {}
        for kw in KEYWORDS:
            # 关键词当纯文本匹配(regex=False),大小写不敏感,APP/app 合并成一条
            cnt = member_text.str.contains(kw, case=False, na=False, regex=False).sum()
            if cnt > 0:
                kw_counts[kw] = int(cnt)
        kw_df = pd.DataFrame(list(kw_counts.items()), columns=['关键词', '对话数']).sort_values('对话数', ascending=False).head(25)
        if not kw_df.empty:
            fig = px.bar(kw_df, x='关键词', y='对话数', color='对话数',
                         color_continuous_scale='Viridis', height=380)
            fig.update_layout(margin=dict(l=40, r=20, t=20, b=80))
            st.plotly_chart(fig, use_container_width=True)
            st.caption(f'基于 {len(cs_filt)} 笔对话的会员发言统计 (每个关键词 = 有多少笔对话的会员提到过它)')

    st.markdown('---')
    with st.expander('ℹ️ 字段说明 / 数据来源'):
        st.markdown('''
- **数据来源**: 包网商提供的 `5月客服对话.xlsx` (32 sheets / 2,535 对话 / 31 天 / 57 客服)
- **首次响应** = 客服第一次回覆的延迟秒数
- **平均响应** = 客服每次回覆的平均延迟秒数
- **总时长** = 整段对话从开始到结束的秒数
- **对话回合数** = 双方互动的轮次
- **评价率** = 已评价对话 / 总对话 (基准: 包网商客服评价率 ~12%,远低于自营 CS 标杆 30%+)
- **会员侧 join** = 把对话的「会员账号」对应到 `raw_member_report` 取 VIP 等级 (未匹配的是访客未登入或会员账号不一致)
- **关键词热点** = 先剔除「系统自动提醒 + 客服模板开场」,只扫**会员发言**,再用业务关键词计数 (非 NLP 聚类,是 keyword count;不去模板会被开场白里的 提款/返水/VIP 等词顶到天花板)
''')


# ── 电访召回（会员召回电话）：上传当月「撥打紀錄總表」Excel，自动出漏斗 + 各专员 + ROI ──
_WINBACK_SKIP_SHEETS = {'话术', '统计结果'}


def _wb_num(x) -> float:
    if x is None:
        return 0.0
    t = str(x).strip().replace(',', '').replace('，', '')
    if t in ('', 'nan', 'None', '-', '—'):
        return 0.0
    try:
        return float(t)
    except ValueError:
        return 0.0


def _wb_yes(x) -> bool:
    return str(x).strip() in ('是', 'Y', 'y', 'TRUE', 'True', '1')


def parse_winback_file(uploaded) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """解析撥打紀錄總表。回传 (per-agent DataFrame, 元信息)。
    统计结果页是块状布局(每位专员占 2 列:标签/数值)；各专员明细页补 召回充值 + 七天回登。"""
    xls = pd.read_excel(uploaded, sheet_name=None, header=None, dtype=str)
    meta = {'sheets': ', '.join(xls.keys())}

    # 1. 官方「统计结果」块状解析
    official: Dict[str, dict] = {}
    stats_raw = xls.get('统计结果')
    if stats_raw is not None and not stats_raw.empty:
        m = stats_raw.values.tolist()
        row0 = m[0]
        for c in range(0, len(row0), 2):
            name = str(row0[c]).strip() if (c < len(row0) and row0[c] is not None) else ''
            if not name or name.lower() in ('none', 'nan'):
                continue
            d = {}
            for r in range(1, len(m)):
                if c + 1 < len(m[r]):
                    lbl = m[r][c]
                    if lbl is not None and str(lbl).strip():
                        d[str(lbl).strip()] = m[r][c + 1]
            official[name] = d

    # 2. 每位专员明细页 → 召回充值 + 七天回登；顺便从拨打日期判月份
    rows = []
    month_votes: Dict[str, int] = {}
    for name, d in official.items():
        detail = None
        for sn, df in xls.items():
            if sn.strip().lower() == name.strip().lower():
                detail = df
                break
        recharge, relogin, n_detail = 0.0, 0, 0
        if detail is not None and not detail.empty:
            mm = detail.values.tolist()
            hdr_idx = None
            for i, row in enumerate(mm[:4]):
                if any(str(c).strip() == '账号' for c in row if c is not None):
                    hdr_idx = i
                    break
            if hdr_idx is not None:
                hdr = [str(c).strip() if c is not None else '' for c in mm[hdr_idx]]

                def _col(key, _hdr=hdr):
                    for j, h in enumerate(_hdr):
                        if key in h:
                            return j
                    return None

                c_acct = _col('账号')
                c_rech = _col('充值金额')
                c_login = _col('七天')
                if c_login is None:
                    c_login = _col('登入')
                c_date = _col('拨打日期')
                if c_date is None:
                    c_date = _col('日期')
                for row in mm[hdr_idx + 1:]:
                    if c_acct is None or c_acct >= len(row) or row[c_acct] is None or str(row[c_acct]).strip() == '':
                        continue
                    n_detail += 1
                    if c_rech is not None and c_rech < len(row):
                        recharge += _wb_num(row[c_rech])
                    if c_login is not None and c_login < len(row) and _wb_yes(row[c_login]):
                        relogin += 1
                    if c_date is not None and c_date < len(row) and row[c_date] is not None:
                        mt = re.match(r'\s*(\d{4})[-/年](\d{1,2})', str(row[c_date]))
                        if mt:
                            ym = f'{int(mt.group(1))}-{int(mt.group(2)):02d}'
                            month_votes[ym] = month_votes.get(ym, 0) + 1
        rows.append({
            '专员': name,
            '名单数': int(_wb_num(d.get('名单数'))),
            '已播数': int(_wb_num(d.get('已播数'))),
            '接通数': int(_wb_num(d.get('接通数'))),
            '有效通话': int(_wb_num(d.get('有效通话'))),
            '申请彩金': int(_wb_num(d.get('是否申请彩金'))),
            '七天回登': relogin,
            '召回充值': recharge,
        })
    if month_votes:
        meta['month'] = max(month_votes.items(), key=lambda kv: kv[1])[0]
    return pd.DataFrame(rows), meta


_CN_MONTH = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6,
             '七': 7, '八': 8, '九': 9, '十': 10, '十一': 11, '十二': 12}


def _winback_label(meta: dict, fname: str) -> str:
    """优先用数据里拨打日期判出的月份；否则从档名解析（X月 / 五月）。"""
    if meta.get('month'):
        y, m = meta['month'].split('-')
        return f'{y}年{int(m)}月'
    mt = re.search(r'(\d{1,2})\s*月', fname or '')
    if mt:
        return f'{int(mt.group(1))}月'
    for cn, num in _CN_MONTH.items():
        if f'{cn}月' in (fname or ''):
            return f'{num}月'
    return (fname or '本期').replace('.xlsx', '')


def _winback_agg(df: pd.DataFrame) -> dict:
    n_dial = int(df['已播数'].sum())
    n_conn = int(df['接通数'].sum())
    return {
        '名单数': int(df['名单数'].sum()),
        '已播数': n_dial,
        '接通数': n_conn,
        '有效通话': int(df['有效通话'].sum()),
        '七天回登': int(df['七天回登'].sum()),
        '召回充值': float(df['召回充值'].sum()),
        '接通率': (n_conn / n_dial) if n_dial else 0.0,
        '有效通话率': (int(df['有效通话'].sum()) / n_conn) if n_conn else 0.0,
        '七天回登率': (int(df['七天回登'].sum()) / int(df['名单数'].sum())) if int(df['名单数'].sum()) else 0.0,
    }


def _winback_ym_from_name(fname: str) -> str:
    """从档名兜底判月份(YYYY-MM)；优先用数据里拨打日期判出的 meta['month']，这里只兜底。"""
    s = fname or ''
    mt = re.search(r'(20\d{2})\D{0,3}(\d{1,2})\s*月', s)
    if mt:
        return f'{int(mt.group(1))}-{int(mt.group(2)):02d}'
    mt2 = re.search(r'(20\d{2})[-/](\d{1,2})', s)
    if mt2:
        return f'{int(mt2.group(1))}-{int(mt2.group(2)):02d}'
    for cn, num in _CN_MONTH.items():
        if f'{cn}月' in s:
            return f'2026-{num:02d}'  # 无年份兜底；正常会被 meta['month'] 覆盖
    return ''


def _bq_winback_month_exists(client, ym: str) -> int:
    try:
        sql = f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.raw_winback` WHERE CAST(`月份` AS STRING)=@m"
        cfg = bigquery.QueryJobConfig(query_parameters=[bigquery.ScalarQueryParameter('m', 'STRING', ym)])
        return int(list(client.query(sql, job_config=cfg).result())[0].n)
    except Exception:
        return 0


def _write_winback(client, new_df, months, source_file):
    """电访按「月份」刷新：去掉这些月的旧行→写新行，其他月一行不动。沙盒禁 DML→整表读改写。"""
    table = f"{BQ_PREFIX}.raw_winback"
    try:
        existing = client.query(f"SELECT * FROM `{table}`").result().to_dataframe()
    except Exception:
        existing = pd.DataFrame()
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    months = [str(m) for m in months]
    if not existing.empty and '月份' in existing.columns:
        keep = existing[~existing['月份'].astype(str).isin(months)].copy()
        # 防掉数据：其他月行数必须原样保留
        for m, cnt in existing['月份'].astype(str).value_counts().to_dict().items():
            if m not in months and int((keep['月份'].astype(str) == m).sum()) != cnt:
                raise RuntimeError(f'安全中止：其他月份 {m} 行数会变，拒绝写入')
        combined = pd.concat([keep, nd], ignore_index=True)
    else:
        combined = nd
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, table, job_config=cfg).result()
    return len(combined)


def _winback_month_view(df: pd.DataFrame, label: str):
    df = df.sort_values('召回充值', ascending=False).reset_index(drop=True)
    a = _winback_agg(df)
    n_list, n_dial, n_conn, n_valid = a['名单数'], a['已播数'], a['接通数'], a['有效通话']
    n_relogin, sum_rech = a['七天回登'], a['召回充值']
    conn_rate, valid_rate, relogin_rate = a['接通率'], a['有效通话率'], a['七天回登率']

    cols = st.columns(5)
    show_metric(cols[0], '名单总数', fmt_num(n_list), help_text='本月分配给电访专员的待召回会员数')
    show_metric(cols[1], '接通率', fmt_pct(conn_rate), help_text='接通数 ÷ 已播数', tone='accent')
    show_metric(cols[2], '有效通话率', fmt_pct(valid_rate), help_text='有效通话 ÷ 接通数（接通后真正聊起来的比例）', tone='accent')
    show_metric(cols[3], '七天回登率', fmt_pct(relogin_rate), help_text='名单里七天内有登入的人数 ÷ 名单总数', tone='accent')
    show_metric(cols[4], '召回充值总额', fmt_num(sum_rech), help_text='名单会员在拨打后（复查口径）的充值金额合计',
                tone=tone_by_sign(sum_rech))

    with st.container(border=True):
        section_header('召回漏斗', '从名单一路漏到接通、有效通话；七天回登与召回充值另算（含未接通但自行回来的会员）。')
        fig = go.Figure(go.Funnel(
            y=['名单/已播', '接通', '有效通话'],
            x=[n_dial, n_conn, n_valid],
            textinfo='value+percent initial',
            marker={'color': [BLUE, PURPLE, GREEN]},
            connector={'line': {'color': 'rgba(150,170,210,0.4)'}},
        ))
        fig.update_layout(height=320, template=TEMPLATE, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width='stretch')

    section_header(f'{label}召回小结', '可直接用于月度汇报。')
    top = df.iloc[0]

    def _hl(text, color=None):
        style = f'color:{color};font-weight:600;' if color else 'font-weight:600;'
        return f'<span style="{style}">{escape(str(text))}</span>'

    line1 = (f'本月对 {_hl(fmt_num(n_list))} 名会员进行电访，接通 {_hl(fmt_num(n_conn))} 通'
             f'（接通率 {_hl(fmt_pct(conn_rate))}），其中有效通话 {_hl(fmt_num(n_valid))} 通。')
    line2 = (f'名单中七天内回登 {_hl(fmt_num(n_relogin))} 人（回登率 {_hl(fmt_pct(relogin_rate))}），'
             f'带回充值 {_hl(fmt_num(sum_rech), GREEN)}。')
    line3 = f'召回充值表现最佳：{_hl(top["专员"])}，{_hl(fmt_num(top["召回充值"]), GREEN)}。'
    st.markdown(
        '<div class="hero-card" style="padding:1.1rem 1.4rem;line-height:2.05;">'
        f'<div>{line1}</div>'
        f'<div>{line2}</div>'
        f'<div style="margin-top:0.35rem;color:#9fb0d0;">{line3}</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    section_header('各专员表现', '按召回充值排序。接通率反映触达能力，召回充值反映实际召回成效。')
    cc = st.columns(2)
    with cc[0]:
        with st.container(border=True):
            figc = go.Figure(go.Bar(
                x=df['专员'], y=df['召回充值'], marker_color=GREEN,
                text=[fmt_num(v) for v in df['召回充值']], textposition='outside',
            ))
            figc.update_layout(height=340, template=TEMPLATE, showlegend=False,
                               title='召回充值', yaxis_title='', margin=dict(t=40))
            st.plotly_chart(figc, width='stretch')
    with cc[1]:
        with st.container(border=True):
            rate = (df['接通数'] / df['已播数'].replace(0, pd.NA)).fillna(0)
            figr = go.Figure(go.Bar(
                x=df['专员'], y=rate, marker_color=BLUE,
                text=[fmt_pct(v) for v in rate], textposition='outside',
            ))
            figr.update_layout(height=340, template=TEMPLATE, showlegend=False,
                               title='接通率', yaxis_tickformat='.0%', margin=dict(t=40))
            st.plotly_chart(figr, width='stretch')

    with st.container(border=True):
        disp = df.copy()
        disp['接通率'] = (disp['接通数'] / disp['已播数'].replace(0, pd.NA)).fillna(0).map(lambda v: f'{v*100:.1f}%')
        disp['召回充值'] = disp['召回充值'].map(lambda v: f'{v:,.0f}')
        disp = disp[['专员', '名单数', '已播数', '接通数', '有效通话', '接通率', '七天回登', '申请彩金', '召回充值']]
        st.dataframe(disp, width='stretch', hide_index=True)
    st.caption('数据来源：上传的「撥打紀錄總表」。召回充值/七天回登为「复查」口径，反映名单会员拨打后一段时间内的实际行为。')


def _winback_compare_view(months: List[Tuple[str, pd.DataFrame]]):
    """多月对比：months = [(label, df), ...]，按月份排好序。"""
    comp = pd.DataFrame([{'月份': lab, **_winback_agg(d)} for lab, d in months])

    section_header('月度对比', '各月关键指标趋势。要看某个月的明细，用下面的选单切换。')
    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            f1 = go.Figure(go.Bar(
                x=comp['月份'], y=comp['召回充值'], marker_color=GREEN,
                text=[fmt_num(v) for v in comp['召回充值']], textposition='outside',
            ))
            f1.update_layout(height=340, template=TEMPLATE, showlegend=False,
                             title='召回充值', margin=dict(t=40))
            st.plotly_chart(f1, width='stretch')
    with c2:
        with st.container(border=True):
            f2 = go.Figure()
            f2.add_trace(go.Scatter(x=comp['月份'], y=comp['接通率'], mode='lines+markers',
                                    name='接通率', line={'color': BLUE, 'width': 2.5}))
            f2.add_trace(go.Scatter(x=comp['月份'], y=comp['七天回登率'], mode='lines+markers',
                                    name='七天回登率', line={'color': PURPLE, 'width': 2.5}))
            f2.update_layout(height=340, template=TEMPLATE, title='接通率 / 七天回登率',
                             yaxis_tickformat='.0%', margin=dict(t=40),
                             legend=dict(orientation='h', y=1.12))
            st.plotly_chart(f2, width='stretch')

    show = comp.copy()
    for c in ['接通率', '有效通话率', '七天回登率']:
        show[c] = show[c].map(lambda v: f'{v*100:.1f}%')
    show['召回充值'] = show['召回充值'].map(lambda v: f'{v:,.0f}')
    show = show[['月份', '名单数', '接通数', '接通率', '有效通话率', '七天回登', '七天回登率', '召回充值']]
    st.dataframe(show, width='stretch', hide_index=True)

    if len(months) >= 2:
        cur, prv = comp.iloc[-1], comp.iloc[-2]
        def _delta(cur_v, prv_v, money=False):
            d = cur_v - prv_v
            arrow = '▲' if d >= 0 else '▼'
            val = fmt_num(abs(d)) if money else f'{abs(d)*100:.1f} 个百分点'
            return f'{arrow} {val}'
        st.markdown(
            '<div class="hero-card" style="padding:1.1rem 1.4rem;line-height:2.05;">'
            f'<div>对比上月（{escape(prv["月份"])} → {escape(cur["月份"])}）：'
            f'召回充值 <span style="color:{GREEN};font-weight:600;">{_delta(cur["召回充值"], prv["召回充值"], money=True)}</span>'
            f'（{fmt_num(prv["召回充值"])} → {fmt_num(cur["召回充值"])}）；'
            f'接通率 {_delta(cur["接通率"], prv["接通率"])}；'
            f'七天回登率 {_delta(cur["七天回登率"], prv["七天回登率"])}。</div>'
            '</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════
# 新注册分析 — 从数据库(raw_member_report)读，看新注册从哪来 / 谁带的 / 质量
# 统一入口：上传一律走「数据上传」页；这页只「读数据库 + 出分析」(跟其他分析页一致)。
# 跨月快照按会员账号去重(取最新)，再按注册时间筛 + 派生 域名/邀请码/首存。
# ══════════════════════════════════════════════════════════════

def _nm_domain(u):
    if pd.isna(u):
        return '(未记录)'
    u = str(u)
    m = re.search(r'https?://([^/:]+)', u)
    if m:
        return m.group(1)
    if u.startswith('null'):
        return '(仅邀请码)'
    return u[:40] if u.strip() else '(未记录)'


def _nm_icode(u):
    if pd.isna(u):
        return None
    # 容错真实数据里的手误变体：i_code= / i_code- / i code- / i%20code- / r_code= 等
    s = str(u).replace('%20', ' ')
    m = re.search(r'[ir][_ ]?code[=\-](\d+)', s)
    return m.group(1) if m else None


def _nm_rank(fdf, col, label):
    if col not in fdf.columns:
        return
    full = (fdf.groupby(col)
            .agg(注册数=('会员账号', 'size'), 有首存=('有首存', 'sum'), 首存额=('首存额', 'sum'))
            .sort_values('注册数', ascending=False))
    total_n = int(full['注册数'].sum())
    n_groups = len(full)
    g = full.head(20).reset_index().rename(columns={'代理n': '代理'})
    g['未充值率'] = ((g['注册数'] - g['有首存']) / g['注册数'] * 100).round(0).astype(int).astype(str) + '%'
    g['首存额'] = g['首存额'].round(0).astype(int)
    st.markdown(f'**{label}**')
    st.dataframe(g, use_container_width=True, hide_index=True)
    if n_groups > 20:
        shown = int(g['注册数'].sum())
        st.caption(f'共 {n_groups} 个、{total_n} 个注册；上表为前 20 名（{shown} 个），'
                   f'其余 {n_groups - 20} 个合计 {total_n - shown} 个（长尾，未列出）。')
    else:
        st.caption(f'共 {n_groups} 个、{total_n} 个注册（已全部列出）。')


def _nm_fill_agent(s):
    """代理列空值统一成「(直客/无代理)」——给分组 / merge 当稳定 key 用。"""
    return s.astype(object).where(s.notna(), '(直客/无代理)').replace('', '(直客/无代理)')


# 这些字段是「逐月可加总」的：每月快照各记当月值，算 cohort 终身净值要跨快照求和，不能取单月。
# (首存金额是一次性属性，不在此列——取去重后单行即可。)
_NM_ADDITIVE = {'公司收入': '净收入', '有效投注额': '有效投注', '红利': '红利n', '返水': '返水n'}


def _nm_prepare(raw: pd.DataFrame) -> pd.DataFrame:
    """从 raw_member_report 准备新注册分析用 df：跨月快照按「会员账号+代理」去重(取最新) + 派生字段。
    去重含代理：同一个账号名挂在不同代理底下=不同的人，不能只按账号合并(跟「代理×会员明细」口径一致)。
    另外把逐月可加总字段(公司收入/有效投注/红利/返水)跨快照求和，挂成 cohort 终身净值列。"""
    df = raw.copy()
    if '会员账号' not in df.columns:
        return df.iloc[0:0]
    dedup_keys = ['会员账号', '代理'] if '代理' in df.columns else ['会员账号']
    if '_snapshot_month' in df.columns:
        df['__sm'] = df['_snapshot_month'].astype(str)
        df = (df.sort_values('__sm')
              .drop_duplicates(subset=dedup_keys, keep='last')
              .drop(columns='__sm'))
    else:
        df = df.drop_duplicates(subset=dedup_keys, keep='first')
    df['注册时间'] = to_datetime_safe(df['注册时间'])
    df = df[df['注册时间'].notna()].copy()
    df['注册日'] = df['注册时间'].dt.strftime('%Y-%m-%d')
    df['首存额'] = pd.to_numeric(df.get('首存金额'), errors='coerce').fillna(0)
    df['有首存'] = df['首存额'] > 0
    df['首投额'] = pd.to_numeric(df.get('首投金额'), errors='coerce').fillna(0)
    df['有首投'] = df['首投额'] > 0
    if '首存时间' in df.columns:
        fdt = to_datetime_safe(df['首存时间'])
        ttf = (fdt - df['注册时间']).dt.total_seconds() / 3600.0
        df['TTF小时'] = ttf.where(df['有首存'] & ttf.ge(0))  # 注册→首存隔多少小时(只算有首存且非负)
    else:
        df['TTF小时'] = pd.NA
    agent_col = df['代理'] if '代理' in df.columns else pd.Series([None] * len(df), index=df.index)
    df['代理n'] = _nm_fill_agent(agent_col)
    url_col = df['注册网址'] if '注册网址' in df.columns else pd.Series([None] * len(df), index=df.index)
    df['域名'] = url_col.map(_nm_domain)
    df['邀请码'] = url_col.map(_nm_icode)

    # ── cohort 终身净值：从原始全量(未去重)按「会员账号+代理」跨快照求和 ──
    # 先按「会员账号+代理+快照月」去重(防同月重复导入被重复计)，再跨月相加，才是这批人到今天为止的真实净值。
    present = {k: v for k, v in _NM_ADDITIVE.items() if k in raw.columns}
    if present:
        base = raw.copy()
        base['代理n'] = _nm_fill_agent(base['代理']) if '代理' in base.columns else '(直客/无代理)'
        if '_snapshot_month' in base.columns:
            sort_col = '_imported_at' if '_imported_at' in base.columns else '_snapshot_month'
            base = (base.sort_values(sort_col)
                    .drop_duplicates(subset=['会员账号', '代理n', '_snapshot_month'], keep='last'))
        for src in present:
            base[src] = pd.to_numeric(base[src], errors='coerce').fillna(0)
        agg = (base.groupby(['会员账号', '代理n'], as_index=False)[list(present)]
               .sum().rename(columns=present))
        df = df.merge(agg, on=['会员账号', '代理n'], how='left')
        for v in present.values():
            df[v] = pd.to_numeric(df[v], errors='coerce').fillna(0)
    return df


def _nm_date_filter(df):
    """这页专用日期筛选：快捷预设(全部/本月/上月/近7天/近30天)或自订日期 + 实时显示。
    用单选(预设 或 自订)，不会出现「月份盖过日期」那种撞。预设相对资料里最新日期算，保证落在数据范围内。"""
    import datetime as _dt
    d = df[df['注册时间'].notna()].copy()
    if d.empty:
        st.warning('没有可用的注册时间。')
        return d
    min_d = d['注册时间'].min().date()
    max_d = d['注册时间'].max().date()
    first_this = max_d.replace(day=1)
    prev_last = first_this - _dt.timedelta(days=1)
    prev_first = prev_last.replace(day=1)
    presets = {
        '全部': (min_d, max_d),
        '本月': (max(min_d, first_this), max_d),
        '上月': (max(min_d, prev_first), min(max_d, prev_last)),
        '近7天': (max(min_d, max_d - _dt.timedelta(days=6)), max_d),
        '近30天': (max(min_d, max_d - _dt.timedelta(days=29)), max_d),
    }
    pick = st.radio('快速选择', list(presets.keys()) + ['自订日期'],
                    horizontal=True, key='nm_dpick')
    if pick == '自订日期':
        c1, c2 = st.columns(2)
        with c1:
            start = st.date_input('开始日期', value=min_d, min_value=min_d, max_value=max_d, key='nm_start')
        with c2:
            end = st.date_input('结束日期', value=max_d, min_value=min_d, max_value=max_d, key='nm_end')
    else:
        start, end = presets[pick]
    if start > end:
        start, end = end, start
    s_ts = pd.Timestamp(start)
    e_ts = pd.Timestamp(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    out = d[(d['注册时间'] >= s_ts) & (d['注册时间'] <= e_ts)].copy()
    st.caption(f'📅 当前显示 {start} ~ {end}，共 {len(out)} 个新注册（已按 会员账号+代理 去重）。改上面就立刻跟着变。')
    return out


_NM_QUALITY_COLS = ['有首存', '首存额', '净收入', '有效投注', '红利n', '返水n']


def _nm_q_cols(df):
    """质量指标需要、且当前 df 里确实存在的列。"""
    return [c for c in _NM_QUALITY_COLS if c in df.columns]


def _nm_int_display(tbl):
    """把整数性质的列转成 Int64，避免 st.dataframe 显示成 10.0 这种带小数点的。"""
    for c in ['注册数', '未充值率%', '人均首存', '人均净收入', '净收入合计']:
        if c in tbl.columns:
            tbl[c] = tbl[c].round().astype('Int64')
    return tbl


def _nm_group_quality(g):
    """对一组新注册算质量指标：转化率 / 未充值率 / 人均首存 / 人均净收入 / 红利套利比值。
    净收入等是去重后挂上的 cohort 终身净值列(跨月累计·公司角度)，直接对当前这批人求和。"""
    n = len(g)
    has = int(g['有首存'].sum())
    fd = float(g['首存额'].sum())
    row = {
        '注册数': n,
        '转化率%': round(has / n * 100, 1) if n else 0.0,
        '未充值率%': round((n - has) / n * 100) if n else 0,
        '人均首存': round(fd / n) if n else 0,
    }
    if '净收入' in g.columns:
        net = float(g['净收入'].sum())
        row['人均净收入'] = round(net / n) if n else 0
        row['净收入合计'] = round(net)
    if {'红利n', '返水n', '有效投注'}.issubset(g.columns):
        to = float(g['有效投注'].sum())
        give = float(g['红利n'].sum()) + float(g['返水n'].sum())
        if to > 0:
            row['套利比值%'] = round(give / to * 100, 1)
        else:
            row['套利比值%'] = (999.0 if give > 0 else None)
    return pd.Series(row)


def _nm_value_by_source(fdf):
    """① 直客 vs 代理来源 价值对比——不只比人数，比转化 / 人均首存 / 人均净收入。"""
    if '用户来源' not in fdf.columns:
        return
    section_header('① 直客 vs 代理来源 — 价值对比',
                   '不只比人数：比转化率、人均首存、人均净收入。直接回答「直客掉到底多痛」。')
    src = fdf.copy()
    src['用户来源'] = src['用户来源'].fillna('(未记录)').astype(str)
    tbl = (src.groupby('用户来源')[_nm_q_cols(src)].apply(_nm_group_quality)
           .reset_index())
    order = {'直客': 0, '普代下线': 1, '官代下线': 2}
    tbl['_o'] = tbl['用户来源'].map(order).fillna(9)
    tbl = tbl.sort_values('_o').drop(columns='_o')
    tbl = _nm_int_display(tbl)
    st.dataframe(tbl, use_container_width=True, hide_index=True)
    direct = src[src['用户来源'] == '直客']
    agent = src[src['用户来源'] != '直客']
    if len(direct) and len(agent) and '净收入' in src.columns:
        d_arpu = direct['净收入'].sum() / len(direct)
        a_arpu = agent['净收入'].sum() / len(agent)
        d_conv = direct['有首存'].mean() * 100
        a_conv = agent['有首存'].mean() * 100
        verdict = ('直客更值钱 → 直客下滑的影响大于人数本身。'
                   if d_arpu > a_arpu else
                   '代理来源人均更高 → 直客量下滑相对可控，但仍要看结构。')
        st.caption(f'📌 直客 {len(direct)} 人、转化 {d_conv:.0f}%、人均净收入 {d_arpu:,.0f}；'
                   f'代理来源 {len(agent)} 人、转化 {a_conv:.0f}%、人均净收入 {a_arpu:,.0f}。{verdict}')
    st.caption('💡 净收入＝公司角度(已扣红利)，负数代表这批人到今天为止公司是亏的；跨月累计。'
               '「直客」指用户来源=直客，代理来源含普代下线＋官代下线。')


def _nm_quality_board(fdf):
    """② 代理质量分层榜——量×质一起看，自动标「疑似刷量 / 红利套利」。"""
    if '代理n' not in fdf.columns:
        return
    section_header('② 代理质量分层榜',
                   '不只看量：转化率＋人均首存＋人均净收入＋未充值率一起看，'
                   '一眼分辨「高量低质 / 疑似刷量」和「量小但真值钱」。')
    vol = st.slider('最少注册数（滤掉长尾，只看够量的代理）', 1, 50, 5, key='nm_q_vol')
    tbl = (fdf.groupby('代理n')[_nm_q_cols(fdf)].apply(_nm_group_quality)
           .reset_index().rename(columns={'代理n': '代理'}))
    tbl = tbl[tbl['注册数'] >= vol].sort_values('注册数', ascending=False)
    if tbl.empty:
        st.caption(f'没有注册数 ≥ {vol} 的代理，把门槛调低一点。')
        return

    def _flag(r):
        marks = []
        if r['注册数'] >= max(vol, 10) and r['未充值率%'] >= 80:
            marks.append('🚩刷量嫌疑')
        if '套利比值%' in r and pd.notna(r['套利比值%']) and r['套利比值%'] >= 100:
            marks.append('⚠️红利套利')
        return ' '.join(marks) if marks else '✅'

    tbl['风险'] = tbl.apply(_flag, axis=1)
    n_flag = int((tbl['风险'] != '✅').sum())
    tbl = _nm_int_display(tbl.head(40))
    st.dataframe(tbl, use_container_width=True, hide_index=True)
    st.caption(
        f'共标出 {n_flag} 个风险代理。🚩刷量嫌疑＝注册量够大(≥10)但未充值率≥80%'
        '(像 qwe8825252 注册46、未充值93%)；⚠️红利套利＝(红利+返水)≥有效投注额(领了优惠几乎不打)。'
        '人均净收入＝这批人到今天为止给公司带来的净收入÷人数(跨月累计·公司角度·负数代表公司亏)。')


def _nm_funnel_ttf(fdf):
    """③ 转化漏斗(注册→首存→首投) + 首存速度 TTF。"""
    section_header('③ 转化漏斗 + 首存速度（TTF）',
                   '注册→首存→首投 三段漏斗，看哪一段漏最多；首存速度比注册量更早预警转化变差。')
    n = len(fdf)
    if n == 0:
        return
    has_fd = int(fdf['有首存'].sum())
    has_fb = int(fdf['有首投'].sum()) if '有首投' in fdf.columns else 0
    # 只画 注册→首存：首存→首投 几乎恒等于 100%(充了钱的人基本都会下注)，当指标没信息量，不单列。
    fig = go.Figure(go.Funnel(
        y=['注册', '首存（充值）'], x=[n, has_fd],
        textposition='inside', textinfo='value+percent initial',
        marker={'color': ['#60a5fa', '#2dd4a7']}))
    fig.update_layout(template='plotly_dark', height=220,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, use_container_width=True)

    ttf = fdf['TTF小时'].dropna() if 'TTF小时' in fdf.columns else pd.Series([], dtype=float)
    reg2fd = has_fd / n
    fd_med = float(fdf.loc[fdf['有首存'], '首存额'].median()) if has_fd else 0
    mc = st.columns(4)
    show_metric(mc[0], '注册→首存 转化', fmt_pct(reg2fd), tone='warn' if reg2fd < 0.5 else 'good',
                help_text=f'有首存人数 ÷ 注册数 ＝ {has_fd} ÷ {n}（这批注册里有多少人真的充了第一笔）')
    show_metric(mc[1], '首存中位额', fmt_num(round(fd_med)) if has_fd else 'N/A',
                help_text='有首存的人，首存金额取中位数（看这批人首存大不大；中位比平均更抗大户拉高）')
    show_metric(mc[2], '首存速度 TTF 中位', f'{ttf.median():.1f} 小时' if len(ttf) else 'N/A',
                help_text='有首存的人，(首存时间 － 注册时间) 取中位数；越短越好')
    fast = (ttf < 1).mean() * 100 if len(ttf) else 0
    show_metric(mc[3], '1 小时内首存占比', f'{fast:.0f}%' if len(ttf) else 'N/A',
                help_text='有首存的人里，注册后 1 小时内就完成首存的占比')

    drop_reg = n - has_fd
    fb_note = ('；充了钱的人几乎都会下注（首存→首投 '
               f'{has_fb / has_fd * 100:.0f}%），所以这步不另列' if has_fd else '')
    st.caption(f'📌 流失全卡在「注册→首存」：{n} 人注册、只有 {has_fd} 人充值，漏掉 {drop_reg} 人'
               f'（{drop_reg / n * 100:.0f}%）{fb_note}。瓶颈是拉新质量 / 首存引导，不是下注意愿。')

    if len(ttf):
        labels = ['<1小时', '1-24小时', '1-7天', '>7天']
        dist = (pd.cut(ttf, bins=[0, 1, 24, 168, float('inf')], labels=labels, right=False)
                .value_counts().reindex(labels).fillna(0).astype(int))
        fig2 = go.Figure(go.Bar(x=labels, y=dist.values, marker_color='#2dd4a7',
                                text=dist.values, textposition='outside'))
        fig2.update_layout(template='plotly_dark', height=240,
                           paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                           margin=dict(l=10, r=10, t=40, b=10),
                           title='首存速度分布（注册到首存隔多久）')
        st.plotly_chart(fig2, use_container_width=True)

    if len(ttf) and '用户来源' in fdf.columns:
        seg = fdf[fdf['TTF小时'].notna()].copy()
        seg['用户来源'] = seg['用户来源'].fillna('(未记录)').astype(str)
        g = seg.groupby('用户来源')['TTF小时'].agg(['median', 'count']).reset_index()
        g.columns = ['用户来源', 'TTF中位(小时)', '首存人数']
        g['TTF中位(小时)'] = g['TTF中位(小时)'].round(1)
        g['首存人数'] = g['首存人数'].astype(int)
        st.caption('各来源首存速度对比（中位 TTF 越短代表转化越顺）：')
        st.dataframe(g.sort_values('首存人数', ascending=False), use_container_width=True, hide_index=True)


def _nm_cohort_pl(fdf):
    """④ Cohort 损益走势：按注册月分组，看每批人到今天为止累计净收入（公司角度）。"""
    section_header('④ Cohort 损益走势（按注册月）',
                   '把新注册按「注册月份」分组，看每一批人到今天为止累计给公司带来多少净收入——'
                   '比注册量更接近「这批客到底值不值」。')
    if '净收入' not in fdf.columns or fdf.empty:
        st.caption('当前数据缺少净收入字段（公司收入），无法算 cohort 损益。')
        return
    df = fdf.copy()
    df['注册月'] = df['注册时间'].dt.strftime('%Y-%m')
    tbl = (df.groupby('注册月')[_nm_q_cols(df)].apply(_nm_group_quality)
           .reset_index().sort_values('注册月'))
    keep = ['注册月', '注册数', '转化率%', '人均首存', '净收入合计', '人均净收入']
    tbl = tbl[[c for c in keep if c in tbl.columns]]
    if tbl.empty:
        st.caption('当前筛选下没有可分组的注册月。')
        return

    colors = ['#2dd4a7' if v >= 0 else '#fb7185' for v in tbl['净收入合计']]
    fig = go.Figure(go.Bar(x=tbl['注册月'], y=tbl['净收入合计'], marker_color=colors))
    fig.update_layout(template='plotly_dark', height=300,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=10, r=10, t=40, b=10),
                      title='各注册月 cohort 到今天的累计净收入（公司角度·绿赚红亏）')
    st.plotly_chart(fig, use_container_width=True)

    fig2 = go.Figure(go.Scatter(x=tbl['注册月'], y=tbl['人均净收入'],
                                mode='lines+markers', line=dict(color='#60a5fa')))
    fig2.update_layout(template='plotly_dark', height=260,
                       paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                       margin=dict(l=10, r=10, t=40, b=10),
                       title='各注册月 cohort 人均净收入')
    st.plotly_chart(fig2, use_container_width=True)

    st.dataframe(_nm_int_display(tbl), use_container_width=True, hide_index=True)
    st.caption('⚠️ 看这页要扣掉「账龄」因素：老月份人均净收入偏高，是因为他们入金时间长、留下来的多是大户；'
               '当月/上月注册的人账龄短，累计自然低，别直接拿绝对值跟老月份比。'
               '真正有意义的是——①近几个月之间的对比 ②哪个月的 cohort 到今天还是「负的」（这批人累计还在亏，要查当月拉来的是什么客）。')


def render_new_member_analysis():
    hero('新注册分析',
         '直接从数据库读你已上传的会员数据，看新注册从哪来、谁带的、质量如何（按会员账号+代理去重）。'
         '要分析新数据？先把会员报表丢进顶部「数据上传」页存一下，这页就读得到。',
         basis='会员报表（口径为注册数的非代理部分·后台导出上传）',
         detail=(
             '**分析范围**：新注册来源、带来的代理、渠道／域名、质量（有无首存）与每日走势。\n\n'
             '**数据来源**：会员报表（报表中心→会员报表，须按注册时间、完整日期、全部页数导出）。\n\n'
             '**计算口径**：去重按会员账号＋代理；平台「注册数」对应会员报表的「非代理」部分。\n\n'
             '**更新方式**：手动上传（上传时会自动做完整度校验）。完整对照见「数据说明」页。'
         ))

    try:
        raw = load_table('raw_member_report')
    except Exception as e:
        st.error(f'读数据库失败：{str(e)[:120]}')
        return
    if raw is None or raw.empty or '会员账号' not in raw.columns:
        st.info('📭 数据库里还没有会员数据。\n\n'
                '先去顶部「数据上传」页，把后台下载的「会员报表」（密码 zip / 分好几份都行）拖进去存一下，'
                '再回这页，就会自动读出来分析——不用在这里上传。')
        return

    df = _nm_prepare(raw)
    if df.empty:
        st.warning('会员数据里没有可用的「注册时间」，无法做新注册分析。')
        return

    total_all = len(df)

    # ── 筛选器 ──
    section_header('筛选', '先圈范围，下面所有总览 / 走势 / 排行 / 交叉都跟着这里走。')
    fdf = _nm_date_filter(df)
    if fdf.empty:
        st.warning('当前日期范围内没有新注册。')
        return
    n_in_range = len(fdf)  # 日期范围内总数（下面的来源/渠道/域名/首存筛选会从这个数往下减）
    c1, c2 = st.columns([2, 1])
    with c1:
        if '用户来源' in fdf.columns:
            fdf = apply_multiselect(fdf, '用户来源', '用户来源（普代下线 / 直客 / 官代下线）',
                                    'nm_src2', options_df=df, auto_include_new=True)
    with c2:
        dep = st.radio('首存', ['全部', '只看有首存', '只看未充值'], horizontal=True, key='nm_dep')
    if dep == '只看有首存':
        fdf = fdf[fdf['有首存']].copy()
    elif dep == '只看未充值':
        fdf = fdf[~fdf['有首存']].copy()
    c3, c4 = st.columns(2)
    with c3:
        if '注册来源' in fdf.columns:
            fdf = apply_multiselect(fdf, '注册来源', '渠道 / 注册来源（多选）',
                                    'nm_chan2', options_df=df, auto_include_new=True)
    with c4:
        fdf = apply_multiselect(fdf, '域名', '域名（多选）', 'nm_dom2',
                                options_df=df, auto_include_new=True)

    if fdf.empty:
        st.warning('当前筛选下没有数据，放宽一下条件。')
        return

    n = len(fdf)
    has = int(fdf['有首存'].sum())
    fd_sum = float(fdf['首存额'].sum())
    n_zhike = int((fdf['代理n'] == '(直客/无代理)').sum())
    n_agent_mem = n - n_zhike
    n_agents = int(fdf[fdf['代理n'] != '(直客/无代理)']['代理n'].nunique())
    uniq_acct = int(fdf['会员账号'].nunique())

    section_header('总览', f'当前筛选：{n} 个新注册（本次上传共 {total_all} 个）')
    if n < n_in_range:
        st.caption(f'⚠️ 日期范围内本有 {n_in_range} 个，下面的「用户来源 / 渠道 / 域名 / 首存」筛选又减掉了 '
                   f'{n_in_range - n} 个，所以总览显示 {n} 个。想看完整 {n_in_range} 个，把这几个筛选都留在「全选」。')
    st.caption('💡 口径提醒：此处「新注册」按『会员账号＋代理』计算（同一账号挂在多个代理下会分别计入），'
               '因此可能比运营日报 / Daybook 的「注册数」（按唯一账号去重）略高。'
               '想知道实际新增了多少「人」→ 以运营日报的去重数为准；想看各代理 / 渠道各拉来多少 → 看这页。')
    mc = st.columns(6)
    show_metric(mc[0], '新注册（账号+代理）', fmt_num(n),
                help_text='按「会员账号＋代理」算：同一账号挂在多个代理下会分别计入，'
                          '用来把人头归到对应代理 / 渠道。')
    show_metric(mc[1], '唯一账号（人）', fmt_num(uniq_acct),
                help_text='去重到「人」（不分代理）。对应平台报表 / Daybook 的注册数。'
                          f'比左边少 {n - uniq_acct}，差额=同一账号挂在多个代理下。')
    show_metric(mc[2], '有首存（充值）', f'{has}（{has / n * 100:.0f}%）', tone='good')
    show_metric(mc[3], '未充值（注册了没充值）', f'{n - has}（{(n - has) / n * 100:.0f}%）',
                tone='warn' if n and (n - has) / n > 0.5 else None)
    show_metric(mc[4], '首存总额', fmt_num(round(fd_sum)))
    show_metric(mc[5], '代理带 / 直客', f'{n_agent_mem} / {n_zhike}',
                help_text=f'有产出代理 {n_agents} 个')

    # ── 每日走势 ──
    section_header('每日新注册走势', '绿=有首存、红=未充值；某天明显高出就是暴增日。')
    daily = (fdf.groupby('注册日')
             .agg(新注册=('会员账号', 'size'), 有首存=('有首存', 'sum')).reset_index())
    daily['未充值'] = daily['新注册'] - daily['有首存']
    fig = go.Figure()
    fig.add_bar(x=daily['注册日'], y=daily['有首存'], name='有首存', marker_color='#2dd4a7')
    fig.add_bar(x=daily['注册日'], y=daily['未充值'], name='未充值', marker_color='#fb7185')
    fig.update_layout(barmode='stack', template='plotly_dark', height=320,
                      paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                      margin=dict(l=10, r=10, t=10, b=10),
                      legend=dict(orientation='h', y=1.1, x=0))
    st.plotly_chart(fig, use_container_width=True)

    # ── 获客质量诊断：直客vs代理 价值对比 + 代理质量分层榜 ──
    _nm_value_by_source(fdf)
    _nm_quality_board(fdf)
    _nm_funnel_ttf(fdf)
    _nm_cohort_pl(fdf)

    # ── 其他来源排行 ──
    section_header('其他来源排行 — 邀请码 / 渠道 / 域名',
                   '各看注册数、有首存、首存额、未充值率。代理质量看上面的「分层榜」。')
    rc1, rc2 = st.columns(2)
    with rc1:
        ic = fdf[fdf['邀请码'].notna()].copy()
        if not ic.empty:
            full = (ic.groupby('邀请码')
                    .agg(注册数=('会员账号', 'size'), 有首存=('有首存', 'sum'), 首存额=('首存额', 'sum'),
                         对应代理=('代理n', lambda s: s.value_counts().index[0]))
                    .sort_values('注册数', ascending=False))
            ic_total = int(full['注册数'].sum())
            ic_groups = len(full)
            g = full.head(20).reset_index()
            g['未充值率'] = ((g['注册数'] - g['有首存']) / g['注册数'] * 100).round(0).astype(int).astype(str) + '%'
            g['首存额'] = g['首存额'].round(0).astype(int)
            st.markdown('**邀请码排行（含对应代理）**')
            st.dataframe(g, use_container_width=True, hide_index=True)
            if ic_groups > 20:
                shown = int(g['注册数'].sum())
                st.caption(f'共 {ic_groups} 个邀请码、{ic_total} 个带码注册；上表为前 20 名（{shown} 个），'
                           f'其余 {ic_groups - 20} 个合计 {ic_total - shown} 个（长尾，未列出）。'
                           f'另有未带邀请码的注册不在此表。')
            else:
                st.caption(f'共 {ic_groups} 个邀请码、{ic_total} 个带码注册（已全部列出）。另有未带邀请码的注册不在此表。')
        else:
            st.markdown('**邀请码排行**')
            st.caption('这批数据里没有带邀请码的注册。')
    with rc2:
        _nm_rank(fdf, '注册来源', '渠道 / 注册来源排行')
    _nm_rank(fdf, '域名', '域名排行')

    # ── 交叉分析 ──
    section_header('交叉分析', '任选两个维度交叉，看「哪个代理走哪个渠道」「哪个域名质量好」这种。')
    dim_map = {'代理': '代理n', '渠道': '注册来源', '域名': '域名',
               '用户来源': '用户来源', '地区': '地区名称', '邀请码': '邀请码'}
    dim_map = {k: v for k, v in dim_map.items() if v in fdf.columns}
    dim_names = list(dim_map.keys())
    xc1, xc2, xc3 = st.columns(3)
    with xc1:
        dimA = st.selectbox('维度 A（行）', dim_names, index=0, key='nm_xa')
    with xc2:
        dimB = st.selectbox('维度 B（列）', dim_names, index=min(1, len(dim_names) - 1), key='nm_xb')
    with xc3:
        xmetric = st.selectbox('交叉看什么', ['注册数', '有首存数', '首存额'], key='nm_xm')
    ca, cb = dim_map[dimA], dim_map[dimB]
    base = fdf.copy()
    ra = base[ca].fillna('(空)').astype(str)
    rb = base[cb].fillna('(空)').astype(str)
    if xmetric == '注册数':
        piv = pd.crosstab(ra, rb)
    elif xmetric == '有首存数':
        piv = pd.crosstab(ra, rb, values=base['有首存'].astype(int), aggfunc='sum').fillna(0).astype(int)
    else:
        piv = pd.crosstab(ra, rb, values=base['首存额'], aggfunc='sum').fillna(0).round(0).astype(int)
    top_rows = piv.sum(axis=1).sort_values(ascending=False).head(15).index
    top_cols = piv.sum(axis=0).sort_values(ascending=False).head(15).index
    piv = piv.loc[top_rows, top_cols]
    st.caption(f'{dimA} × {dimB} — {xmetric}（各取前 15，按总量排）')
    st.dataframe(piv, use_container_width=True)

    # ── 查单一代理 ──
    section_header('查单一代理', '输入或选一个代理账号，看他名下新增了哪些会员、多少充值。')
    agent_opts = ['(选一个)'] + fdf[fdf['代理n'] != '(直客/无代理)']['代理n'].value_counts().index.tolist()
    pick = st.selectbox('代理账号', agent_opts, key='nm_pick_agent')
    if pick != '(选一个)':
        sub = fdf[fdf['代理n'] == pick].copy()
        am = st.columns(4)
        show_metric(am[0], '新增会员', fmt_num(len(sub)))
        show_metric(am[1], '有首存', f"{int(sub['有首存'].sum())}", tone='good')
        show_metric(am[2], '未充值', f"{int((~sub['有首存']).sum())}", tone='warn')
        show_metric(am[3], '首存总额', fmt_num(round(float(sub['首存额'].sum()))))
        detail_cols = [c for c in ['会员账号', '注册时间', '首存额', '注册来源', '域名', '邀请码',
                                   '地区名称', '用户来源', 'VIP等级', '会员状态']
                       if c in sub.columns]
        st.dataframe(sub[detail_cols].sort_values('首存额', ascending=False),
                     use_container_width=True, hide_index=True)


def _fin_minutes(df, t_start, t_end):
    s = to_datetime_safe(df[t_start]); e = to_datetime_safe(df[t_end])
    return (e - s).dt.total_seconds() / 60


def _fin_deposit_view(dep):
    if dep is None or dep.empty or '订单状态' not in dep.columns:
        st.info('暂无充值数据。')
        return
    dep, _s, _e, _m = date_range_picker(dep, '完成时间', 'fd_dep')
    if dep.empty:
        st.warning('该日期范围内无充值数据。')
        return
    n = len(dep)
    succ = int((dep['订单状态'] == '存款成功').sum())
    cancel = int((dep['订单状态'] == '已取消').sum())
    wait = None
    if '完成时间' in dep.columns:
        wait = _fin_minutes(dep[dep['订单状态'] == '存款成功'], '订单时间', '完成时间')
        wait = wait[(wait >= 0) & (wait.notna())]
    section_header('充值总览', f'共 {n} 笔（成功率 = 成功÷总；掉单率 = 已取消÷总）')
    c = st.columns(4)
    show_metric(c[0], '充值笔数', fmt_num(n))
    show_metric(c[1], '成功率', fmt_pct(succ / n) if n else 'N/A', tone='good')
    show_metric(c[2], '掉单率（已取消）', fmt_pct(cancel / n) if n else 'N/A',
                tone='warn' if n and cancel / n > 0.15 else None)
    show_metric(c[3], '平均到账（成功单）', f'{wait.mean():.1f} 分' if wait is not None and len(wait) else 'N/A',
                tone='accent', help_text='完成时间 − 订单时间')
    if '支付方式' in dep.columns:
        rows = []
        for pm, g in dep.groupby('支付方式'):
            gn = len(g); gs = int((g['订单状态'] == '存款成功').sum()); gc = int((g['订单状态'] == '已取消').sum())
            gw = _fin_minutes(g[g['订单状态'] == '存款成功'], '订单时间', '完成时间') if '完成时间' in g.columns else None
            gw = gw[(gw >= 0) & (gw.notna())] if gw is not None else None
            rows.append({'支付方式': pm, '笔数': gn, '_succ': gs / gn if gn else 0, '_drop': gc / gn if gn else 0,
                         '平均到账(分)': round(gw.mean(), 1) if gw is not None and len(gw) else 0})
        ch = pd.DataFrame(rows).sort_values('笔数', ascending=False)
        disp = ch.copy()
        disp['成功率'] = disp['_succ'].map(lambda v: f'{v*100:.0f}%')
        disp['掉单率'] = disp['_drop'].map(lambda v: f'{v*100:.0f}%')
        disp = disp[['支付方式', '笔数', '成功率', '掉单率', '平均到账(分)']]
        section_header('分渠道（支付方式）', '成功率低的渠道＝客户在这关掉单。建议主推高成功率渠道 + 优化低成功率渠道的引导。')
        st.dataframe(disp, use_container_width=True, hide_index=True)
        low = ch[(ch['笔数'] >= 20) & (ch['_succ'] < 0.6)].sort_values('笔数', ascending=False)
        if not low.empty:
            st.warning('⚠️ 成功率偏低（<60%、≥20 笔）的渠道：'
                       + '、'.join(f"{r['支付方式']}({r['_succ']*100:.0f}%)" for _, r in low.iterrows())
                       + ' —— 客户多在这几关掉单，建议主推高成功率渠道 + 优化引导话术。')
    if '取消原因' in dep.columns:
        rc = dep[dep['订单状态'] == '已取消']['取消原因'].astype(str).str.strip()
        rc = rc[~rc.str.lower().isin(['nan', 'none', ''])]
        if len(rc):
            section_header('掉单原因 Top')
            vc = rc.value_counts().head(8).reset_index()
            vc.columns = ['取消原因', '笔数']
            st.dataframe(vc, use_container_width=True, hide_index=True)


def _fin_withdraw_view(wd):
    if wd is None or wd.empty or '订单状态' not in wd.columns:
        st.info('暂无提款数据。')
        return
    wd, _s, _e, _m = date_range_picker(wd, '完成时间', 'fd_wd')
    if wd.empty:
        st.warning('该日期范围内无提款数据。')
        return
    n = len(wd)
    succ = int((wd['订单状态'] == '提款成功').sum())
    reject = int((wd['订单状态'] == '审核拒绝').sum())
    fail = int((wd['订单状态'] == '提款失败').sum())
    wait = None
    if '完成时间' in wd.columns:
        wait = _fin_minutes(wd[wd['订单状态'] == '提款成功'], '申请时间', '完成时间')
        wait = wait[(wait >= 0) & (wait.notna())]
    section_header('提款总览', f'共 {n} 笔')
    c = st.columns(4)
    show_metric(c[0], '提款笔数', fmt_num(n))
    show_metric(c[1], '成功率', fmt_pct(succ / n) if n else 'N/A', tone='good')
    show_metric(c[2], '拒绝率（审核拒绝）', fmt_pct(reject / n) if n else 'N/A',
                tone='warn' if n and reject / n > 0.1 else None)
    show_metric(c[3], '平均出款（成功单）', f'{wait.mean():.1f} 分' if wait is not None and len(wait) else 'N/A',
                tone='accent', help_text='完成时间 − 申请时间')
    if wait is not None and len(wait):
        c2 = st.columns(4)
        show_metric(c2[0], '出款中位', f'{wait.median():.1f} 分')
        show_metric(c2[1], '90% 在', f'{wait.quantile(.9):.1f} 分内')
        show_metric(c2[2], '超 1 小时', fmt_num(int((wait > 60).sum())),
                    tone='warn' if int((wait > 60).sum()) else None)
        show_metric(c2[3], '提款失败笔数', fmt_num(fail))
        w2 = wd[wd['订单状态'] == '提款成功'].copy()
        w2['处理分钟'] = _fin_minutes(w2, '申请时间', '完成时间')
        slow = w2[w2['处理分钟'] > 60].sort_values('处理分钟', ascending=False)
        if not slow.empty:
            section_header('出款慢单（>1 小时）', '大额 / 风控长尾，值得逐笔查。')
            cols = [c for c in ['订单号', '会员账号', '会员等级', '订单金额', '申请时间', '完成时间', '处理分钟']
                    if c in slow.columns]
            sd = slow[cols].head(50).copy()
            sd['处理分钟'] = sd['处理分钟'].round(0).astype(int)
            st.dataframe(sd, use_container_width=True, hide_index=True)


def render_finance_channel():
    hero('存取款分析',
         '充值 / 提款各渠道的成功率、掉单率、平均处理时间。数据存数据库、这页从库里读——'
         '要更新去顶部「数据上传」把后台「存款管理 / 提款管理 历史记录」Csv 拖进去（以后接日报机器人自动更新）。',
         basis='存款/提款历史记录订单（程序每日 11:00 自动抓取）',
         detail=(
             '**分析范围**：充值／提款各渠道的成功率、掉单率、平均处理时间与慢单。\n\n'
             '**数据来源**：财务管理→存款管理／提款管理→历史记录。\n\n'
             '**计算口径**：按「完成时间」统计；处理时长＝完成时间−订单时间（存）／申请时间（提）。\n\n'
             '**更新方式**：程序每日 11:00 自动抓取；亦可于「数据上传」页拖入历史记录 Csv 手动补。完整对照见「数据说明」页。'
         ))
    try:
        dep = load_table('raw_finance_deposit')
    except Exception:
        dep = None
    try:
        wd = load_table('raw_finance_withdraw')
    except Exception:
        wd = None
    if (dep is None or dep.empty) and (wd is None or wd.empty):
        st.info('📭 数据库里还没有存取款数据。\n\n去顶部「数据上传」，把后台「存款管理 / 提款管理 → 历史记录」导出的 Csv 拖进去存一下，再回这页就能看。')
        return
    tab1, tab2 = st.tabs(['💰 充值（存款）', '🏧 提款（取款）'])
    with tab1:
        _fin_deposit_view(dep)
    with tab2:
        _fin_withdraw_view(wd)


def render_winback():
    hero('电访召回',
         '会员召回电话的成效。数据存进数据库、这页从库里读——要新增月份，去顶部「数据上传」把那份「撥打紀錄總表」拖进去存一次即可；这页就能单月看明细、多月看趋势。',
         basis='撥打紀錄總表（电访团队提供·上传）',
         detail=(
             '**分析范围**：电访召回漏斗（名单／接通／有效通话／七天回登／召回充值）、各专员表现与月度对比。\n\n'
             '**数据来源**：撥打紀錄總表（非后台报表，由电访团队提供 xlsx）。\n\n'
             '**更新方式**：人工提供后于「数据上传」页上传（按月份刷新）。完整对照见「数据说明」页。'
         ))

    try:
        wb = load_table('raw_winback')
    except Exception:
        wb = None
    stored = []
    if wb is not None and not wb.empty and '月份' in wb.columns:
        for c in ['名单数', '已播数', '接通数', '有效通话', '申请彩金', '七天回登', '召回充值']:
            if c in wb.columns:
                wb[c] = pd.to_numeric(wb[c], errors='coerce').fillna(0)
        for ym, g in wb.groupby('月份'):
            s = str(ym)
            lab = f"{s.split('-')[0]}年{int(s.split('-')[1])}月" if '-' in s and s.split('-')[1].isdigit() else s
            stored.append((s, lab, g.reset_index(drop=True)))
        stored.sort(key=lambda x: x[0])

    with st.expander('📤 临时看一份（只看、不写库）'):
        files = st.file_uploader('上传「撥打紀錄總表」（.xlsx，可多个）', type=['xlsx'],
                                 accept_multiple_files=True, key='winback_upload')
    adhoc = []
    if files:
        for f in files:
            try:
                df, meta = parse_winback_file(f)
                if not df.empty:
                    adhoc.append((_winback_label(meta, f.name), meta.get('month', ''), df))
            except Exception as e:
                st.warning(f'{f.name}：{e}')

    if adhoc:
        adhoc.sort(key=lambda x: x[1] or x[0])
        labeled = [(lab, df) for lab, _ym, df in adhoc]
        st.caption('当前显示：临时上传的档（未写入数据库；要永久存请去「数据上传」页）。')
    elif stored:
        labeled = [(lab, df) for _ym, lab, df in stored]
    else:
        st.info('📭 数据库里还没有电访数据。\n\n'
                '去顶部「数据上传」页，把后台那份「撥打紀錄總表（X月）.xlsx」拖进去存一下，再回这页就能看——'
                '存一次永久留底，不用每次重传。（也可以用上面「临时看一份」先看效果。）')
        return

    if len(labeled) == 1:
        _winback_month_view(labeled[0][1], labeled[0][0])
        return

    _winback_compare_view(labeled)
    section_header('单月明细', '选一个月看完整漏斗与各专员表现。')
    pick = st.selectbox('选择月份', [lab for lab, _ in labeled],
                        index=len(labeled) - 1, key='winback_pick')
    _winback_month_view(dict(labeled)[pick], pick)


# ══════════════════════════════════════════════════════════════
# 数据上传页 — 自助把月度报表写进 BigQuery
# 复用 import_tool 的解析/清洗逻辑；写入走 get_client() 的服务账号。
# 铁律：只「追加」+ 用 _source_file 防重复，绝不覆盖/删除既有数据。
# 注：标准 10 张月报先上（append 安全）；红利/客服对话之后补（需 read-modify-write）。
# ══════════════════════════════════════════════════════════════

def _import_tool():
    """惰性载入同目录的 import_tool（部署时与 dashboard.py 一起进 repo）。"""
    import os
    import sys
    here = os.path.dirname(os.path.abspath(__file__))
    if here not in sys.path:
        sys.path.insert(0, here)
    import import_tool
    return import_tool


def _bq_count_source_file(client, table_name: str, source_file: str) -> int:
    """既有表里这个 _source_file 已写过几行（防同档重复）。表不存在/无此栏 → 0。"""
    sql = (
        f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table_name}` "
        f"WHERE _source_file = @sf"
    )
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter('sf', 'STRING', source_file)
    ])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _identify_by_content(it, tmp_path):
    """档名认不出时，靠栏位内容认报表类型（裸档名 top.xlsx、或重命名过的档都能认）。
    只用各报表「最有辨识度」的栏位组合，避免误判。"""
    try:
        cols = set(str(c).strip() for c in it.read_file(tmp_path).columns)
    except Exception:
        return None

    def has(*xs):
        return all(x in cols for x in xs)
    if has('用户名', '个人输赢') or has('用户名', '杀数'):
        return ('raw_top_report', 'Top报表')
    if has('会员账号', '注册时间', '是否为代理'):
        return ('raw_member_report', '会员报表')
    if has('注册数', '公司输赢', '首存人数'):
        return ('raw_platform_report', '平台报表')
    if has('红利', '返水', '代理佣金'):
        return ('raw_finance_report', '财务报表')
    return None


def _parse_standard_report(it, tmp_path: str, source_file: str, table_override=None):
    """复用 import_tool 解析单个标准月报。返回 (table, display, df_data)；无法识别→(None,None,None)。
    table_override=(table, display)：档名认不出时由内容识别传入。"""
    table, display = it.identify_report_type(source_file)
    if table is None and table_override:
        table, display = table_override
    if table is None:
        return None, None, None
    df = it.read_file(tmp_path)
    summary_mask = df.apply(lambda r: it.is_summary_row(r.values), axis=1)
    df_clean = df[~summary_mask].copy()
    zero_mask = df_clean.apply(
        lambda r: it.is_all_zero_data_row(r.values, df_clean.columns), axis=1)
    df_data = df_clean[~zero_mask].copy()
    # 快照月份（TOP / 会员）
    if table == 'raw_top_report':
        month = it.extract_top_month(source_file)
        if month:
            df_data['_snapshot_month'] = month
            snap = it.month_to_snapshot_date(month)
            if snap:
                df_data['_snapshot_date'] = snap
    elif table == 'raw_member_report':
        month = it.infer_member_snapshot_month(df_data, tmp_path)
        if month:
            df_data['_snapshot_month'] = month
            snap = it.month_to_snapshot_date(month)
            if snap:
                df_data['_snapshot_date'] = snap
    return table, display, df_data


def _append_standard_report(client, df_data, table_name: str, source_file: str) -> int:
    """append 写入（绝不覆盖）。沿用 import_tool 的 schema 容错。"""
    payload = df_data.copy()
    payload['_imported_at'] = pd.Timestamp.now()
    payload['_source_file'] = source_file
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION],
    )
    job = client.load_table_from_dataframe(
        payload, f"{BQ_PREFIX}.{table_name}", job_config=cfg)
    job.result()
    return job.output_rows


# ---- 红利 / 客服对话（第二批，全部 append-only，绝不 TRUNCATE，不依赖 DML）----

def _parse_bonus_df(it, csv_path):
    """单档红利清洗（复用 import_bonus_records 逻辑）。返回 df（含 订单号/红利金额/_snapshot_month/活动名称）。"""
    import pandas as pd
    df = None
    for enc in ('utf-8', 'gb18030', 'gbk', 'big5'):
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            break
        except Exception:
            continue
    if df is None:
        raise ValueError('红利 CSV 编码识别失败')
    if '流水倍数(倍)' in df.columns:
        df = df.rename(columns={'流水倍数(倍)': '流水倍数'})
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = (df[col].astype(str)
                       .str.replace(r'^="(.*)"$', r'\1', regex=True).str.strip())
            df[col] = df[col].replace({'nan': None, 'None': None, '': None})
    if '红利标题' in df.columns:
        df['活动名称'] = df['红利标题'].fillna('').replace('', None)
        mask = df['活动名称'].isna()
        if '申请备注' in df.columns:
            df.loc[mask, '活动名称'] = df.loc[mask, '申请备注']
        df['活动名称'] = df['活动名称'].fillna('未知')
    if '申请时间' in df.columns:
        dt = pd.to_datetime(df['申请时间'], errors='coerce')
        df['_snapshot_month'] = dt.dt.strftime('%Y-%m')
    if '订单号' in df.columns:
        df = df.drop_duplicates(subset=['订单号'], keep='first')
    for col in ('红利金额', '流水倍数'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def _parse_cs_df(it, xlsx_path, source_file):
    """单档客服对话清洗（复用 import_cs_conversations 逻辑，多 sheet）。月份从 开始时间 推断。"""
    import re as _re
    import pandas as pd
    xl = pd.ExcelFile(xlsx_path)
    frames = []
    for sh in [s for s in xl.sheet_names if s != '总表']:
        try:
            d = pd.read_excel(xlsx_path, sheet_name=sh)
        except Exception:
            continue
        if d.empty:
            continue
        d['_sheet'] = sh
        frames.append(d)
    if not frames:
        raise ValueError('客服对话 xlsx 没读到数据 sheet')
    merged = pd.concat(frames, ignore_index=True)
    for c in ('开始时间', '结束时间'):
        if c in merged.columns:
            merged[c] = pd.to_datetime(merged[c], errors='coerce')
    month = ''
    if '开始时间' in merged.columns and merged['开始时间'].notna().any():
        mode = merged['开始时间'].dt.strftime('%Y%m').mode()
        if len(mode):
            month = mode.iloc[0]
    if not month:
        m = _re.search(r'(20\d{2})\D?([01]\d)', source_file)
        month = f'{m.group(1)}{m.group(2)}' if m else ''
    merged['_snapshot_month'] = month
    for c in ('首次响应', '平均响应', '总时长', '访客消息数', '客服消息数', '撤回消息数', '对话回合数'):
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors='coerce')
    str_cols = ['终端', '访客ID', '对话ID', '新对话ID', '会员账号', '地区', '接待客服',
                '访客IP', '网站名称', '是否邀请评价', '满意度评价', '评价内容',
                '服务主题', '备注', '机器人标识', '对话内容', '_sheet']
    for c in str_cols:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).replace({'nan': None, 'None': None, '': None})
    UNHAPPY = {'非常不满意', '不满意'}

    def _ext(row):
        if str(row.get('满意度评价') or '') not in UNHAPPY:
            return ''
        return it.extract_unhappy_reason(row)
    merged['_extracted_issue'] = [_ext(merged.iloc[i]) for i in range(len(merged))]
    return merged


def _existing_bonus_orders(client):
    try:
        rows = client.query(
            f"SELECT `订单号` FROM `{BQ_PREFIX}.raw_bonus_report`").result()
        return set(str(r['订单号']) for r in rows)
    except Exception:
        return set()


def _append_bonus(client, df, source_file, existing_orders=None):
    """只追加「新订单号」的红利行，既有订单一律不动。返回真正写入行数。"""
    import pandas as pd
    if existing_orders is None:
        existing_orders = _existing_bonus_orders(client)
    if '订单号' in df.columns and existing_orders:
        new = df[~df['订单号'].astype(str).isin(existing_orders)].copy()
    else:
        new = df.copy()
    if len(new) == 0:
        return 0
    new['_source_file'] = source_file
    new['_imported_at'] = pd.Timestamp.now()
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION])
    client.load_table_from_dataframe(
        new, f"{BQ_PREFIX}.raw_bonus_report", job_config=cfg).result()
    return len(new)


def _cs_basename_loaded(client, basename: str) -> int:
    sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.raw_cs_conversations` "
           f"WHERE _source_file = @b OR ENDS_WITH(_source_file, @sb)")
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter('b', 'STRING', basename),
        bigquery.ScalarQueryParameter('sb', 'STRING', '/' + basename)])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _append_cs(client, df, source_file):
    import pandas as pd
    payload = df.copy()
    payload['_source_file'] = source_file
    payload['_imported_at'] = pd.Timestamp.now()
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        autodetect=True,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION])
    client.load_table_from_dataframe(
        payload, f"{BQ_PREFIX}.raw_cs_conversations", job_config=cfg).result()
    return len(payload)


# ---- 代理佣金 单线/团队版（第三批，append-safe by 佣金月份；绝不依赖 DML）----
_COMM_NUMERIC = {'存款金额', '提款金额', '总输赢', '场馆费', '红利', '代理冲销', '返水', '账户调整',
    '存款手续费基数', '提款手续费基数', '存款手续费', '提款手续费', '补单输赢', '净输赢', '上月结余',
    '冲账调整', '冲正后净输赢', '佣金调整', '佣金', '已发放佣金', '剩余佣金', '申请发放佣金', 'VIP专享'}
_COMM_PERCENT = {'佣金比例', '二次佣金比例'}
_COMM_INT = {'团队人数', '下级人数', '注册人数', '首存人数', '活跃人数', '新增活跃人数', '有效活跃人数'}
_COMM_DATE = {'成为代理时间', '加入团队时间'}
_COMM_BOOL = {'是否在团队', '是否取消代理资格', '是否为主线'}


def _comm_num(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().replace(',', '')
    if s in ('', '-', 'None', 'nan'):
        return None
    try:
        return float(s)
    except Exception:
        return None


def _comm_pct(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().rstrip('%')
    if s in ('', '-', 'None', 'nan'):
        return None
    try:
        return float(s) / 100.0
    except Exception:
        return None


def _comm_int(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def _comm_bool(v):
    s = str(v).strip() if v is not None else ''
    return True if s == '是' else (False if s == '否' else None)


def _commission_table(cols):
    cset = set(str(c).strip() for c in cols)
    if ('线别' in cset or '主线/副线' in cset) and '团队名称' in cset:
        return 'raw_agent_commission_team'
    if '是否在团队' in cset:
        return 'raw_agent_commission_single'
    return None


def _parse_commission(it, path):
    import pandas as pd
    df = it.read_file(path)
    if '主线/副线' in df.columns:
        df = df.rename(columns={'主线/副线': '线别'})
    for c in df.columns:
        cs = str(c).strip()
        if cs in _COMM_NUMERIC:
            df[c] = df[c].map(_comm_num)
        elif cs in _COMM_PERCENT:
            df[c] = df[c].map(_comm_pct)
        elif cs in _COMM_INT:
            df[c] = df[c].map(_comm_int)
        elif cs in _COMM_DATE:
            df[c] = pd.to_datetime(df[c], errors='coerce')
        elif cs in _COMM_BOOL:
            df[c] = df[c].map(_comm_bool)
        else:
            df[c] = df[c].map(lambda v: None if v is None or str(v).strip() in ('', 'nan', 'None') else str(v).strip())
    keep = [c for c in df.columns if c not in ('_source_file', '_imported_at')]
    return df.dropna(subset=keep, how='all')


def _commission_months(df):
    if '佣金月份' in df.columns:
        return sorted(set(df['佣金月份'].dropna().astype(str)))
    return []


def _commission_month_exists(client, table, months):
    if not months:
        return 0
    sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
           f"WHERE CAST(`佣金月份` AS STRING) IN UNNEST(@ms)")
    cfg = bigquery.QueryJobConfig(query_parameters=[
        bigquery.ArrayQueryParameter('ms', 'STRING', months)])
    try:
        return list(client.query(sql, job_config=cfg).result())[0].n
    except Exception:
        return 0


def _write_commission_safe(client, new_df, table, source_files):
    """读现有→去掉 new 的月份→concat→统一类型→TRUNCATE 写回。
    保留其他月份不动、刷新上传的月份；写回前校验「其他月份行数不变」防掉数据。返回写入总行数。"""
    import pandas as pd
    new_months = set(new_df['佣金月份'].dropna().astype(str)) if '佣金月份' in new_df.columns else set()
    existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    if '佣金月份' in existing.columns and new_months:
        keep = existing[~existing['佣金月份'].astype(str).isin(new_months)]
    else:
        keep = existing
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_files
    combined = pd.concat([keep, nd], ignore_index=True)
    for c in list(combined.columns):
        cs = str(c).strip()
        if cs in (_COMM_NUMERIC | _COMM_PERCENT | _COMM_INT):
            combined[c] = pd.to_numeric(combined[c], errors='coerce')
        elif cs in _COMM_DATE:
            combined[c] = pd.to_datetime(combined[c].astype(str), errors='coerce')
        elif cs in _COMM_BOOL:
            combined[c] = combined[c].map(
                lambda v: True if v in (True, 'True', '是') else (False if v in (False, 'False', '否') else None)
            ).astype('boolean')
    # 防掉数据：每个「没在上传的旧月份」行数必须原样保留
    if '佣金月份' in existing.columns:
        ex_counts = existing['佣金月份'].astype(str).value_counts().to_dict()
        cb_counts = combined['佣金月份'].astype(str).value_counts().to_dict()
        for m, cnt in ex_counts.items():
            if m not in new_months and cb_counts.get(m, 0) != cnt:
                raise RuntimeError(f'安全中止：旧月份 {m} 行数会变（{cnt}→{cb_counts.get(m, 0)}），拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return len(combined)


# ── 按月快照报表（会员 / TOP）同月替换 + 删除月份（read-modify-truncate，免费版真删）──

SNAPSHOT_TABLES = {'raw_member_report', 'raw_top_report'}  # 有 _snapshot_month、同月应替换


def _replace_by_snapshot_month(client, new_df, table, source_file):
    """会员/TOP 等按月快照表：同月「按会员账号+代理合并」（不是整月覆盖）。
    同一个月里：上传档有的(账号+代理)→用新的更新；只在旧档有的→保留；新的→加进来；其他月份完全不动。
    去重必须含代理：同名挂不同代理=不同人，只按账号会把另一个代理下的同名会员误删（跟 member_count 一致）。
    这样分批传同一个月（如 6/1-12 再传 6/13-30）会累加合并、不丢数据。
    返回 (months_str, updated, added, total)。"""
    months = (set(new_df['_snapshot_month'].dropna().astype(str))
              if '_snapshot_month' in new_df.columns else set())
    existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    has_acct = '会员账号' in existing.columns and '会员账号' in new_df.columns
    has_agent = has_acct and '代理' in existing.columns and '代理' in new_df.columns

    def _merge_key(d):  # 账号+代理；没代理列才退回纯账号
        if has_agent:
            return d['会员账号'].astype(str) + '\x01' + d['代理'].astype(str)
        return d['会员账号'].astype(str)

    new_keys = set(_merge_key(new_df)) if has_acct else set()
    if '_snapshot_month' in existing.columns and months and has_acct:
        # 去掉「同月 且 (账号+代理)在新档里」的旧行（被新行取代）；同月没在新档的 + 其他月，全保留
        same_month = existing['_snapshot_month'].astype(str).isin(months)
        old_mask = same_month & _merge_key(existing).isin(new_keys)
        updated = int(old_mask.sum())
        kept_same_month = int((same_month & ~old_mask).sum())  # 同月、新档没有 → 保留的旧会员
        added = len(new_df) - updated
    elif '_snapshot_month' in existing.columns and months:
        # 没有会员账号栏（如某些 TOP）→ 退回整月替换
        old_mask = existing['_snapshot_month'].astype(str).isin(months)
        updated = int(old_mask.sum())
        added = len(new_df)
    else:
        old_mask = pd.Series(False, index=existing.index)
        updated = 0
        added = len(new_df)
    keep = existing[~old_mask].copy()
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    combined = pd.concat([keep, nd], ignore_index=True)
    # 防掉数据：没在上传的「其他月份」行数必须原样保留
    if '_snapshot_month' in existing.columns:
        ex_counts = existing['_snapshot_month'].astype(str).value_counts().to_dict()
        cb_counts = combined['_snapshot_month'].astype(str).value_counts().to_dict()
        for m, cnt in ex_counts.items():
            if m not in months and cb_counts.get(m, 0) != cnt:
                raise RuntimeError(f'安全中止：其他月份 {m} 行数会变（{cnt}→{cb_counts.get(m, 0)}），拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return ','.join(sorted(months)) if months else '?', updated, added, len(combined)


def _date_col_of(df) -> Optional[str]:
    """日报类报表的日期栏：日期 或 时间（取前 10 字 = YYYY-MM-DD，realtime「..23~24」也 OK）。"""
    if '日期' in df.columns:
        return '日期'
    if '时间' in df.columns:
        return '时间'
    if '订单时间' in df.columns:   # 存款订单
        return '订单时间'
    if '申请时间' in df.columns:   # 提款订单
        return '申请时间'
    return None


def _date_keys(series) -> set:
    ks = set(series.astype(str).str[:10])
    ks.discard('')
    ks.discard('nan')
    ks.discard('None')
    return ks


def _looks_month_aggregated(df) -> bool:
    """日报类报表若日期栏大多只有 YYYY-MM（没有「日」），多半是后台误用「按月」颗粒度导出的
    「整月汇总」而非「日明细」。导进去后按日期范围筛会整月对不上（跟其他月日明细不一致）。"""
    dc = _date_col_of(df)
    if dc is None:
        return False
    s = df[dc].dropna().astype(str).str.strip()
    s = s[(s != '') & (s.str.lower() != 'nan') & (s != 'none')]
    if s.empty:
        return False
    month_only = s.str.match(r'^\d{4}-\d{2}$')
    return float(month_only.mean()) > 0.8


def _replace_by_date_range(client, new_df, table, source_file):
    """日报类（平台/财务/游戏/场馆/分析/推广/代理/实时注单）：按日期替换。
    去掉「现有表里日期在新档日期集合内」的旧行→写新行，其它日期一行不动。
    重传同一天/同月 = 覆盖那些日期，永不重复。返回 (date_range_str, removed, written, total)。"""
    dc = _date_col_of(new_df)
    if dc is None:
        # 没日期栏 → 退回按档名追加（理论上日报类都有日期，不该走到这）
        n = _append_standard_report(client, new_df, table, source_file)
        return '(无日期栏)', 0, n, None
    new_dates = _date_keys(new_df[dc])
    new_months = {d[:7] for d in new_dates}
    existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    if dc in existing.columns and new_dates:
        ex_key = existing[dc].astype(str).str[:10]
        # 覆盖同一天的旧行；并顺手清掉「同月的整月汇总行」(如旧的 2026-05)，
        # 否则它的 key 不等于任何 2026-05-DD、会跟新的日明细并存造成重复/口径混乱
        stale_month_agg = ex_key.str.match(r'^\d{4}-\d{2}$') & ex_key.str[:7].isin(new_months)
        old_mask = ex_key.isin(new_dates) | stale_month_agg
    else:
        old_mask = pd.Series(False, index=existing.index)
    removed = int(old_mask.sum())
    keep = existing[~old_mask].copy()
    # 防掉数据：保留的行数必须 = 现有 - 删除
    if len(keep) != len(existing) - removed:
        raise RuntimeError('安全中止：行数对不上，拒绝写入')
    nd = new_df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    combined = pd.concat([keep, nd], ignore_index=True)
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    rng = (f"{min(new_dates)} ~ {max(new_dates)}" if len(new_dates) > 1
           else (next(iter(new_dates)) if new_dates else '?'))
    return rng, removed, len(nd), len(combined)


def _table_period_col(table: str):
    """这张表用哪个栏当「月份/期间」键。没有→None（只能按上传档名删）。"""
    if table in SNAPSHOT_TABLES:
        return '_snapshot_month'
    if table in ('raw_agent_commission_single', 'raw_agent_commission_team'):
        return '佣金月份'
    if table in ('raw_winback', 'raw_agent_settlement_monthly'):
        return '月份'
    return None


def _bq_periods(client, table: str):
    """回传 (period_col, {期间: 行数})。有月份栏就按月份；否则按 _source_file（上传批次）。"""
    col = _table_period_col(table)
    key = col if col else '_source_file'
    try:
        sql = (f"SELECT CAST(`{key}` AS STRING) AS p, COUNT(*) AS n "
               f"FROM `{BQ_PREFIX}.{table}` GROUP BY p ORDER BY p")
        rows = list(client.query(sql).result())
        return key, {(r.p if r.p is not None else '(空)'): r.n for r in rows}
    except Exception:
        return key, {}


def _bq_delete_periods(client, table: str, key: str, periods):
    """真删：读现有→去掉指定期间→TRUNCATE 写回。返回 (removed, remaining)。"""
    periods = set(str(p) for p in periods)
    existing = client.query(f"SELECT * FROM `{BQ_PREFIX}.{table}`").result().to_dataframe()
    if key not in existing.columns:
        return 0, len(existing)
    mask = existing[key].astype(str).isin(periods)
    removed = int(mask.sum())
    keep = existing[~mask].copy()
    if len(keep) != len(existing) - removed:
        raise RuntimeError('安全中止：行数对不上，拒绝写入')
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(keep, f"{BQ_PREFIX}.{table}", job_config=cfg).result()
    return removed, len(keep)


def _try_parse_agent_monthly(tmp_path):
    """市代「整理资料」分页：每代理每月 充值/投注/输赢/挂账。返回 (df, months) 或 None。
    认法：xlsx 里有一张分页含「代理帐号 + 月份 + (累计挂账额度 或 发放佣金)」栏。"""
    import os
    if os.path.splitext(tmp_path)[1].lower() not in ('.xlsx', '.xls'):
        return None
    try:
        xl = pd.ExcelFile(tmp_path)
    except Exception:
        return None
    target = None
    for sh in xl.sheet_names:
        try:
            cols = set(str(c).strip() for c in pd.read_excel(tmp_path, sheet_name=sh, nrows=0).columns)
        except Exception:
            continue
        if {'代理帐号', '月份'}.issubset(cols) and ('累计挂账额度' in cols or '发放佣金' in cols):
            target = sh
            break
    if target is None:
        return None
    df = pd.read_excel(tmp_path, sheet_name=target, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['代理帐号'].notna() & (df['代理帐号'].astype(str).str.strip() != '')].copy()
    mm = pd.to_datetime(df['月份'], errors='coerce')
    df = df[mm.notna()].copy()
    if df.empty:
        return None
    df['月份'] = pd.to_datetime(df['月份'], errors='coerce').dt.strftime('%Y-%m')
    text_cols = {'代理名称', '代理帐号', '发展情况', '开户日期', '月份', '来源栏位'}
    for c in df.columns:
        if c in text_cols:
            df[c] = df[c].astype(str).str.strip()
        else:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    months = sorted(df['月份'].dropna().unique().tolist())
    return df, months


def _write_agent_monthly(client, df, months, source_file):
    """写 raw_agent_settlement_monthly：按月份替换（读改 WRITE_TRUNCATE，沙盒禁 DELETE）。返回写入行数。"""
    table = 'raw_agent_settlement_monthly'
    nd = df.copy()
    nd['_imported_at'] = pd.Timestamp.now()
    nd['_source_file'] = source_file
    full = f'{BQ_PREFIX}.{table}'
    try:
        existing = client.query(f'SELECT * FROM `{full}`').result().to_dataframe()
    except Exception:
        existing = pd.DataFrame()
    if not existing.empty and '月份' in existing.columns:
        keep = existing[~existing['月份'].astype(str).isin([str(m) for m in months])].copy()
        combined = pd.concat([keep, nd], ignore_index=True)
    else:
        combined = nd
    cfg = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
    client.load_table_from_dataframe(combined, full, job_config=cfg).result()
    return len(nd)


def _try_parse_agent_settlement(tmp_path):
    """平哥「X月代理帐.xlsx」（位置式：R1/R2 累计挂帐摘要 + R4/R5 发放摘要 + R9+ 明细 9 栏）。
    返回 (summary_records, detail_records) 或 None。月份在档外，写入时再补。"""
    import os
    if os.path.splitext(tmp_path)[1].lower() not in ('.xlsx', '.xls'):
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, data_only=True)
    except Exception:
        return None
    sheet = next((s for s in wb.sheetnames if ('代理帐' in s or '代理账' in s)), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    if len(rows) < 10:
        return None
    r7 = [str(c).strip() if c is not None else '' for c in (rows[7] or [])]
    if '名称' not in r7 or '总代' not in r7:   # 不是这种位置式格式
        return None

    def _num(v):
        return float(v) if isinstance(v, (int, float)) else None
    summary = []
    if len(rows) > 2:
        l1, v1 = rows[1] or [], rows[2] or []
        for i in range(1, 7):
            if i < len(l1) and l1[i] and i < len(v1) and isinstance(v1[i], (int, float)):
                summary.append({'项目': str(l1[i]).strip(), '金额': float(v1[i])})
    if len(rows) > 5:
        l2, v2 = rows[4] or [], rows[5] or []
        for i in range(3, 7):
            if i < len(l2) and l2[i] and i < len(v2) and isinstance(v2[i], (int, float)):
                summary.append({'项目': str(l2[i]).strip(), '金额': float(v2[i])})
    detail = []
    for r in rows[8:]:
        if not any(v not in (None, '') for v in r):
            continue
        treatment = r[6] if len(r) > 6 else None
        name = r[1] if len(r) > 1 else None
        agent = r[2] if len(r) > 2 else None
        if not treatment or not name or not agent:
            continue
        detail.append({
            '回款状态': str(r[0]).strip() if r[0] else None,
            '名称': str(name).strip(), '总代账号': str(agent).strip(),
            '先前挂账业绩': _num(r[3]) if len(r) > 3 else None,
            '本月业绩': _num(r[4]) if len(r) > 4 else None,
            '业绩总计': _num(r[5]) if len(r) > 5 else None,
            '适用待遇': str(treatment).strip(),
            '比例': _num(r[7]) if len(r) > 7 else None,
            '实际佣金': _num(r[8]) if len(r) > 8 else None,
        })
    if not summary and not detail:
        return None
    return summary, detail


def _write_agent_settlement(client, summary_records, detail_records, month, source_file=''):
    """写 raw_agent_settlement_summary + _detail：按月份替换（读改 WRITE_TRUNCATE）。返回 (摘要行, 明细行)。"""
    def _write_one(table, records):
        if not records:
            return 0
        df = pd.DataFrame(records)
        df['月份'] = month
        df['_imported_at'] = pd.Timestamp.now().isoformat()   # 跟原表 _imported_at 同为字符串，避免 Timestamp 类型冲突
        df['_source_file'] = source_file or f'代理帐_{month}.xlsx'
        full = f'{BQ_PREFIX}.{table}'
        try:
            existing = client.query(f'SELECT * FROM `{full}`').result().to_dataframe()
        except Exception:
            existing = pd.DataFrame()
        if not existing.empty and '月份' in existing.columns:
            keep = existing[existing['月份'].astype(str) != str(month)].copy()
            combined = pd.concat([keep, df], ignore_index=True)
        else:
            combined = df
        cfg = bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
        client.load_table_from_dataframe(combined, full, job_config=cfg).result()
        return len(df)
    ns = _write_one('raw_agent_settlement_summary', summary_records)
    nd = _write_one('raw_agent_settlement_detail', detail_records)
    return ns, nd


def _classify_and_parse(it, client, src, tmp_path):
    """识别+解析一个上传单元。返回 entry dict（kind/display/df/rows/is_new/status）。"""
    import os
    # 1) 标准月报（先文件名匹配 FILE_MAP；认不出再靠栏位内容认）
    table, display = it.identify_report_type(src)
    content_override = None
    if table is None:
        content_override = _identify_by_content(it, tmp_path)
        if content_override:
            table, display = content_override
    if table is not None:
        _, disp, df = _parse_standard_report(it, tmp_path, src, table_override=content_override)
        # 会员 / TOP 是按月快照：同月「按会员账号合并」（分批传同月会累加、不丢；换档名也不会变两份）
        if table in SNAPSHOT_TABLES:
            mth = (str(df['_snapshot_month'].iloc[0])
                   if '_snapshot_month' in df.columns and len(df) else '')
            if mth in ('', 'None', 'nan', '?'):
                # 抓不到月份（如 TOP 报表 top.xlsx 裸档名、内容无日期栏）→ 让上传页选月份再写
                return {'src': src, 'kind': 'snapshot', 'table': table, 'display': disp,
                        'df': df, 'rows': len(df), 'is_new': len(df) > 0,
                        'snapshot_month': None, 'need_month': True, 'exists': 0,
                        'status': '🗓 需选月份（档名/内容判不出是哪个月，请在下方选）'}
            try:
                sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                       f"WHERE CAST(`_snapshot_month` AS STRING)=@m")
                cfg = bigquery.QueryJobConfig(query_parameters=[
                    bigquery.ScalarQueryParameter('m', 'STRING', mth)])
                exists = list(client.query(sql, job_config=cfg).result())[0].n
            except Exception:
                exists = 0
            return {'src': src, 'kind': 'snapshot', 'table': table, 'display': disp,
                    'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'snapshot_month': mth,
                    'exists': exists,
                    'status': (f'🔄 合并到 {mth}（同月已有 {exists} 行，按会员账号合并：重复的更新、新会员追加、其他月不动）'
                               if exists else f'🆕 新增 {mth}')}
        # 日报类（有日期/时间栏）：按日期替换——重传同一天/同月就覆盖那些日期，永不重复
        dc = _date_col_of(df)
        if dc:
            nd_dates = _date_keys(df[dc])
            rng = (f"{min(nd_dates)} ~ {max(nd_dates)}" if len(nd_dates) > 1
                   else (next(iter(nd_dates)) if nd_dates else '?'))
            try:
                sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                       f"WHERE SUBSTR(CAST(`{dc}` AS STRING),1,10) IN UNNEST(@d)")
                cfg = bigquery.QueryJobConfig(query_parameters=[
                    bigquery.ArrayQueryParameter('d', 'STRING', list(nd_dates))])
                exists = list(client.query(sql, job_config=cfg).result())[0].n
            except Exception:
                exists = 0
            month_agg = _looks_month_aggregated(df)
            base_status = (f'🔄 替换 {rng}（覆盖这段日期已有的 {exists} 行，其他日期不动）'
                           if exists else f'🆕 新增 {rng}')
            if month_agg:
                base_status = '⚠️ 疑似「整月汇总」非日明细（日期只有年月、没到日）— ' + base_status
            return {'src': src, 'kind': 'standard', 'table': table, 'display': disp,
                    'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'date_range': rng,
                    'date_col': dc, 'date_keys': sorted(nd_dates), 'exists': exists,
                    'warn_month_agg': month_agg, 'status': base_status}
        # 没日期/时间栏（理论上日报类都有）→ 退回按档名去重
        dup = _bq_count_source_file(client, table, src)
        return {'src': src, 'kind': 'standard', 'table': table, 'display': disp,
                'df': df, 'rows': len(df), 'is_new': dup == 0,
                'status': (f'⏭ 已上传过（{dup} 行），会跳过' if dup else '🆕 待写入')}
    # 1.34) 代理结算月报（平哥「X月代理帐.xlsx」位置式：累计挂帐摘要 + 适用待遇/退成明细）→ 需选月份、按月替换
    sett = _try_parse_agent_settlement(tmp_path)
    if sett is not None:
        summ, det = sett
        return {'src': src, 'kind': 'settlement', 'display': '代理结算月报(平哥)',
                'summary_records': summ, 'detail_records': det, 'df': None,
                'rows': len(det), 'is_new': (len(summ) + len(det)) > 0,
                'need_month': True, 'snapshot_month': None,
                'status': f'🗓 需选月份（代理结算月报：{len(summ)} 摘要 + {len(det)} 明细）'}
    # 1.35) 代理结算月度（市代「整理资料」分页：每代理每月 充值/投注/输赢/挂账）→ 按月份替换
    am = _try_parse_agent_monthly(tmp_path)
    if am is not None:
        am_df, am_months = am
        try:
            sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.raw_agent_settlement_monthly` "
                   f"WHERE CAST(`月份` AS STRING) IN UNNEST(@m)")
            cfg = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ArrayQueryParameter('m', 'STRING', am_months)])
            exists = list(client.query(sql, job_config=cfg).result())[0].n
        except Exception:
            exists = 0
        rng = f'{am_months[0]} ~ {am_months[-1]}' if len(am_months) > 1 else (am_months[0] if am_months else '?')
        return {'src': src, 'kind': 'agent_monthly', 'table': 'raw_agent_settlement_monthly',
                'display': '代理结算月度(市代)', 'df': am_df, 'rows': len(am_df),
                'is_new': len(am_df) > 0, 'months': am_months,
                'status': (f'🔄 按月替换 {rng}（{len(am_months)} 个月，覆盖同月已有 {exists} 行，其他月不动）'
                           if exists else f'🆕 新增 {rng}（{len(am_months)} 个月）')}
    # 1.4) 存取款订单（存款管理/提款管理 历史记录 Csv）→ 按日期替换
    if ('存款管理' in src) or ('提款管理' in src):
        is_dep = '存款管理' in src
        table = 'raw_finance_deposit' if is_dep else 'raw_finance_withdraw'
        disp = '存款订单' if is_dep else '提款订单'
        fdf = None
        for enc in ('utf-8-sig', 'gbk', 'utf-8', 'big5'):
            try:
                fdf = pd.read_csv(tmp_path, dtype=str, encoding=enc)
                break
            except Exception:
                continue
        if fdf is None or fdf.empty:
            return {'src': src, 'kind': 'none', 'display': disp, 'df': None, 'rows': 0,
                    'is_new': False, 'status': '⚠️ 读不出该 Csv'}
        for c in fdf.columns:
            fdf[c] = fdf[c].astype(str).str.replace('\t', '', regex=False).str.strip()
        # 去掉栏名里的括号单位（如 订单金额(元)→订单金额）；BigQuery 栏名不允许括号
        fdf.columns = [re.sub(r'[（(].*?[）)]', '', str(c)).strip() for c in fdf.columns]
        dc = '订单时间' if is_dep else '申请时间'
        if dc not in fdf.columns:
            return {'src': src, 'kind': 'none', 'display': disp, 'df': None, 'rows': 0,
                    'is_new': False, 'status': f'⚠️ 缺日期栏「{dc}」，认不出'}
        nd_dates = _date_keys(fdf[dc])
        rng = (f"{min(nd_dates)} ~ {max(nd_dates)}" if len(nd_dates) > 1
               else (next(iter(nd_dates)) if nd_dates else '?'))
        try:
            sql = (f"SELECT COUNT(*) AS n FROM `{BQ_PREFIX}.{table}` "
                   f"WHERE SUBSTR(CAST(`{dc}` AS STRING),1,10) IN UNNEST(@d)")
            cfg = bigquery.QueryJobConfig(query_parameters=[
                bigquery.ArrayQueryParameter('d', 'STRING', list(nd_dates))])
            exists = list(client.query(sql, job_config=cfg).result())[0].n
        except Exception:
            exists = 0
        return {'src': src, 'kind': 'standard', 'table': table, 'display': disp, 'df': fdf,
                'rows': len(fdf), 'is_new': len(fdf) > 0, 'date_range': rng,
                'date_col': dc, 'date_keys': sorted(nd_dates), 'exists': exists,
                'status': (f'🔄 替换 {rng}（覆盖这段日期已有的 {exists} 行）' if exists else f'🆕 新增 {rng}')}

    # 1.5) 电访召回「撥打紀錄總表」→ 解析成各专员月度汇总，存 raw_winback（按月份刷新）
    if ('撥打' in src) or ('拨打' in src) or ('電訪' in src) or ('电访' in src):
        try:
            wdf, wmeta = parse_winback_file(tmp_path)
        except Exception as e:
            return {'src': src, 'kind': 'none', 'display': '电访召回(解析失败)', 'df': None,
                    'rows': 0, 'is_new': False, 'status': f'⚠️ 撥打紀錄總表解析失败：{str(e)[:50]}'}
        ym = wmeta.get('month') or _winback_ym_from_name(src)
        if wdf is None or wdf.empty or not ym:
            return {'src': src, 'kind': 'none', 'display': '电访召回', 'df': None, 'rows': 0,
                    'is_new': False, 'status': '⚠️ 没读到专员数据或判不出月份，跳过'}
        wdf = wdf.copy()
        wdf['月份'] = ym
        exists = _bq_winback_month_exists(client, ym)
        return {'src': src, 'kind': 'winback', 'table': 'raw_winback', 'display': '电访召回',
                'df': wdf, 'rows': len(wdf), 'is_new': True, 'months': [ym],
                'status': (f'🔄 刷新 {ym}（覆盖该月旧的电访统计）' if exists else f'🆕 新增 {ym}')}

    # 2) 客服对话（文件名带「客服对话」）
    if '客服对话' in src:
        df = _parse_cs_df(it, tmp_path, src)
        dup = _cs_basename_loaded(client, os.path.basename(src))
        return {'src': src, 'kind': 'cs', 'display': '客服对话', 'df': df,
                'rows': len(df), 'is_new': dup == 0,
                'status': ('⏭ 已上传过，会跳过' if dup else '🆕 待写入')}
    # 3) 读列判断 红利 / 客服对话（红利 zip 内层档名是纯数字，认不出，只能看列）
    ext = os.path.splitext(src)[1].lower()
    try:
        if ext in ('.xlsx', '.xls'):
            import pandas as pd
            cols = set(str(c).strip() for c in pd.read_excel(tmp_path, nrows=5).columns)
        else:
            cols = set(str(c).strip() for c in it.read_file(tmp_path).columns)
    except Exception:
        cols = set()
    ct = _commission_table(cols)
    if ct:
        df = _parse_commission(it, tmp_path)
        months = _commission_months(df)
        exists = _commission_month_exists(client, ct, months)
        disp = '代理佣金(团队版)' if ct.endswith('team') else '代理佣金(单线版)'
        mlbl = ','.join(months) if months else '?'
        return {'src': src, 'kind': 'commission', 'table': ct, 'display': disp,
                'df': df, 'rows': len(df), 'is_new': len(df) > 0, 'months': months,
                'status': (f'🔄 刷新 {mlbl}（其他月不动）' if exists > 0 else f'🆕 新增 {mlbl}')}
    if {'满意度评价', '对话内容'} & cols:
        df = _parse_cs_df(it, tmp_path, src)
        dup = _cs_basename_loaded(client, os.path.basename(src))
        return {'src': src, 'kind': 'cs', 'display': '客服对话', 'df': df,
                'rows': len(df), 'is_new': dup == 0,
                'status': ('⏭ 已上传过，会跳过' if dup else '🆕 待写入')}
    if {'订单号', '红利金额'} <= cols or {'红利标题', '申请时间'} <= cols:
        df = _parse_bonus_df(it, tmp_path)
        existing = _existing_bonus_orders(client)
        new_n = (int((~df['订单号'].astype(str).isin(existing)).sum())
                 if '订单号' in df.columns else len(df))
        return {'src': src, 'kind': 'bonus', 'display': '红利记录', 'df': df,
                'rows': len(df), 'is_new': new_n > 0, '_existing_orders': existing,
                'status': (f'🆕 待写入（新订单 {new_n} 笔）' if new_n > 0
                           else '⏭ 订单全部已存在，会跳过')}
    return {'src': src, 'kind': 'none', 'display': '⚠️ 未识别', 'df': None,
            'rows': 0, 'is_new': False, 'status': '跳过（认不出类型）'}


def _recover_zip_name(name: str, is_utf8: bool) -> str:
    """zipfile 对非 UTF-8 条目名用 cp437 解码→中文变乱码；还原回 GBK/Big5。"""
    if is_utf8:
        return name
    for enc in ('gbk', 'big5'):
        try:
            return name.encode('cp437').decode(enc)
        except Exception:
            continue
    return name


def _expand_upload_units(f, zip_password: str):
    """把一个上传档展开成若干 (source_name, tmp_path, err)。
    普通档→1 个；zip→解压里面每个 csv/xlsx（用密码 + 还原中文名）。调用方负责删 tmp_path。"""
    import os
    import tempfile
    import zipfile

    name = f.name
    ext = os.path.splitext(name)[1].lower()
    units = []
    if ext == '.zip':
        zpath = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as zt:
                zt.write(f.getbuffer())
                zpath = zt.name
            zf = zipfile.ZipFile(zpath)
            pwd = zip_password.encode() if zip_password else None
            data_members = [
                m for m in zf.infolist()
                if not m.is_dir()
                and os.path.splitext(m.filename)[1].lower() in ('.csv', '.xlsx', '.xls')
            ]
            if not data_members:
                units.append((name, None, 'zip 里没有 csv/xlsx'))
            for m in data_members:
                inner_ext = os.path.splitext(m.filename)[1].lower()
                src = os.path.basename(
                    _recover_zip_name(m.filename, bool(m.flag_bits & 0x800)))
                try:
                    data = zf.read(m, pwd=pwd)
                except RuntimeError as e:
                    units.append((src, None, f'解压失败（密码错？）：{str(e)[:40]}'))
                    continue
                with tempfile.NamedTemporaryFile(delete=False, suffix=inner_ext) as itmp:
                    itmp.write(data)
                    units.append((src, itmp.name, None))
        except Exception as e:
            units.append((name, None, f'zip 错误：{str(e)[:50]}'))
        finally:
            if zpath:
                try:
                    os.unlink(zpath)
                except Exception:
                    pass
    else:
        suffix = ext or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(f.getbuffer())
            units.append((name, tmp.name, None))
    return units


def render_agent_market_monthly():
    hero('市代月度结算',
         '市场代理每月的 产值 / 输赢 / 红利 / 返水 / 集团分成 / 发放佣金 / 平台净营利 / 累计挂账，可看走势、两月对比。',
         source_badge='数据源：市代「整理资料」(自助上传)',
         basis='市代「整理资料」表（数据上传页拖 xlsx 即入库；月份读表内「月份」栏）',
         detail=(
             '**分析范围**：市场代理每月的产值、输赢、红利/返水、集团分成、发放佣金、平台净营利、累计挂账，'
             '含月度走势与两月对比。\n\n'
             '**数据来源**：市代「整理资料」表（每代理每月一行），自助上传至「数据上传」页 → `raw_agent_settlement_monthly`，按月份去重。\n\n'
             '**跟「代理佣金」页的区别**：那页是「代理佣金 单线/团队」+ 平哥结算月报（不同数据源）；这页只用你上传的市代月度表。'))
    try:
        am = load_table('raw_agent_settlement_monthly')
    except Exception:
        am = pd.DataFrame()
    if am is None or am.empty or '月份' not in am.columns:
        st.info('📭 尚无市代月度数据。请到「🗂 数据上传」页把市代「整理资料」xlsx 拖进去，这页就读得到。')
        return
    am_months = sorted(am['月份'].astype(str).unique().tolist(), reverse=True)

    section_header('单月总览', '选月份看当月各代理合计')
    sel_am = st.selectbox('📅 月份', am_months, key='amp_month')
    cur = am[am['月份'].astype(str) == sel_am].copy()
    for c in ('注册人数', '首存人数', '活跃人数', '总充', '总提', '有效投注额', '总输赢',
              '红利', '返水', '集团分成', '发放佣金', '平台净营利', '累计挂账额度'):
        if c in cur.columns:
            cur[c] = pd.to_numeric(cur[c], errors='coerce')

    def _s(c):
        return float(cur[c].sum()) if c in cur.columns else 0.0
    k1 = st.columns(4)
    show_metric(k1[0], '市场代理数', fmt_num(cur['代理帐号'].nunique()))
    show_metric(k1[1], '有效投注额', fmt_num(round(_s('有效投注额'))))
    show_metric(k1[2], '发放佣金', fmt_num(round(_s('发放佣金'))), tone='warn')
    show_metric(k1[3], '平台净营利', fmt_num(round(_s('平台净营利'))), tone=tone_by_sign(_s('平台净营利')))
    k2 = st.columns(4)
    show_metric(k2[0], '累计挂账（合计）', fmt_num(round(_s('累计挂账额度'))), tone='warn',
                help_text='截至该月各市场代理的累计挂账余额合计（负值＝代理端尚有赤字）')
    show_metric(k2[1], '集团分成', fmt_num(round(_s('集团分成'))))
    show_metric(k2[2], '红利', fmt_num(round(_s('红利'))), tone='warn')
    show_metric(k2[3], '返水', fmt_num(round(_s('返水'))), tone='warn')

    section_header('各市场代理明细', f'{sel_am} · 按平台净营利排序')
    disp_cols = [c for c in ['代理名称', '代理帐号', '活跃人数', '总充', '有效投注额', '总输赢',
                             '红利', '返水', '集团分成', '发放佣金', '平台净营利', '累计挂账额度', '发展情况']
                 if c in cur.columns]
    sort_col = '平台净营利' if '平台净营利' in cur.columns else disp_cols[0]
    st.dataframe(cur[disp_cols].sort_values(sort_col, ascending=False), width='stretch', hide_index=True)

    # 月度走势
    am2 = am.copy()
    for c in ('有效投注额', '总输赢', '红利', '返水', '集团分成', '发放佣金', '平台净营利', '累计挂账额度'):
        if c in am2.columns:
            am2[c] = pd.to_numeric(am2[c], errors='coerce')
    trend = (am2.groupby('月份')
             .agg(发放佣金=('发放佣金', 'sum'), 平台净营利=('平台净营利', 'sum'),
                  累计挂账=('累计挂账额度', 'sum'), 有效投注额=('有效投注额', 'sum'))
             .reset_index().sort_values('月份'))
    section_header('月度走势', '各月合计（你上传的全部月份）')
    c_a, c_b = st.columns(2)
    with c_a:
        with st.container(border=True):
            fig = px.bar(trend, x='月份', y=['发放佣金', '平台净营利'], barmode='group',
                         template=TEMPLATE, title='发放佣金 vs 平台净营利')
            fig.update_layout(height=320, legend_title_text='', xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig, width='stretch')
    with c_b:
        with st.container(border=True):
            fig2 = px.line(trend, x='月份', y='累计挂账', markers=True, template=TEMPLATE, title='累计挂账走势')
            fig2.update_layout(height=320, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig2, width='stretch')

    # 两月对比
    section_header('两月对比', '选两个月，看整体与各代理的变化（变化 = B − A）')
    cc = st.columns(2)
    ma = cc[0].selectbox('A 月（基准）', am_months, index=min(1, len(am_months) - 1), key='amp_cmpA')
    mb = cc[1].selectbox('B 月（对比）', am_months, index=0, key='amp_cmpB')

    def _msum(m, c):
        d = am2[am2['月份'].astype(str) == m]
        return float(pd.to_numeric(d[c], errors='coerce').sum()) if c in d.columns and len(d) else 0.0
    cmp_rows = []
    for c in ['有效投注额', '总输赢', '红利', '返水', '集团分成', '发放佣金', '平台净营利', '累计挂账额度']:
        if c in am2.columns:
            a, b = _msum(ma, c), _msum(mb, c)
            cmp_rows.append({'指标': c, ma: round(a), mb: round(b), '变化(B−A)': round(b - a)})
    st.dataframe(pd.DataFrame(cmp_rows), width='stretch', hide_index=True)

    keyc = '平台净营利' if '平台净营利' in am2.columns else '发放佣金'
    da = am2[am2['月份'].astype(str) == ma][['代理帐号', '代理名称', keyc]].rename(columns={keyc: f'A({ma})'})
    db = am2[am2['月份'].astype(str) == mb][['代理帐号', keyc]].rename(columns={keyc: f'B({mb})'})
    mg = pd.merge(da, db, on='代理帐号', how='outer')
    mg[f'A({ma})'] = pd.to_numeric(mg[f'A({ma})'], errors='coerce').fillna(0)
    mg[f'B({mb})'] = pd.to_numeric(mg[f'B({mb})'], errors='coerce').fillna(0)
    mg['变化'] = mg[f'B({mb})'] - mg[f'A({ma})']
    mg = mg.sort_values('变化', ascending=False)
    section_header('各代理变化', f'{keyc}：{ma} → {mb}（涨幅排前、跌幅排后）')
    st.dataframe(mg, width='stretch', hide_index=True,
                 column_config={f'A({ma})': st.column_config.NumberColumn(format='%.0f'),
                                f'B({mb})': st.column_config.NumberColumn(format='%.0f'),
                                '变化': st.column_config.NumberColumn(format='%.0f')})


@st.cache_data(ttl=600)
def _data_health_rows():
    """逐表查数据最新日期 + 行数（一条 UNION 查询，省调用）。kind: auto=每日自动 / manual=手动上传。"""
    CFG = [
        ('存款订单 · 存取款分析', 'raw_finance_deposit', '完成时间', 'auto'),
        ('提款订单 · 存取款分析', 'raw_finance_withdraw', '完成时间', 'auto'),
        ('平台报表 · 经营总览', 'raw_platform_report', '日期', 'manual'),
        ('财务报表 · 经营总览', 'raw_finance_report', '时间', 'manual'),
        ('游戏报表场馆 · 游戏&场馆', 'raw_game_report_venue', '时间', 'manual'),
        ('游戏分析 · 游戏&场馆', 'raw_game_analysis', '日期', 'manual'),
        ('推广报表 · 代理&渠道', 'raw_promotion_report', '日期', 'manual'),
        ('代理报表 · 代理&渠道', 'raw_agent_report', '日期', 'manual'),
        ('即时注单 · 实时波动', 'raw_realtime_bet', '时间', 'manual'),
        ('红利记录 · 红利分析', 'raw_bonus_report', '申请时间', 'manual'),
        ('会员报表 · 会员/新注册', 'raw_member_report', '注册时间', 'manual'),
        ('TOP报表 · 会员价值', 'raw_top_report', '_snapshot_month', 'manual'),
        ('客服对话 · 客服分析', 'raw_cs_conversations', '_snapshot_month', 'manual'),
        ('电访 · 电访召回', 'raw_winback', '月份', 'manual'),
        ('代理结算汇总 · 代理佣金', 'raw_agent_settlement_summary', '月份', 'manual'),
        ('代理结算月度(市代) · 代理佣金', 'raw_agent_settlement_monthly', '月份', 'manual'),
    ]
    client = get_client()
    # 先查哪些表真的存在，避免某张表缺失（沙盒过期/从未上传）让整条 UNION 查询 404 拖垮整页
    existing = None
    try:
        edf = client.query(
            f"SELECT table_name FROM `{BQ_PREFIX}`.INFORMATION_SCHEMA.TABLES"
        ).result().to_dataframe()
        existing = set(edf['table_name'].tolist())
    except Exception:
        existing = None  # 查不到就退回旧行为（全部纳入查询）
    rows = []
    parts = []
    for name, tbl, col, kind in CFG:
        if existing is not None and tbl not in existing:
            rows.append((name, kind, None, 0))  # 表不存在 → 该行单独标无数据，不拖垮其他表
            continue
        parts.append(
            f"SELECT '{name}' AS name, '{kind}' AS kind, "
            f"CAST(MAX(SUBSTR(CAST(`{col}` AS STRING),1,10)) AS STRING) AS max_d, "
            f"COUNT(*) AS n FROM `{BQ_PREFIX}.{tbl}`"
        )
    if parts:
        sql = "\nUNION ALL\n".join(parts)
        try:
            df = client.query(sql).result().to_dataframe()
            for _, r in df.iterrows():
                rows.append((r['name'], r['kind'], r['max_d'], int(r['n'])))
        except Exception as e:
            rows.append((f'查询失败：{str(e)[:80]}', 'auto', None, 0))
    # 运营日报谷歌表（自动）
    try:
        daily = load_daily_ops(_recent_month_labels(3))
        if daily is not None and not daily.empty and '日期' in daily.columns:
            md = str(daily['日期'].max())[:10]
            rows.append(('运营日报谷歌表 · 近期走势', 'auto', md, int(len(daily))))
    except Exception:
        pass
    return rows


def render_data_health():
    import datetime as _dt
    hero('数据健康',
         '一览各报表数据更新到几号、是否滞后，便于及时补数与交接。绿＝最新，黄＝稍旧，红＝需尽快更新。',
         source_badge='数据健康检查')
    today = _dt.date.today()
    cur_ym = (today.year, today.month)

    def _ym(s):
        s = str(s)
        if len(s) >= 7 and s[4] == '-':
            return (int(s[:4]), int(s[5:7]))
        if len(s) >= 6 and s[:6].isdigit():
            return (int(s[:4]), int(s[4:6]))
        return None

    def _month_gap(s):
        ym = _ym(s)
        if not ym:
            return None
        return (cur_ym[0] - ym[0]) * 12 + (cur_ym[1] - ym[1])

    out = []
    red = yellow = 0
    for name, kind, max_d, n in _data_health_rows():
        if not max_d or max_d == 'None':
            status, note = '❌ 无数据', '表为空或查询失败'
            red += 1
        elif kind == 'auto':
            try:
                d = _dt.date.fromisoformat(max_d)
                gap = (today - d).days
            except Exception:
                gap = None
            if gap is None:
                status, note = '⚠️ 待查', '日期解析失败'
                yellow += 1
            elif gap <= 1:
                status, note = '✅ 最新', '自动更新中'
            elif gap <= 3:
                status, note = f'⚠️ 滞后 {gap} 天', '检查排程是否正常'
                yellow += 1
            else:
                status, note = f'❌ 滞后 {gap} 天', '排程可能失败，需检查'
                red += 1
        else:  # manual
            mg = _month_gap(max_d)
            if mg is None:
                status, note = '⚠️ 待查', '日期解析失败'
                yellow += 1
            elif mg <= 0:
                status, note = '✅ 本月已更新', '手动上传'
            elif mg == 1:
                status, note = '⚠️ 只到上月', '本月待上传'
                yellow += 1
            else:
                status, note = f'❌ 落后 {mg} 个月', '需补上传'
                red += 1
        cat = '🤖 自动' if kind == 'auto' else '📤 手动上传'
        out.append({'报表': name, '类别': cat, '数据最新': max_d or '—',
                    '状态': status, '说明': note, '_sev': (2 if status.startswith('❌') else 1 if status.startswith('⚠️') else 0)})

    if red:
        st.error(f'有 {red} 张表需尽快更新（红）。', icon='🚨')
    elif yellow:
        st.warning(f'有 {yellow} 张表稍旧（黄），其余正常。', icon='⚠️')
    else:
        st.success('所有数据均为最新 ✅', icon='✅')

    df = pd.DataFrame(out).sort_values('_sev', ascending=False).drop(columns='_sev')
    st.dataframe(df, use_container_width=True, hide_index=True)
    st.caption('判定：自动表期望更新到昨日（滞后>3天标红）；手动表期望本月已上传（只到上月标黄、更早标红）。'
               '需要补哪些、从后台哪里导，见「数据说明」页。缓存 10 分钟，刷新页面可重查。')


def render_data_source_guide():
    hero('数据说明',
         '本页汇整面板各页的数据来源、对应后台位置与更新方式，供日常维护与交接查阅。',
         source_badge='面板数据地图')
    st.markdown(
        '#### 数据更新方式（三类）\n'
        '- **自动**：「近期走势(日报)」读取运营日报谷歌表；「存取款分析」由程序每日 11:00 抓取。两者无需人工操作。\n'
        '- **手动上传**：其余多数页面。后台导出报表后，于「数据上传」页上传；同期重传将自动覆盖，不产生重复。\n'
        '- **人工提供**：电访（电访团队提供）、客服对话（客服系统导出），取得后经上传页入库。\n'
        '  （代理结算月报、市代月度虽由客服主管/平哥提供，但现已可直接拖上传页自助入库，归手动上传。）')

    st.markdown('#### 后台报表对应导出位置')
    backend = pd.DataFrame([
        ['经营报表', '报表中心 → 经营报表', '手动上传'],
        ['财务报表', '报表中心 → 财务报表', '手动上传'],
        ['游戏报表(场馆)', '报表中心 → 游戏报表(场馆)', '手动上传'],
        ['游戏分析', '报表中心 → 游戏分析（导出选「日报」颗粒度）', '手动上传'],
        ['推广报表', '报表中心 → 推广报表', '手动上传'],
        ['代理报表', '报表中心 → 代理报表（密码 zip）', '手动上传'],
        ['会员报表', '报表中心 → 会员报表（按注册时间、完整日期、全部页数导出）', '手动上传'],
        ['TOP报表', '报表中心 → TOP报表', '手动上传'],
        ['即时注单', '报表中心 → 即时注单', '手动上传'],
        ['红利记录', '会员管理 → VIP记录管理 → 红利记录', '手动上传'],
        ['代理佣金(单线/团队)', '代理管理 → 佣金管理 → 发放佣金（设佣金月份，导出 Csv）', '手动上传'],
        ['存款/提款 历史记录', '财务管理 → 存款管理/提款管理 → 历史记录', '自动 11:00'],
        ['运营日报(平台)', '由程序每日 10:00 写入谷歌表', '自动 10:00'],
        ['代理结算月报(平哥)', '客服主管提供「X月代理帐.xlsx」→ 拖数据上传页自助入库（选月份）', '手动上传'],
        ['代理结算月度(市代)', '市代「整理资料」表导出 xlsx（每代理每月产值/挂账）', '手动上传'],
        ['客服对话', '非后台报表，客服系统导出 xlsx', '人工提供'],
        ['撥打紀錄總表(电访)', '非后台报表，电访团队提供 xlsx', '人工提供'],
    ], columns=['报表', '后台位置 / 来源', '更新方式'])
    st.dataframe(backend, use_container_width=True, hide_index=True)

    st.markdown('#### 各页面对应报表与更新方式')
    pages = pd.DataFrame([
        ['🅰️ 财务结果', '经营总览', '经营报表＋财务报表＋红利记录＋代理结算', '手动上传'],
        ['🅰️ 财务结果', '近期走势(日报)', '运营日报谷歌表「平台报表」分页', '自动（每日10:00）'],
        ['🅰️ 财务结果', '存取款分析', '存款/提款 历史记录', '自动（每日11:00）'],
        ['🅰️ 财务结果', '红利分析', '红利记录', '手动上传'],
        ['🅰️ 财务结果', '红利 ROI & 代理质量', '红利记录＋代理报表', '手动上传'],
        ['🅰️ 财务结果', '代理佣金 & 退成', '代理佣金(单线/团队)＋代理结算月报(平哥)', '手动上传'],
        ['🅱️ 会员价值', '会员结构 & ARPU', '会员报表＋TOP报表', '手动上传'],
        ['🅱️ 会员价值', '投注分析', '注单明细(按月)；来源待确认，目前仅含 4 月', '手动上传'],
        ['🅱️ 会员价值', '客服分析', '客服对话＋会员报表', '人工提供'],
        ['🅱️ 会员价值', '电访召回', '撥打紀錄總表', '人工提供'],
        ['🅱️ 会员价值', '实时波动 & DAU', '即时注单', '手动上传'],
        ['🅲 代理/渠道', '代理团队 & 渠道', '代理报表＋推广报表', '手动上传'],
        ['🅲 代理/渠道', '新注册分析', '会员报表（口径为注册数的非代理部分）', '手动上传'],
        ['🅲 代理/渠道', '代理 × 会员 明细', '会员报表＋代理报表', '手动上传'],
        ['🅲 代理/渠道', '市代月度结算', '市代「整理资料」表（每代理每月）', '手动上传'],
        ['🅲 代理/渠道', '游戏 & 场馆', '游戏报表(场馆)＋游戏分析', '手动上传'],
    ], columns=['分组', '页面', '对应报表', '更新方式'])
    st.dataframe(pages, use_container_width=True, hide_index=True)

    st.info('各页面的补数月份与口径说明，见该页顶部「数据详情」。导出注意事项：日期字段须正确、日期范围完整、导出全部页数。'
            '「投注分析」所用注单明细来源待确认（目前仅含 4 月）。')


def _validate_entry(entry, client):
    """上传完整性校验（只提示不阻断）。返回 ⚠️ 提示字符串列表。
    四项：①会员报表非代理数 vs 平台注册数 ②行数 vs 库中同期骤降 ③日期区间缺天 ④关键数值全空。"""
    warns = []
    df = entry.get('df')
    table = entry.get('table')
    if df is None or len(df) == 0:
        return warns

    # ① 会员报表完整度：最新注册月「非代理」数 vs 平台报表同月同区间「注册数」
    if table == 'raw_member_report' and '注册时间' in df.columns and '是否为代理' in df.columns:
        try:
            rt = to_datetime_safe(df['注册时间'])
            valid = rt.notna()
            if valid.any():
                latest_m = rt[valid].dt.strftime('%Y-%m').max()
                in_m = valid & (rt.dt.strftime('%Y-%m') == latest_m)
                max_day = rt[in_m].max().strftime('%Y-%m-%d')
                file_nonagent = int((in_m & (df['是否为代理'].astype(str) == '非代理')).sum())
                # 比对基准：优先用日报谷歌表（当月最新、自动），否则退 BQ 平台报表（历史月）
                plat = 0
                try:
                    daily = load_daily_ops(_recent_month_labels(3))
                    if daily is not None and not daily.empty and {'日期', '注册数'}.issubset(daily.columns):
                        dd = daily.copy()
                        ds = dd['日期'].astype(str)
                        sub = dd[(ds.str[:7] == latest_m) & (ds.str[:10] <= max_day)]
                        if len(sub):
                            plat = int(pd.to_numeric(sub['注册数'], errors='coerce').fillna(0).sum())
                except Exception:
                    plat = 0
                if plat == 0:
                    try:
                        sql = (f"SELECT SUM(SAFE_CAST(REPLACE(CAST(`注册数` AS STRING),',','') AS FLOAT64)) AS n "
                               f"FROM `{BQ_PREFIX}.raw_platform_report` "
                               f"WHERE SUBSTR(CAST(`日期` AS STRING),1,7)=@m "
                               f"AND SUBSTR(CAST(`日期` AS STRING),1,10)<=@d")
                        cfg = bigquery.QueryJobConfig(query_parameters=[
                            bigquery.ScalarQueryParameter('m', 'STRING', latest_m),
                            bigquery.ScalarQueryParameter('d', 'STRING', max_day)])
                        rows = list(client.query(sql, job_config=cfg).result())
                        plat = int(rows[0].n) if rows and rows[0].n else 0
                    except Exception:
                        plat = 0
                if plat > 0 and file_nonagent < 0.8 * plat:
                    warns.append(f'⚠️ 完整度：{latest_m} 非代理 {file_nonagent} 笔，平台同期注册 {plat} 笔'
                                 f'（仅 {file_nonagent/plat*100:.0f}%）——这份可能不完整，导出疑似带筛选或没导全部页数')
        except Exception:
            pass

    # ② 行数骤降：本次 vs 库中同期已有
    exists = entry.get('exists')
    if isinstance(exists, int) and exists >= 20 and entry['rows'] < 0.7 * exists:
        warns.append(f'⚠️ 行数偏少：本次 {entry["rows"]} 行，库中同期已有 {exists} 行'
                     f'（{entry["rows"]/exists*100:.0f}%）——是否只导了部分？')

    # ③ 日期缺口：日明细类，区间内缺天
    keys = entry.get('date_keys')
    if keys and len(keys) >= 2:
        try:
            import datetime as _dt
            ds = [_dt.date.fromisoformat(k) for k in keys if len(k) == 10 and k[4] == '-']
            if len(ds) >= 2:
                full = {ds[0] + _dt.timedelta(days=i) for i in range((max(ds) - min(ds)).days + 1)}
                miss = sorted(full - set(ds))
                if miss:
                    shown = '、'.join(d.isoformat() for d in miss[:5]) + ('…' if len(miss) > 5 else '')
                    warns.append(f'⚠️ 区间内缺 {len(miss)} 天：{shown}（导出是否跳过了某些日期？）')
        except Exception:
            pass

    # ④ 关键数值全空：常见金额列若全为 0 / 空
    for col in ('有效投注额', '公司输赢', '订单金额', '存款额', '红利金额'):
        if col in df.columns:
            num = pd.to_numeric(df[col].astype(str).str.replace(',', '', regex=False), errors='coerce')
            if num.notna().sum() == 0 or (num.fillna(0) == 0).all():
                warns.append(f'⚠️ 「{col}」整列为 0 或空——数据疑似异常，请确认导出是否正确')
            break
    return warns


def render_data_upload():
    import os

    hero('数据上传（= 把报表永久保存进数据库）',
         '这页是把数据「存起来累积」用的：把后台下载的「任何一份」月度报表直接拖进来——平台 / 财务 / 游戏 / 游戏场馆 / '
         '游戏分析 / 推广 / 代理 / 会员 / TOP / 实时注单 / 红利 / 客服对话 / 代理佣金（单线·团队）/ 电访（撥打紀錄總表）全部通吃，'
         '系统自动认出是哪张、清洗、【保存进数据库（BigQuery）永久留存】。一次拖一堆不同的、密码 zip、分好几份都行；'
         '只「新增」绝不动库里旧数据。（拖进来 = 存进数据库；只想看分析不存，去对应的分析页。）')

    it = _import_tool()
    files = st.file_uploader(
        '上传月度报表（可一次多个；支持 .xlsx / .csv / .zip）',
        type=['xlsx', 'xls', 'csv', 'zip'], accept_multiple_files=True, key='dataup_files')
    zip_pw = st.text_input(
        '压缩档解压密码（只有上传 .zip 才需要填，例如 代理 / 会员 / 推广 的密码档）',
        type='password', key='dataup_zip_pw', help='密码只在本次使用，不会保存。')

    if not files:
        st.info(
            '👆 把后台下载的报表原档拖进来 = 自动保存进数据库（永久留存）。'
            '支持：平台 / 财务 / 游戏 / 游戏分析 / 推广 / 代理 / 会员 / TOP / 实时注单 / 红利 / 客服对话 / 代理佣金(单线·团队) / 电访(撥打紀錄總表)。\n\n'
            '· 文件名保持后台下载时的原样——系统靠文件名/内容认出是哪张报表，请别改名。\n\n'
            '· 密码压缩档（.zip）也能直接拖进来，在上面填一次解压密码即可，不用自己先解压。\n\n'
            '· 同一个档案重复上传会自动跳过；红利按「订单号」去重只补新订单；代理佣金按「佣金月份」刷新当月、保留其他月——都不会变成两份，也不动其他旧数据。')
        return

    client = get_client()
    preview_rows = []
    parsed = []  # 每个 entry 是 _classify_and_parse 返回的 dict
    upload_warnings = []  # [(文件, [⚠️...])]
    for f in files:
        for src, tmp_path, err in _expand_upload_units(f, zip_pw):
            if err:
                preview_rows.append({'文件': src, '识别类型': '❌ 出错',
                                     '行数': 0, '状态': err, '校验': '—'})
                continue
            try:
                entry = _classify_and_parse(it, client, src, tmp_path)
                try:
                    warns = _validate_entry(entry, client)
                except Exception:
                    warns = []
                if warns:
                    upload_warnings.append((src, warns))
                preview_rows.append({'文件': src, '识别类型': entry['display'],
                                     '行数': entry['rows'], '状态': entry['status'],
                                     '校验': ('；'.join(w.replace('⚠️ ', '') for w in warns) if warns else '✅ 正常')})
                if entry['is_new'] and entry['rows'] > 0:
                    parsed.append(entry)
            except Exception as e:
                preview_rows.append({'文件': src, '识别类型': '❌ 出错',
                                     '行数': 0, '状态': str(e)[:60], '校验': '—'})
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    section_header('识别结果', '先核对识别对不对，再按下面的按钮写入。')
    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    if upload_warnings:
        lines = []
        for src, warns in upload_warnings:
            fname = str(src).replace('\\', '/').split('/')[-1]
            for w in warns:
                lines.append(f'· **{fname}**：{w.replace("⚠️ ", "")}')
        st.warning('**完整性检查发现以下情况**（仅提示，不阻止写入；确认无误可照常写入）：\n\n' + '\n\n'.join(lines))

    if not parsed:
        st.warning('没有需要写入的新数据（都已上传过或未识别）。')
        return

    st.markdown(f'**准备写入 {len(parsed)} 个文件**（新增 / 刷新当月，不动其他旧数据）。')
    # 抓不到月份的快照报表（如 TOP 报表 top.xlsx 裸档名、无日期栏）→ 让用户选月份再写
    need_month = [e for e in parsed if e.get('need_month')]
    month_ok = True
    if need_month:
        import datetime as _dt
        _y, _mo = _dt.date.today().year, _dt.date.today().month
        _months = []
        for _ in range(13):
            _months.append(f'{_y:04d}-{_mo:02d}')
            _mo -= 1
            if _mo == 0:
                _mo = 12; _y -= 1
        opts = ['— 请选择 —'] + _months
        st.markdown('**下列报表判不出是哪个月，请选月份**（例如 TOP 报表档名只有 `top.xlsx`，没有日期）：')
        for e in need_month:
            fname = str(e['src']).replace('\\', '/').split('/')[-1]
            pick = st.selectbox(f'📄 {fname}（{e["display"]}）是哪个月？', opts, key=f'dataup_month_{fname}')
            if pick == '— 请选择 —':
                month_ok = False
            else:
                e['picked_month'] = pick   # 'YYYY-MM'
                if e.get('kind') == 'snapshot' and e.get('df') is not None:
                    ym = pick.replace('-', '')
                    e['df']['_snapshot_month'] = ym
                    snap = it.month_to_snapshot_date(ym)
                    if snap:
                        e['df']['_snapshot_date'] = snap
                    e['snapshot_month'] = ym
        if not month_ok:
            st.info('请先为上面每个报表选好月份，再写入。')
    # 守门：侦测到「整月汇总」误传(本该日明细) → 当场拦下、要二次确认，避免静默污染
    warn_entries = [e for e in parsed if e.get('warn_month_agg')]
    proceed = month_ok
    if warn_entries:
        st.warning('⚠️ 下列文件看起来是「整月汇总」而不是「日明细」（日期栏只有年月、没有具体到日）：\n\n'
                   + '\n'.join('· ' + str(e['src']).replace('\\', '/').split('/')[-1] for e in warn_entries)
                   + '\n\n这种文件导进去后，按日期范围筛选会整月对不上（跟其他月的日明细不一致）。'
                     '建议回后台改用「按日」颗粒度重新导出（档名会带完整起讫日，如 …【2026-05-01,2026-05-31】）。'
                     '真的要传汇总版才勾下面这格。')
        proceed = proceed and st.checkbox('我知道这是整月汇总、仍要写入', key='dataup_force_monthagg')
    if st.button('✅ 确认写入 BigQuery', type='primary', key='dataup_confirm', disabled=not proceed):
        import os as _os
        ok, fail = [], []
        others = [e for e in parsed if e['kind'] != 'commission']
        comm = [e for e in parsed if e['kind'] == 'commission']
        comm_tables = sorted(set(e['table'] for e in comm))
        total_steps = max(1, len(others) + len(comm_tables))
        prog = st.progress(0.0)
        done = 0
        # 非佣金：逐档 append
        for e in others:
            src, kind = e['src'], e['kind']
            try:
                if kind == 'standard':
                    if e.get('date_range'):
                        # 日报类：按日期替换（覆盖重叠日期，不重复）
                        rng, removed, written, total = _replace_by_date_range(client, e['df'], e['table'], src)
                        ok.append(f"{e['display']}：已替换 {rng}（覆盖旧 {removed} 行、写新 {written} 行，"
                                  f"表共 {total} 行，其他日期不动）")
                    elif _bq_count_source_file(client, e['table'], src) > 0:
                        fail.append(f'{src}：已存在，跳过')
                    else:
                        n = _append_standard_report(client, e['df'], e['table'], src)
                        ok.append(f"{e['display']}「{src}」：+{n} 行")
                elif kind == 'snapshot':
                    mth, updated, added, total = _replace_by_snapshot_month(client, e['df'], e['table'], src)
                    ok.append(f"{e['display']}：已合并 {mth}（按会员账号：更新 {updated} 个、新增 {added} 个，"
                              f"表共 {total} 行，其他月一行不动）")
                elif kind == 'bonus':
                    n = _append_bonus(client, e['df'], src, e.get('_existing_orders'))
                    if n == 0:
                        fail.append(f'{src}：订单全部已存在，跳过')
                    else:
                        ok.append(f"红利「{src}」：+{n} 笔新订单")
                elif kind == 'cs':
                    if _cs_basename_loaded(client, _os.path.basename(src)) > 0:
                        fail.append(f'{src}：已存在，跳过')
                    else:
                        n = _append_cs(client, e['df'], src)
                        ok.append(f"客服对话「{src}」：+{n} 条")
                elif kind == 'winback':
                    total = _write_winback(client, e['df'], e.get('months', []), src)
                    ok.append(f"电访召回：已存 {','.join(e.get('months', []))}（表共 {total} 行，其他月不动）")
                elif kind == 'agent_monthly':
                    months = e.get('months', [])
                    n = _write_agent_monthly(client, e['df'], months, src)
                    rng = f"{months[0]}~{months[-1]}" if len(months) > 1 else (months[0] if months else '?')
                    ok.append(f"代理结算月度(市代)：已按月替换 {rng}（{len(months)} 个月，写 {n} 行，其他月不动）")
                elif kind == 'settlement':
                    mth = e.get('picked_month', '?')
                    ns, nd = _write_agent_settlement(client, e.get('summary_records', []),
                                                     e.get('detail_records', []), mth, src)
                    ok.append(f"代理结算月报(平哥) {mth}：摘要 {ns} 行 + 明细 {nd} 行（按月替换，其他月不动）")
            except Exception as ex:
                fail.append(f'{src}：{str(ex)[:80]}')
            done += 1
            prog.progress(done / total_steps)
        # 佣金：按表分组，整表读改写（保留其他月份）
        for table in comm_tables:
            entries = [e for e in comm if e['table'] == table]
            disp = entries[0]['display']
            try:
                new_df = pd.concat([e['df'] for e in entries], ignore_index=True)
                srcs = ' / '.join(_os.path.basename(e['src']) for e in entries)
                months = sorted(set(m for e in entries for m in e.get('months', [])))
                total = _write_commission_safe(client, new_df, table, srcs)
                ok.append(f"{disp}：已更新 {','.join(months)}（表共 {total} 行，其他月保留）")
            except Exception as ex:
                fail.append(f'{disp}：{str(ex)[:90]}')
            done += 1
            prog.progress(done / total_steps)
        if ok:
            st.success('写入成功：\n\n' + '\n\n'.join('· ' + s for s in ok))
        if fail:
            st.error('未写入：\n\n' + '\n\n'.join('· ' + s for s in fail))
        try:
            st.cache_data.clear()
        except Exception:
            pass
        st.info('完成。切到对应分析页（可能要刷新一下）就能看到新数据了。')


_DM_TABLES = {
    '会员报表': 'raw_member_report', 'TOP报表': 'raw_top_report',
    '平台报表': 'raw_platform_report', '财务报表': 'raw_finance_report',
    '游戏报表': 'raw_game_report', '游戏场馆': 'raw_game_report_venue',
    '游戏分析': 'raw_game_analysis', '推广报表': 'raw_promotion_report',
    '代理报表': 'raw_agent_report', '实时注单': 'raw_realtime_bet',
    '红利记录': 'raw_bonus_report', '客服对话': 'raw_cs_conversations',
    '代理佣金(单线)': 'raw_agent_commission_single', '代理佣金(团队)': 'raw_agent_commission_team',
    '电访召回': 'raw_winback',
    '存款订单': 'raw_finance_deposit', '提款订单': 'raw_finance_withdraw',
    '代理结算月度(市代)': 'raw_agent_settlement_monthly',
}


def render_data_manage():
    hero('删除数据（真删 BigQuery · 控制库大小）',
         '选一张表 + 要删的月份/批次 → 真的从 BigQuery 把那批删掉（表会变小、省空间），不是只是不显示。'
         '用「读出来 → 去掉 → 整张写回」的安全删法（这项目没开 billing、免费版禁直接 DELETE，但这招照样真删）。删了就没了，谨慎操作。')
    try:
        client = get_client()
    except Exception as e:
        st.error(f'连不上数据库：{str(e)[:100]}')
        return
    name = st.selectbox('选一张表', list(_DM_TABLES.keys()), key='dm_table')
    table = _DM_TABLES[name]
    key, periods = _bq_periods(client, table)
    if not periods:
        st.info('这张表目前没有数据（或读取失败）。')
        return
    label = '月份' if key in ('_snapshot_month', '佣金月份') else ('上传批次（档名）' if key == '_source_file' else key)
    section_header(f'{name} 现有数据（按{label}）', f'下面是这张表里现有的{label}和各自行数，勾选要删的。')
    dfp = pd.DataFrame([{label: k, '行数': v} for k, v in periods.items()])
    st.dataframe(dfp, use_container_width=True, hide_index=True)
    sel = st.multiselect(f'选要删的{label}（可多选）', list(periods.keys()), key='dm_sel')
    if sel:
        total_del = sum(periods[s] for s in sel)
        st.warning(f'⚠️ 将从 BigQuery 真的删掉：「{name}」的 {len(sel)} 个{label}，共 {total_del} 行。'
                   f'删了无法复原（要回来只能重新上传）。')
        confirm = st.text_input('确认请输入「删除」二字', key='dm_confirm')
        if st.button('🗑 确认从 BigQuery 删除', type='primary', key='dm_btn',
                     disabled=(confirm.strip() != '删除')):
            with st.spinner('删除中（读出 → 去掉 → 写回）…'):
                try:
                    removed, remaining = _bq_delete_periods(client, table, key, sel)
                    st.success(f'✅ 已从 BigQuery 删除 {removed} 行；「{name}」现在剩 {remaining} 行（表已变小）。')
                    try:
                        st.cache_data.clear()
                    except Exception:
                        pass
                except Exception as ex:
                    st.error(f'删除失败：{str(ex)[:140]}')


def main():
    # 两层导航：大类 → 细项
    GROUPS = {
        '🅰️ 财务结果': [
            ('经营总览', render_overview),
            ('近期走势(日报)', render_recent_trend),
            ('存取款分析', render_finance_channel),
            ('红利分析', render_bonus_analysis),
            ('红利 ROI & 代理质量', render_bonus_roi_agent_quality),
            ('代理佣金 & 退成', render_agent_commission),
        ],
        '🅱️ 会员价值': [
            ('会员结构 & ARPU', render_member_value),
            ('投注分析', render_bet_analysis),
            ('客服分析', render_cs_analysis),
            ('电访召回', render_winback),
            ('实时波动 & DAU', render_realtime),
        ],
        '🅲 代理 / 渠道': [
            ('代理团队 & 渠道', render_channel_agent),
            ('新注册分析', render_new_member_analysis),
            ('代理 × 会员 明细', render_agent_member_matrix),
            ('市代月度结算', render_agent_market_monthly),
            ('游戏 & 场馆', render_game_venue),
        ],
        '🗂 数据上传': [
            ('🩺 数据健康', render_data_health),
            ('📖 数据说明', render_data_source_guide),
            ('月度报表上传', render_data_upload),
            ('删除数据', render_data_manage),
        ],
    }
    group = st.radio(
        '大类', list(GROUPS.keys()),
        horizontal=True, label_visibility='collapsed', key='nav_group',
    )
    sub_options = [name for name, _ in GROUPS[group]]
    sub_renderers = {name: fn for name, fn in GROUPS[group]}
    st.markdown(
        '<div style="height:0.4rem;"></div>',
        unsafe_allow_html=True,
    )
    sub = st.radio(
        '细项', sub_options,
        horizontal=True, label_visibility='collapsed', key=f'nav_sub_{group}',
    )
    sub_renderers[sub]()
    st.divider()
    st.caption(f'运营数据面板 {APP_VERSION}（{APP_VERSION_DATE}）· 更新内容见仓库 CHANGELOG.md')


if __name__ == '__main__':
    main()
